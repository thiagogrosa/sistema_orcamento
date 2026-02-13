#!/usr/bin/env python3
"""
Gerador de Planilha Interna HVAC

Gera Excel com custos detalhados para controle interno.
Inclui estrutura hierarquica e listas consolidadas para compras.
Possui colunas para controle orcado x realizado.
"""

from datetime import date
from pathlib import Path
from typing import Dict, Any, Optional, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .utils import (
    carregar_configs,
    formatar_moeda,
    criar_pasta_cliente,
    gerar_nome_arquivo,
    BASE_DIR
)


# Estilos
AZUL_ARMANT = "00A0E3"
CINZA_CLARO = "F0F0F0"
CINZA_MEDIO = "E0E0E0"
VERDE_POSITIVO = "C6EFCE"
VERMELHO_NEGATIVO = "FFC7CE"


def criar_estilos(wb: Workbook) -> Dict[str, NamedStyle]:
    """Cria estilos padronizados para a planilha"""

    borda_fina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    estilos = {}

    # Cabecalho
    estilo_cabecalho = NamedStyle(name="cabecalho")
    estilo_cabecalho.font = Font(bold=True, color="FFFFFF", size=10)
    estilo_cabecalho.fill = PatternFill(start_color=AZUL_ARMANT, end_color=AZUL_ARMANT, fill_type="solid")
    estilo_cabecalho.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    estilo_cabecalho.border = borda_fina
    wb.add_named_style(estilo_cabecalho)
    estilos["cabecalho"] = estilo_cabecalho

    # Grupo
    estilo_grupo = NamedStyle(name="grupo")
    estilo_grupo.font = Font(bold=True, size=10)
    estilo_grupo.fill = PatternFill(start_color=CINZA_MEDIO, end_color=CINZA_MEDIO, fill_type="solid")
    estilo_grupo.border = borda_fina
    wb.add_named_style(estilo_grupo)
    estilos["grupo"] = estilo_grupo

    # Item normal
    estilo_item = NamedStyle(name="item")
    estilo_item.font = Font(size=9)
    estilo_item.border = borda_fina
    estilo_item.alignment = Alignment(vertical="center")
    wb.add_named_style(estilo_item)
    estilos["item"] = estilo_item

    # Subitem (materiais, MO, ferramentas)
    estilo_subitem = NamedStyle(name="subitem")
    estilo_subitem.font = Font(size=8, color="666666")
    estilo_subitem.fill = PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type="solid")
    estilo_subitem.border = borda_fina
    wb.add_named_style(estilo_subitem)
    estilos["subitem"] = estilo_subitem

    # Total
    estilo_total = NamedStyle(name="total")
    estilo_total.font = Font(bold=True, size=11)
    estilo_total.fill = PatternFill(start_color=CINZA_MEDIO, end_color=CINZA_MEDIO, fill_type="solid")
    estilo_total.border = borda_fina
    wb.add_named_style(estilo_total)
    estilos["total"] = estilo_total

    # Moeda
    estilo_moeda = NamedStyle(name="moeda")
    estilo_moeda.font = Font(size=9)
    estilo_moeda.border = borda_fina
    estilo_moeda.number_format = '#,##0.00'
    estilo_moeda.alignment = Alignment(horizontal="right")
    wb.add_named_style(estilo_moeda)
    estilos["moeda"] = estilo_moeda

    return estilos


def criar_cabecalho_planilha(ws: Worksheet, numero_orcamento: str, cliente: str, data: str):
    """Adiciona cabecalho padrao na planilha"""
    ws["A1"] = f"Referencia: ORC {numero_orcamento}"
    ws["A1"].font = Font(bold=True, size=12, color=AZUL_ARMANT)

    ws["A2"] = f"Cliente: {cliente}"
    ws["A2"].font = Font(size=10)

    ws["A3"] = f"Data: {data}"
    ws["A3"].font = Font(size=10)

    # Linha em branco
    return 5  # Proxima linha disponivel


def criar_aba_resumo(
    wb: Workbook,
    precificado: Dict[str, Any],
    numero_orcamento: str,
    cliente: str
) -> Worksheet:
    """
    Cria aba com resumo hierarquico por item

    Estrutura:
    - Grupos de servico
      - Itens de servico
        - Materiais
        - Mao de obra
        - Ferramentas
    """
    ws = wb.active
    ws.title = "Resumo por Item"

    # Cabecalho
    linha = criar_cabecalho_planilha(ws, numero_orcamento, cliente, date.today().strftime("%d/%m/%Y"))

    # Cabecalho da tabela
    colunas = [
        ("A", "Item", 8),
        ("B", "Descricao", 45),
        ("C", "Tipo", 12),
        ("D", "Unidade", 8),
        ("E", "Qtd", 8),
        ("F", "Custo Unit. Orcado", 15),
        ("G", "Custo Total Orcado", 15),
        ("H", "Custo Unit. Real", 15),
        ("I", "Custo Total Real", 15),
        ("J", "Diferenca (R$)", 14),
        ("K", "Diferenca (%)", 12)
    ]

    for col, titulo, largura in colunas:
        celula = ws[f"{col}{linha}"]
        celula.value = titulo
        celula.style = "cabecalho"
        ws.column_dimensions[col].width = largura

    linha += 1
    linha_inicio_dados = linha

    # Processa itens
    itens = precificado.get("itens_precificados", [])
    agrupamento = precificado.get("agrupamento", [])

    # Se tem agrupamento
    if agrupamento:
        for idx_grupo, grupo in enumerate(agrupamento, 1):
            # Linha do grupo
            ws[f"A{linha}"] = str(idx_grupo)
            ws[f"B{linha}"] = grupo.get("nome", f"GRUPO {idx_grupo}")
            ws[f"C{linha}"] = "GRUPO"

            for col in "ABCDEFGHIJK":
                ws[f"{col}{linha}"].style = "grupo"

            linha += 1

            # Itens do grupo
            for item_id in grupo.get("itens_ids", []):
                for item in itens:
                    if item.get("id") == item_id:
                        linha = adicionar_item_hierarquico(ws, item, idx_grupo, linha)
                        break
    else:
        # Sem agrupamento - lista direta
        for idx, item in enumerate(itens, 1):
            linha = adicionar_item_hierarquico(ws, item, idx, linha)

    # Linha de total
    linha += 1
    ws[f"A{linha}"] = ""
    ws[f"B{linha}"] = "TOTAL GERAL"
    ws[f"F{linha}"] = ""
    ws[f"G{linha}"] = f"=SUM(G{linha_inicio_dados}:G{linha-2})"
    ws[f"H{linha}"] = ""
    ws[f"I{linha}"] = f"=SUM(I{linha_inicio_dados}:I{linha-2})"
    ws[f"J{linha}"] = f"=I{linha}-G{linha}"
    ws[f"K{linha}"] = f"=IF(G{linha}=0,0,(J{linha}/G{linha})*100)"

    for col in "ABCDEFGHIJK":
        ws[f"{col}{linha}"].style = "total"

    # Formata colunas de moeda
    for row in ws.iter_rows(min_row=linha_inicio_dados, max_row=linha, min_col=6, max_col=10):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Formatacao condicional para diferenca
    for row in range(linha_inicio_dados, linha + 1):
        celula_dif = ws[f"J{row}"]
        celula_dif.number_format = '#,##0.00'

    return ws


def adicionar_item_hierarquico(ws: Worksheet, item: Dict, idx_grupo: int, linha: int) -> int:
    """Adiciona item com seus subitens (materiais, MO, ferramentas)"""

    item_num = f"{idx_grupo}.{item.get('id', 1)}"

    # Linha do servico
    ws[f"A{linha}"] = item_num
    ws[f"B{linha}"] = item.get("descricao", "")
    ws[f"C{linha}"] = "SERVICO"
    ws[f"D{linha}"] = item.get("unidade", "pc")
    ws[f"E{linha}"] = item.get("quantidade", 1)
    ws[f"F{linha}"] = ""  # Unitario calculado
    ws[f"G{linha}"] = item.get("custo_direto", 0)
    ws[f"H{linha}"] = ""  # Real - editavel
    ws[f"I{linha}"] = ""  # Total real - formula
    ws[f"J{linha}"] = f"=IF(I{linha}=\"\",\"\",I{linha}-G{linha})"
    ws[f"K{linha}"] = f"=IF(OR(G{linha}=0,J{linha}=\"\"),\"\",(J{linha}/G{linha})*100)"

    for col in "ABCDEFGHIJK":
        ws[f"{col}{linha}"].style = "item"

    linha += 1

    # Materiais
    for idx_mat, mat in enumerate(item.get("materiais", []), 1):
        ws[f"A{linha}"] = f"{item_num}.{idx_mat}"
        ws[f"B{linha}"] = mat.get("descricao", mat.get("codigo", ""))
        ws[f"C{linha}"] = "MATERIAL"
        ws[f"D{linha}"] = mat.get("unidade", "UN")
        ws[f"E{linha}"] = mat.get("quantidade", 0)
        ws[f"F{linha}"] = mat.get("preco_unitario", 0)
        ws[f"G{linha}"] = mat.get("custo", 0)
        ws[f"H{linha}"] = ""
        ws[f"I{linha}"] = f"=IF(H{linha}=\"\",\"\",H{linha}*E{linha})"
        ws[f"J{linha}"] = f"=IF(I{linha}=\"\",\"\",I{linha}-G{linha})"
        ws[f"K{linha}"] = f"=IF(OR(G{linha}=0,J{linha}=\"\"),\"\",(J{linha}/G{linha})*100)"

        for col in "ABCDEFGHIJK":
            ws[f"{col}{linha}"].style = "subitem"

        linha += 1

    # Mao de obra
    for idx_mo, mo in enumerate(item.get("mao_de_obra", []), 1):
        ws[f"A{linha}"] = f"{item_num}.MO{idx_mo}"
        ws[f"B{linha}"] = mo.get("descricao", mo.get("codigo", ""))
        ws[f"C{linha}"] = "MAO_OBRA"
        ws[f"D{linha}"] = "h"
        ws[f"E{linha}"] = mo.get("quantidade", 0)
        ws[f"F{linha}"] = mo.get("preco_unitario", 0)
        ws[f"G{linha}"] = mo.get("custo", 0)
        ws[f"H{linha}"] = ""
        ws[f"I{linha}"] = f"=IF(H{linha}=\"\",\"\",H{linha}*E{linha})"
        ws[f"J{linha}"] = f"=IF(I{linha}=\"\",\"\",I{linha}-G{linha})"
        ws[f"K{linha}"] = f"=IF(OR(G{linha}=0,J{linha}=\"\"),\"\",(J{linha}/G{linha})*100)"

        for col in "ABCDEFGHIJK":
            ws[f"{col}{linha}"].style = "subitem"

        linha += 1

    # Ferramentas
    for idx_fer, fer in enumerate(item.get("ferramentas", []), 1):
        ws[f"A{linha}"] = f"{item_num}.FE{idx_fer}"
        ws[f"B{linha}"] = fer.get("descricao", fer.get("codigo", ""))
        ws[f"C{linha}"] = "FERRAMENTA"
        ws[f"D{linha}"] = "h"
        ws[f"E{linha}"] = fer.get("quantidade", 0)
        ws[f"F{linha}"] = fer.get("preco_unitario", 0)
        ws[f"G{linha}"] = fer.get("custo", 0)
        ws[f"H{linha}"] = ""
        ws[f"I{linha}"] = f"=IF(H{linha}=\"\",\"\",H{linha}*E{linha})"
        ws[f"J{linha}"] = f"=IF(I{linha}=\"\",\"\",I{linha}-G{linha})"
        ws[f"K{linha}"] = f"=IF(OR(G{linha}=0,J{linha}=\"\"),\"\",(J{linha}/G{linha})*100)"

        for col in "ABCDEFGHIJK":
            ws[f"{col}{linha}"].style = "subitem"

        linha += 1

    return linha


def criar_aba_materiais(
    wb: Workbook,
    precificado: Dict[str, Any],
    numero_orcamento: str,
    cliente: str
) -> Worksheet:
    """Cria aba com lista consolidada de materiais"""

    ws = wb.create_sheet("Materiais")

    # Cabecalho
    linha = criar_cabecalho_planilha(ws, numero_orcamento, cliente, date.today().strftime("%d/%m/%Y"))

    # Cabecalho da tabela
    colunas = [
        ("A", "Codigo", 15),
        ("B", "Descricao", 40),
        ("C", "Unidade", 10),
        ("D", "Qtd Total", 12),
        ("E", "Custo Unit. Orcado", 16),
        ("F", "Custo Total Orcado", 16),
        ("G", "Fornecedor Ref.", 25),
        ("H", "Custo Unit. Real", 16),
        ("I", "Custo Total Real", 16),
        ("J", "Diferenca", 14)
    ]

    for col, titulo, largura in colunas:
        celula = ws[f"{col}{linha}"]
        celula.value = titulo
        celula.style = "cabecalho"
        ws.column_dimensions[col].width = largura

    linha += 1
    linha_inicio = linha

    # Consolida materiais
    materiais_consolidados = {}

    for item in precificado.get("itens_precificados", []):
        for mat in item.get("materiais", []):
            codigo = mat.get("codigo", "")
            if codigo in materiais_consolidados:
                materiais_consolidados[codigo]["quantidade"] += mat.get("quantidade", 0)
                materiais_consolidados[codigo]["custo"] += mat.get("custo", 0)
            else:
                materiais_consolidados[codigo] = {
                    "codigo": codigo,
                    "descricao": mat.get("descricao", ""),
                    "unidade": mat.get("unidade", "UN"),
                    "quantidade": mat.get("quantidade", 0),
                    "preco_unitario": mat.get("preco_unitario", 0),
                    "custo": mat.get("custo", 0),
                    "fornecedor": mat.get("fornecedor_referencia", "")
                }

    # Adiciona linhas
    for mat in sorted(materiais_consolidados.values(), key=lambda x: x["codigo"]):
        ws[f"A{linha}"] = mat["codigo"]
        ws[f"B{linha}"] = mat["descricao"]
        ws[f"C{linha}"] = mat["unidade"]
        ws[f"D{linha}"] = mat["quantidade"]
        ws[f"E{linha}"] = mat["preco_unitario"]
        ws[f"F{linha}"] = mat["custo"]
        ws[f"G{linha}"] = mat["fornecedor"]
        ws[f"H{linha}"] = ""
        ws[f"I{linha}"] = f"=IF(H{linha}=\"\",\"\",H{linha}*D{linha})"
        ws[f"J{linha}"] = f"=IF(I{linha}=\"\",\"\",I{linha}-F{linha})"

        for col in "ABCDEFGHIJ":
            ws[f"{col}{linha}"].style = "item"

        linha += 1

    # Total
    linha += 1
    ws[f"A{linha}"] = ""
    ws[f"B{linha}"] = "TOTAL"
    ws[f"F{linha}"] = f"=SUM(F{linha_inicio}:F{linha-2})"
    ws[f"I{linha}"] = f"=SUM(I{linha_inicio}:I{linha-2})"
    ws[f"J{linha}"] = f"=I{linha}-F{linha}"

    for col in "ABCDEFGHIJ":
        ws[f"{col}{linha}"].style = "total"

    return ws


def criar_aba_mao_obra(
    wb: Workbook,
    precificado: Dict[str, Any],
    numero_orcamento: str,
    cliente: str
) -> Worksheet:
    """Cria aba com lista consolidada de mao de obra"""

    ws = wb.create_sheet("Mao de Obra")

    # Cabecalho
    linha = criar_cabecalho_planilha(ws, numero_orcamento, cliente, date.today().strftime("%d/%m/%Y"))

    # Cabecalho da tabela
    colunas = [
        ("A", "Categoria", 20),
        ("B", "Descricao", 35),
        ("C", "Horas Totais", 14),
        ("D", "Custo/Hora Orcado", 16),
        ("E", "Custo Total Orcado", 16),
        ("F", "Custo/Hora Real", 16),
        ("G", "Custo Total Real", 16),
        ("H", "Diferenca", 14)
    ]

    for col, titulo, largura in colunas:
        celula = ws[f"{col}{linha}"]
        celula.value = titulo
        celula.style = "cabecalho"
        ws.column_dimensions[col].width = largura

    linha += 1
    linha_inicio = linha

    # Consolida mao de obra
    mo_consolidada = {}

    for item in precificado.get("itens_precificados", []):
        for mo in item.get("mao_de_obra", []):
            codigo = mo.get("codigo", "")
            if codigo in mo_consolidada:
                mo_consolidada[codigo]["quantidade"] += mo.get("quantidade", 0)
                mo_consolidada[codigo]["custo"] += mo.get("custo", 0)
            else:
                mo_consolidada[codigo] = {
                    "codigo": codigo,
                    "descricao": mo.get("descricao", ""),
                    "quantidade": mo.get("quantidade", 0),
                    "preco_unitario": mo.get("preco_unitario", 0),
                    "custo": mo.get("custo", 0)
                }

    # Adiciona linhas
    for mo in sorted(mo_consolidada.values(), key=lambda x: x["codigo"]):
        ws[f"A{linha}"] = mo["codigo"]
        ws[f"B{linha}"] = mo["descricao"]
        ws[f"C{linha}"] = mo["quantidade"]
        ws[f"D{linha}"] = mo["preco_unitario"]
        ws[f"E{linha}"] = mo["custo"]
        ws[f"F{linha}"] = ""
        ws[f"G{linha}"] = f"=IF(F{linha}=\"\",\"\",F{linha}*C{linha})"
        ws[f"H{linha}"] = f"=IF(G{linha}=\"\",\"\",G{linha}-E{linha})"

        for col in "ABCDEFGH":
            ws[f"{col}{linha}"].style = "item"

        linha += 1

    # Total
    linha += 1
    ws[f"A{linha}"] = ""
    ws[f"B{linha}"] = "TOTAL"
    ws[f"C{linha}"] = f"=SUM(C{linha_inicio}:C{linha-2})"
    ws[f"E{linha}"] = f"=SUM(E{linha_inicio}:E{linha-2})"
    ws[f"G{linha}"] = f"=SUM(G{linha_inicio}:G{linha-2})"
    ws[f"H{linha}"] = f"=G{linha}-E{linha}"

    for col in "ABCDEFGH":
        ws[f"{col}{linha}"].style = "total"

    return ws


def criar_aba_ferramentas(
    wb: Workbook,
    precificado: Dict[str, Any],
    numero_orcamento: str,
    cliente: str
) -> Worksheet:
    """Cria aba com lista consolidada de ferramentas"""

    ws = wb.create_sheet("Ferramentas")

    # Cabecalho
    linha = criar_cabecalho_planilha(ws, numero_orcamento, cliente, date.today().strftime("%d/%m/%Y"))

    # Cabecalho da tabela
    colunas = [
        ("A", "Codigo", 15),
        ("B", "Descricao", 35),
        ("C", "Horas Uso", 12),
        ("D", "Custo/Hora Orcado", 16),
        ("E", "Custo Total Orcado", 16),
        ("F", "Custo/Hora Real", 16),
        ("G", "Custo Total Real", 16),
        ("H", "Diferenca", 14)
    ]

    for col, titulo, largura in colunas:
        celula = ws[f"{col}{linha}"]
        celula.value = titulo
        celula.style = "cabecalho"
        ws.column_dimensions[col].width = largura

    linha += 1
    linha_inicio = linha

    # Consolida ferramentas
    fer_consolidadas = {}

    for item in precificado.get("itens_precificados", []):
        for fer in item.get("ferramentas", []):
            codigo = fer.get("codigo", "")
            if codigo in fer_consolidadas:
                fer_consolidadas[codigo]["quantidade"] += fer.get("quantidade", 0)
                fer_consolidadas[codigo]["custo"] += fer.get("custo", 0)
            else:
                fer_consolidadas[codigo] = {
                    "codigo": codigo,
                    "descricao": fer.get("descricao", ""),
                    "quantidade": fer.get("quantidade", 0),
                    "preco_unitario": fer.get("preco_unitario", 0),
                    "custo": fer.get("custo", 0)
                }

    # Adiciona linhas
    for fer in sorted(fer_consolidadas.values(), key=lambda x: x["codigo"]):
        ws[f"A{linha}"] = fer["codigo"]
        ws[f"B{linha}"] = fer["descricao"]
        ws[f"C{linha}"] = fer["quantidade"]
        ws[f"D{linha}"] = fer["preco_unitario"]
        ws[f"E{linha}"] = fer["custo"]
        ws[f"F{linha}"] = ""
        ws[f"G{linha}"] = f"=IF(F{linha}=\"\",\"\",F{linha}*C{linha})"
        ws[f"H{linha}"] = f"=IF(G{linha}=\"\",\"\",G{linha}-E{linha})"

        for col in "ABCDEFGH":
            ws[f"{col}{linha}"].style = "item"

        linha += 1

    # Total
    if linha > linha_inicio:
        linha += 1
        ws[f"A{linha}"] = ""
        ws[f"B{linha}"] = "TOTAL"
        ws[f"C{linha}"] = f"=SUM(C{linha_inicio}:C{linha-2})"
        ws[f"E{linha}"] = f"=SUM(E{linha_inicio}:E{linha-2})"
        ws[f"G{linha}"] = f"=SUM(G{linha_inicio}:G{linha-2})"
        ws[f"H{linha}"] = f"=G{linha}-E{linha}"

        for col in "ABCDEFGH":
            ws[f"{col}{linha}"].style = "total"

    return ws


def gerar_planilha_interna(
    precificado: Dict[str, Any],
    numero_orcamento: str,
    output_path: Optional[str] = None,
    configs: Optional[Dict] = None
) -> Dict[str, Any]:
    """
    Gera planilha Excel interna com custos detalhados

    Args:
        precificado: Dados do orcamento precificado
        numero_orcamento: Numero do orcamento
        output_path: Caminho de saida (gera automatico se nao informado)
        configs: Configuracoes

    Returns:
        Dicionario com resultado:
        {
            "sucesso": bool,
            "arquivo_xlsx": str,
            "erro": str (se falhou)
        }
    """
    try:
        # Carrega configs
        if configs is None:
            configs = carregar_configs()

        # Dados do cliente
        dados_cliente = precificado.get("dados_cliente", {})
        cliente_nome = dados_cliente.get("razao_social") or precificado.get("cliente", "Cliente")

        # Cria workbook
        wb = Workbook()

        # Cria estilos
        criar_estilos(wb)

        # Cria abas
        criar_aba_resumo(wb, precificado, numero_orcamento, cliente_nome)
        criar_aba_materiais(wb, precificado, numero_orcamento, cliente_nome)
        criar_aba_mao_obra(wb, precificado, numero_orcamento, cliente_nome)
        criar_aba_ferramentas(wb, precificado, numero_orcamento, cliente_nome)

        # Define caminho de saida
        if output_path is None:
            cliente_dir = criar_pasta_cliente(cliente_nome)

            # Extrai revisao do numero
            revisao = "R00"
            if "-R" in numero_orcamento:
                revisao = numero_orcamento.split("-")[-1]

            # Detecta tipo de servico
            tipo_servico = "instalacao"
            for item in precificado.get("itens_precificados", []):
                if item.get("tipo_servico"):
                    tipo_servico = item["tipo_servico"]
                    break

            nome_arquivo = gerar_nome_arquivo(
                numero_orcamento,
                cliente_nome,
                tipo_servico,
                revisao,
                "_interno"
            )
            output_path = cliente_dir / f"{nome_arquivo}.xlsx"
        else:
            output_path = Path(output_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Salva
        wb.save(str(output_path))

        return {
            "sucesso": True,
            "arquivo_xlsx": str(output_path)
        }

    except Exception as e:
        return {
            "sucesso": False,
            "erro": str(e)
        }


def main():
    """CLI para teste"""
    import argparse
    import json

    parser = argparse.ArgumentParser(description="Gera planilha interna HVAC")
    parser.add_argument("--input", "-i", required=True, help="Arquivo precificado.json")
    parser.add_argument("--output", "-o", help="Arquivo Excel de saida")
    parser.add_argument("--numero", "-n", required=True, help="Numero do orcamento")

    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        precificado = json.load(f)

    resultado = gerar_planilha_interna(
        precificado,
        args.numero,
        args.output
    )

    if resultado["sucesso"]:
        print(f"Planilha gerada: {resultado['arquivo_xlsx']}")
    else:
        print(f"Erro: {resultado.get('erro')}")


if __name__ == "__main__":
    main()
