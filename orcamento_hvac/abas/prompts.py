"""
Aba PROMPTS - Prompts para criação de novos itens com LLM.
"""
from openpyxl.styles import Font, PatternFill, Alignment


def criar(wb, estilos):
    """Cria a aba PROMPTS com os prompts para LLM."""
    ws = wb.create_sheet('PROMPTS')

    conteudo = [
        ('TÍTULO', 'PROMPTS PARA CRIAÇÃO DE NOVOS ITENS COM LLM'),
        ('', ''),
        ('TEXTO', 'Copie o prompt correspondente, substitua o texto entre colchetes e use com ChatGPT, Claude ou outro LLM.'),
        ('', ''),
        ('SEÇÃO', 'PROMPT 1 - CRIAR MATERIAL'),
        ('', ''),
        ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar um novo material na minha planilha de custos.

CONTEXTO DA NOMENCLATURA:
- Tubos: TUB_[POLEGADA]_[FLEX/RIG]
- Isolamentos: ISO_[POLEGADA]_[ELA/POL]_E[ESPESSURA_MM]
- Cabos: CAB_[TIPO]_[SEÇÃO]
- Dreno: DRN_[TIPO]_[ESPECIFICAÇÃO]
- Suportes: SUP_[TIPO]_[ESPECIFICAÇÃO]
- Acabamento: ACA_[TIPO]_[ESPECIFICAÇÃO]
- Gás: GAS_[TIPO] / Solda: SOL_[TIPO]
- Disjuntores: DISJ_[M/B]_[AMPERAGEM]
- Alvenaria: ALV_[TIPO]

MATERIAL A CADASTRAR: [DESCREVA O MATERIAL AQUI]

Responda em formato de tabela com:
1. Código sugerido (seguindo a nomenclatura)
2. Categoria
3. Descrição completa
4. Unidade de medida
5. Preço estimado de mercado (Porto Alegre/RS)'''),
        ('', ''),
        ('SEÇÃO', 'PROMPT 2 - CRIAR MÃO DE OBRA'),
        ('', ''),
        ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar uma nova função de mão de obra.

CONTEXTO DA NOMENCLATURA:
- Formato: MO_[FUNÇÃO]
- Exemplos: MO_TEC (técnico), MO_AJU (ajudante), MO_ELE (eletricista)

FUNÇÃO A CADASTRAR: [DESCREVA A FUNÇÃO AQUI]

Responda em formato de tabela com:
1. Código sugerido
2. Categoria (Instalação, Elétrica, Civil, Adicional, Deslocamento)
3. Descrição completa
4. Unidade (H, VZ, DIA)
5. Custo estimado por unidade (Porto Alegre/RS)'''),
        ('', ''),
        ('SEÇÃO', 'PROMPT 3 - CRIAR FERRAMENTA'),
        ('', ''),
        ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar uma nova ferramenta com cálculo de depreciação.

CONTEXTO DA NOMENCLATURA:
- Formato: FER_[TIPO]
- Exemplos: FER_VACUO, FER_MANIF, FER_PERF

FERRAMENTA A CADASTRAR: [DESCREVA A FERRAMENTA AQUI]

Responda em formato de tabela com:
1. Código sugerido
2. Categoria (Vácuo, Manifold, Solda, Furação, Acesso, Teste, Elétrica, Diversos)
3. Descrição completa
4. Valor de aquisição estimado (R$)
5. Vida útil estimada em HORAS de uso
6. Justificativa da vida útil'''),
        ('', ''),
        ('SEÇÃO', 'PROMPT 4 - CRIAR EQUIPAMENTO'),
        ('', ''),
        ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar um novo equipamento de climatização.

CONTEXTO DA NOMENCLATURA:
- Splits Hi-Wall: EQP_HW_[CAPACIDADE]K
- Splits Piso-Teto: EQP_PT_[CAPACIDADE]K
- Cassete: EQP_CASS_[CAPACIDADE]K
- Bombas: EQP_BOMB_[P/M/G]

EQUIPAMENTO A CADASTRAR: [DESCREVA O EQUIPAMENTO AQUI]

Responda em formato de tabela com:
1. Código sugerido
2. Categoria (Split Hi-Wall, Split Piso-Teto, Cassete, Bomba Dreno, etc.)
3. Descrição completa
4. Capacidade em BTUs (se aplicável)
5. Unidade (UN)
6. Preço estimado de mercado (Porto Alegre/RS)'''),
        ('', ''),
        ('SEÇÃO', 'PROMPT 5 - CRIAR COMPOSIÇÃO'),
        ('', ''),
        ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso criar uma nova composição de serviço.

CONTEXTO DA NOMENCLATURA:
- Formato: COMP_[SERVIÇO]_[ESPECIFICAÇÃO]
- Exemplos: COMP_INST_9K, COMP_DRN_PVC, COMP_FURO

MATERIAIS DISPONÍVEIS (exemplos):
TUB_14_FLEX, TUB_38_FLEX, ISO_14_ELA_E9, CAB_PP_25, DRN_CRIS_34, SUP_MF_400, ACA_CAN_50, GAS_R410A, SOL_PRATA, DISJ_M_16

MÃO DE OBRA DISPONÍVEL:
MO_TEC (técnico R$65/h), MO_AJU (ajudante R$35/h), MO_ELE (eletricista R$55/h), MO_PED (pedreiro R$50/h)

FERRAMENTAS DISPONÍVEIS:
FER_VACUO, FER_MANIF, FER_SOLDA, FER_PERF, FER_SERRA_65, FER_ESCADA, FER_MULT, FER_MANUAL

SERVIÇO A CRIAR: [DESCREVA O SERVIÇO AQUI]

Responda com:
1. Código sugerido
2. Descrição completa do serviço
3. Lista de insumos em formato de tabela:
   | Tipo | Código | Quantidade | Unidade |
4. Tempo estimado de execução
5. Observações técnicas importantes'''),
    ]

    row = 1
    for tipo, texto in conteudo:
        if tipo == 'TÍTULO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = Font(bold=True, size=16, color='2E86AB')
            cell.alignment = Alignment(horizontal='center')
            row += 1
        elif tipo == 'SEÇÃO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = estilos['section_font']
            cell.fill = estilos['section_fill']
            row += 1
        elif tipo == 'PROMPT':
            ws.merge_cells(f'A{row}:H{row + texto.count(chr(10))}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = Font(name='Consolas', size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = PatternFill('solid', fgColor='F8F9F9')
            row += texto.count(chr(10)) + 2
        elif tipo == 'TEXTO':
            ws.cell(row=row, column=1, value=texto)
            row += 1
        else:
            row += 1

    ws.column_dimensions['A'].width = 120

    return ws


def preencher(wb, estilos):
    """
    Preenche a aba PROMPTS existente no template.

    Usa a mesma lógica do criar() mas com aba existente.
    """
    ws = wb['PROMPTS']

    # Desfazer mesclagens existentes antes de limpar
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Limpar conteúdo existente
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Mesmo conteúdo do criar()
    conteudo = [
        ('TÍTULO', 'PROMPTS PARA CRIAÇÃO DE NOVOS ITENS COM LLM'),
        ('', ''),
        ('TEXTO', 'Copie o prompt correspondente, substitua o texto entre colchetes e use com ChatGPT, Claude ou outro LLM.'),
        ('', ''),
        ('SEÇÃO', 'PROMPT 1 - CRIAR MATERIAL'),
        ('', ''),
        ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar um novo material na minha planilha de custos.

CONTEXTO DA NOMENCLATURA:
- Tubos: TUB_[POLEGADA]_[FLEX/RIG]
- Isolamentos: ISO_[POLEGADA]_[ELA/POL]_E[ESPESSURA_MM]
- Cabos: CAB_[TIPO]_[SEÇÃO]
- Dreno: DRN_[TIPO]_[ESPECIFICAÇÃO]
- Suportes: SUP_[TIPO]_[ESPECIFICAÇÃO]
- Acabamento: ACA_[TIPO]_[ESPECIFICAÇÃO]
- Gás: GAS_[TIPO] / Solda: SOL_[TIPO]
- Disjuntores: DISJ_[M/B]_[AMPERAGEM]
- Alvenaria: ALV_[TIPO]

MATERIAL A CADASTRAR: [DESCREVA O MATERIAL AQUI]

Responda em formato de tabela com:
1. Código sugerido (seguindo a nomenclatura)
2. Categoria
3. Descrição completa
4. Unidade de medida
5. Preço estimado de mercado (Porto Alegre/RS)'''),
        ('', ''),
        ('SEÇÃO', 'PROMPT 2 - CRIAR COMPOSIÇÃO'),
        ('', ''),
        ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso criar uma nova composição de serviço.

CONTEXTO DA NOMENCLATURA:
- Formato: COMP_[SERVIÇO]_[ESPECIFICAÇÃO]
- Exemplos: COMP_INST_9K, COMP_DRN_PVC, COMP_FURO

SERVIÇO A CRIAR: [DESCREVA O SERVIÇO AQUI]

Responda com:
1. Código sugerido
2. Descrição completa do serviço
3. Lista de insumos em formato de tabela:
   | Tipo | Código | Quantidade Base | Quantidade/Metro |
4. Tempo estimado de execução
5. Observações técnicas importantes'''),
    ]

    row = 1
    for tipo, texto in conteudo:
        if tipo == 'TÍTULO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = Font(bold=True, size=16, color='2E86AB')
            cell.alignment = Alignment(horizontal='center')
            row += 1
        elif tipo == 'SEÇÃO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = estilos['section_font']
            cell.fill = estilos['section_fill']
            row += 1
        elif tipo == 'PROMPT':
            ws.merge_cells(f'A{row}:H{row + texto.count(chr(10))}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = Font(name='Consolas', size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.fill = PatternFill('solid', fgColor='F8F9F9')
            row += texto.count(chr(10)) + 2
        elif tipo == 'TEXTO':
            ws.cell(row=row, column=1, value=texto)
            row += 1
        else:
            row += 1

    ws.column_dimensions['A'].width = 120

    return ws
