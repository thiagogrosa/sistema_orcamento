"""
Aba INSTRUCOES - Guia de uso da planilha.
"""
from openpyxl.styles import Font, PatternFill, Alignment


def criar(wb, estilos):
    """Cria a aba INSTRUCOES com o guia de uso."""
    ws = wb.active
    ws.title = 'INSTRUCOES'

    conteudo = [
        ('TÍTULO', 'PLANILHA DE CUSTOS - INSTALAÇÃO DE EQUIPAMENTOS SPLIT'),
        ('', ''),
        ('SEÇÃO', '1. VISÃO GERAL'),
        ('TEXTO', 'Esta planilha foi desenvolvida para auxiliar na composição de custos e elaboração de orçamentos'),
        ('TEXTO', 'para serviços de instalação de equipamentos de ar-condicionado tipo Split.'),
        ('TEXTO', ''),
        ('TEXTO', 'A estrutura modular permite:'),
        ('TEXTO', '• Calcular custos precisos de materiais, mão de obra e ferramentas'),
        ('TEXTO', '• Criar composições de serviços reutilizáveis'),
        ('TEXTO', '• Montar orçamentos completos de forma rápida'),
        ('TEXTO', '• Atualizar preços de forma centralizada'),
        ('', ''),
        ('SEÇÃO', '2. ESTRUTURA DAS ABAS'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'MATERIAIS'),
        ('TEXTO', 'Contém todos os insumos consumíveis: tubos, cabos, conexões, isolamentos, etc.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'MAO_DE_OBRA'),
        ('TEXTO', 'Profissionais e seus custos por hora: técnico, ajudante, eletricista, pedreiro.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'NEGOCIO'),
        ('TEXTO', 'Configurações de BDI (Benefícios e Despesas Indiretas) ou Markup para cálculo'),
        ('TEXTO', 'do multiplicador aplicado aos custos diretos. Define a margem de lucro e despesas'),
        ('TEXTO', 'indiretas que compõem o preço de venda final.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'FERRAMENTAS'),
        ('TEXTO', 'Ferramentas com cálculo de depreciação. O custo por hora é calculado automaticamente'),
        ('TEXTO', 'dividindo o valor de aquisição pela vida útil em horas.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'EQUIPAMENTOS'),
        ('TEXTO', 'Equipamentos de climatização: splits de diversas capacidades, bombas de dreno.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'COMPOSICOES'),
        ('TEXTO', 'Serviços compostos por materiais, mão de obra e ferramentas. O total de cada'),
        ('TEXTO', 'composição é calculado automaticamente e aparece na linha do código.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'ESCOPO'),
        ('TEXTO', 'Aba para montagem de orçamentos. Permite selecionar composições, equipamentos,'),
        ('TEXTO', 'materiais ou mão de obra avulsa e calcular o total automaticamente.'),
        ('', ''),
        ('SEÇÃO', '3. SISTEMA DE NOMENCLATURA'),
        ('TEXTO', ''),
        ('TEXTO', 'Os códigos seguem um padrão intuitivo para facilitar a identificação:'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'TUBULAÇÃO DE COBRE'),
        ('TEXTO', 'Formato: TUB_[POLEGADA]_[TIPO]'),
        ('TEXTO', '• FLEX = Flexível (espessura ~1/32" ou 0,79mm)'),
        ('TEXTO', '• RIG = Rígido (espessura ~1/16" ou 1,58mm)'),
        ('TEXTO', 'Exemplos: TUB_14_FLEX (1/4" flexível), TUB_34_RIG (3/4" rígido)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'ISOLAMENTO TÉRMICO'),
        ('TEXTO', 'Formato: ISO_[POLEGADA]_[TIPO]_E[ESPESSURA]'),
        ('TEXTO', '• ELA = Elastomérico (Armaflex, K-Flex)'),
        ('TEXTO', '• POL = Polietileno blindado'),
        ('TEXTO', '• E9, E10, E13, E19, E25 = Espessura em mm'),
        ('TEXTO', 'Exemplos: ISO_14_ELA_E9 (p/ 1/4", elastom. 9mm), ISO_38_POL_E10 (p/ 3/8", polie. 10mm)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'CABOS ELÉTRICOS'),
        ('TEXTO', 'Formato: CAB_[TIPO]_[SEÇÃO]'),
        ('TEXTO', 'Exemplos: CAB_PP_25 (PP 3x2,5mm²), CAB_FLEX_40 (Flexível 4mm²), CAB_COM (comunicação)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'DRENO'),
        ('TEXTO', 'Formato: DRN_[TIPO]_[DIÂMETRO]'),
        ('TEXTO', 'Exemplos: DRN_CRIS_34 (mangueira cristal 3/4"), DRN_PVC_25 (tubo PVC 25mm)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'SUPORTES'),
        ('TEXTO', 'Formato: SUP_[TIPO]_[ESPECIFICAÇÃO]'),
        ('TEXTO', 'Exemplos: SUP_MF_400 (mão francesa 400mm), SUP_CALCO (calços borracha)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'ACABAMENTO'),
        ('TEXTO', 'Formato: ACA_[TIPO]_[ESPECIFICAÇÃO]'),
        ('TEXTO', 'Exemplos: ACA_CAN_50 (canaleta 50mm), ACA_ESPUMA (espuma expansiva)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'GÁS E SOLDA'),
        ('TEXTO', 'Formatos: GAS_[TIPO] / SOL_[TIPO]'),
        ('TEXTO', 'Exemplos: GAS_R410A, GAS_R22, SOL_PRATA, SOL_FLUXO'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'DISJUNTORES'),
        ('TEXTO', 'Formato: DISJ_[M/B]_[AMPERAGEM]'),
        ('TEXTO', '• M = Monopolar, B = Bipolar'),
        ('TEXTO', 'Exemplos: DISJ_M_16 (mono 16A), DISJ_B_25 (bi 25A)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'MÃO DE OBRA'),
        ('TEXTO', 'Formato: MO_[FUNÇÃO]'),
        ('TEXTO', 'Exemplos: MO_TEC (técnico), MO_AJU (ajudante), MO_ELE (eletricista)'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'FERRAMENTAS'),
        ('TEXTO', 'Formato: FER_[TIPO]'),
        ('TEXTO', 'Exemplos: FER_VACUO, FER_MANIF, FER_PERF, FER_ESCADA'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'EQUIPAMENTOS'),
        ('TEXTO', 'Formato: EQP_[TIPO]_[CAPACIDADE]'),
        ('TEXTO', '• HW = Hi-Wall, PT = Piso-Teto, BOMB = Bomba dreno'),
        ('TEXTO', 'Exemplos: EQP_HW_9K, EQP_PT_36K, EQP_BOMB_P'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'COMPOSIÇÕES'),
        ('TEXTO', 'Formato: COMP_[SERVIÇO]_[ESPECIFICAÇÃO]'),
        ('TEXTO', 'Exemplos: COMP_INST_9K, COMP_DRN_PVC, COMP_FURO, COMP_ALV_3M'),
        ('', ''),
        ('SEÇÃO', '4. COMO INSERIR NOVOS ITENS'),
        ('TEXTO', ''),
        ('TEXTO', 'IMPORTANTE: Cada aba possui uma coluna "Seleção" oculta (última coluna) que'),
        ('TEXTO', 'concatena código + descrição para os dropdowns. Esta coluna usa fórmulas,'),
        ('TEXTO', 'então ao adicionar novos itens, copie a fórmula da linha anterior.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'Materiais, Mão de Obra, Equipamentos:'),
        ('TEXTO', '1. Vá até a última linha preenchida da aba correspondente'),
        ('TEXTO', '2. Insira uma nova linha abaixo'),
        ('TEXTO', '3. Preencha o código seguindo a nomenclatura padrão'),
        ('TEXTO', '4. Preencha os demais campos'),
        ('TEXTO', '5. Copie a fórmula da coluna Seleção (última coluna, oculta) da linha anterior'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'Ferramentas:'),
        ('TEXTO', '1. Preencha código, categoria, descrição, valor de aquisição e vida útil em HORAS'),
        ('TEXTO', '2. A coluna Custo/Hora será calculada automaticamente (=Valor/Vida Útil)'),
        ('TEXTO', '3. Copie a fórmula da coluna Seleção da linha anterior'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'Composições:'),
        ('TEXTO', '1. Insira uma linha de cabeçalho com código e descrição da composição'),
        ('TEXTO', '2. Nas linhas abaixo, selecione o TIPO (MAT, MO, FER, EQP) no dropdown'),
        ('TEXTO', '3. Selecione o código do item no dropdown da coluna seguinte'),
        ('TEXTO', '4. Informe a quantidade'),
        ('TEXTO', '5. Os campos descrição, unidade, preço e subtotal são automáticos'),
        ('TEXTO', '6. Na linha do cabeçalho, a coluna TOTAL soma todos os subtotais'),
        ('TEXTO', '7. Copie a fórmula da coluna Seleção (L) da linha de cabeçalho anterior'),
        ('', ''),
        ('SEÇÃO', '5. COMO CONFIGURAR BDI/MARKUP'),
        ('TEXTO', ''),
        ('TEXTO', '1. Acesse a aba NEGOCIO'),
        ('TEXTO', '2. Escolha o método de cálculo:'),
        ('TEXTO', '   • BDI: Para composição detalhada de custos indiretos'),
        ('TEXTO', '   • MARKUP: Para aplicação de percentual único'),
        ('TEXTO', '3. Se escolher BDI, ajuste os componentes:'),
        ('TEXTO', '   • Administração Central: custos de escritório, gerência'),
        ('TEXTO', '   • Despesas Financeiras: juros, taxas bancárias'),
        ('TEXTO', '   • Lucro: margem de lucro desejada'),
        ('TEXTO', '   • Impostos: ISS, PIS, COFINS (varia por município)'),
        ('TEXTO', '   • Riscos e Imprevistos: contingências'),
        ('TEXTO', '4. Se escolher MARKUP, defina o percentual único'),
        ('TEXTO', '5. O multiplicador é calculado automaticamente'),
        ('TEXTO', '6. O multiplicador é aplicado automaticamente no orçamento'),
        ('TEXTO', ''),
        ('SEÇÃO', '6. COMO MONTAR UM ORÇAMENTO'),
        ('TEXTO', ''),
        ('TEXTO', '1. Configure BDI/MARKUP na aba NEGOCIO (veja seção anterior)'),
        ('TEXTO', '2. Preencha os dados do cliente no cabeçalho'),
        ('TEXTO', '3. Na tabela de itens, selecione o TIPO:'),
        ('TEXTO', '   • COMP = Composição de serviço'),
        ('TEXTO', '   • EQP = Equipamento'),
        ('TEXTO', '   • MAT = Material avulso'),
        ('TEXTO', '   • MO = Mão de obra avulsa'),
        ('TEXTO', '   • FER = Ferramenta (custo por hora)'),
        ('TEXTO', '4. Selecione o código correspondente'),
        ('TEXTO', '5. Informe a quantidade e variável (metros de linha, cabo, etc. - se aplicável)'),
        ('TEXTO', '6. A descrição mostra automaticamente a metragem: "com 5 metros de linha frigorígena"'),
        ('TEXTO', '7. Os campos descrição, preço e total são calculados automaticamente'),
        ('TEXTO', '8. O BDI/MARKUP é aplicado automaticamente sobre o subtotal'),
        ('TEXTO', '9. Aplique desconto se necessário'),
        ('TEXTO', '10. O TOTAL GERAL é calculado automaticamente'),
        ('', ''),
        ('SEÇÃO', '7. DICAS E BOAS PRÁTICAS'),
        ('TEXTO', ''),
        ('TEXTO', '• Mantenha os preços atualizados nas abas de origem (MATERIAIS, MAO_DE_OBRA, etc.)'),
        ('TEXTO', '• Todas as composições e orçamentos serão atualizados automaticamente'),
        ('TEXTO', '• Use a aba PROMPTS para criar novos itens com ajuda de IA'),
        ('TEXTO', '• Faça backup da planilha periodicamente'),
        ('TEXTO', '• Revise as composições semestralmente para ajustar quantidades e tempos'),
    ]

    row = 1
    for tipo, texto in conteudo:
        if tipo == 'TÍTULO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = Font(bold=True, size=16, color='2E86AB')
            cell.alignment = Alignment(horizontal='center')
            row += 1
        elif tipo == 'SEÇÃO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = estilos['section_font']
            cell.fill = estilos['section_fill']
            row += 1
        elif tipo == 'SUBTÍTULO':
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = estilos['subtitle_font']
            row += 1
        elif tipo == 'TEXTO':
            ws.cell(row=row, column=1, value=texto)
            row += 1
        else:
            row += 1

    ws.column_dimensions['A'].width = 100

    return ws


def preencher(wb, estilos):
    """
    Preenche a aba INSTRUCOES existente no template.

    Limpa conteúdo anterior e repopula com os dados.
    """
    ws = wb['INSTRUCOES']

    # Desfazer mesclagens existentes antes de limpar
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Limpar conteúdo existente
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Reaproveitar lógica do criar()
    conteudo = [
        ('TÍTULO', 'PLANILHA DE CUSTOS - INSTALAÇÃO DE EQUIPAMENTOS SPLIT'),
        ('', ''),
        ('SEÇÃO', '1. VISÃO GERAL'),
        ('TEXTO', 'Esta planilha foi desenvolvida para auxiliar na composição de custos e elaboração de orçamentos'),
        ('TEXTO', 'para serviços de instalação de equipamentos de ar-condicionado tipo Split.'),
        ('TEXTO', ''),
        ('TEXTO', 'A estrutura modular permite:'),
        ('TEXTO', '• Calcular custos precisos de materiais, mão de obra e ferramentas'),
        ('TEXTO', '• Criar composições de serviços reutilizáveis'),
        ('TEXTO', '• Montar orçamentos completos de forma rápida'),
        ('TEXTO', '• Atualizar preços de forma centralizada'),
        ('', ''),
        ('SEÇÃO', '2. ESTRUTURA DAS ABAS'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'MATERIAIS'),
        ('TEXTO', 'Contém todos os insumos consumíveis: tubos, cabos, conexões, isolamentos, etc.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'MAO_DE_OBRA'),
        ('TEXTO', 'Profissionais e seus custos por hora: técnico, ajudante, eletricista, pedreiro.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'NEGOCIO'),
        ('TEXTO', 'Configurações de BDI (Benefícios e Despesas Indiretas) ou Markup para cálculo'),
        ('TEXTO', 'do multiplicador aplicado aos custos diretos. Define a margem de lucro e despesas'),
        ('TEXTO', 'indiretas que compõem o preço de venda final.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'FERRAMENTAS'),
        ('TEXTO', 'Ferramentas com cálculo de depreciação. O custo por hora é calculado automaticamente'),
        ('TEXTO', 'dividindo o valor de aquisição pela vida útil em horas.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'EQUIPAMENTOS'),
        ('TEXTO', 'Equipamentos de climatização: splits de diversas capacidades, bombas de dreno.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'COMPOSICOES'),
        ('TEXTO', 'Serviços compostos por materiais, mão de obra e ferramentas. O total de cada'),
        ('TEXTO', 'composição é calculado automaticamente e aparece na linha do código.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'ESCOPO'),
        ('TEXTO', 'Aba para montagem de orçamentos. Permite selecionar composições, equipamentos,'),
        ('TEXTO', 'materiais ou mão de obra avulsa e calcular o total automaticamente.'),
        ('', ''),
        ('SEÇÃO', '3. MACROS DISPONÍVEIS'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'AtualizarSelecao'),
        ('TEXTO', 'Atualiza as fórmulas da coluna Seleção na aba ativa. Use após adicionar novos itens'),
        ('TEXTO', 'em qualquer catálogo ou composição.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'NovaComposicao'),
        ('TEXTO', 'Abre um formulário para criar uma nova composição de serviço com formatação automática.'),
        ('TEXTO', ''),
        ('SUBTÍTULO', 'AtualizarTotaisComposicao'),
        ('TEXTO', 'Recalcula as fórmulas de Total Base e Total/Metro da composição atual.'),
        ('TEXTO', ''),
        ('SEÇÃO', '4. SISTEMA DE NOMENCLATURA'),
        ('TEXTO', ''),
        ('TEXTO', 'Os códigos seguem um padrão intuitivo para facilitar a identificação.'),
        ('TEXTO', 'Veja a aba PROMPTS para detalhes e prompts de criação com IA.'),
    ]

    row = 1
    for tipo, texto in conteudo:
        if tipo == 'TÍTULO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = Font(bold=True, size=16, color='2E86AB')
            cell.alignment = Alignment(horizontal='center')
            row += 1
        elif tipo == 'SEÇÃO':
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = estilos['section_font']
            cell.fill = estilos['section_fill']
            row += 1
        elif tipo == 'SUBTÍTULO':
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = estilos['subtitle_font']
            row += 1
        elif tipo == 'TEXTO':
            ws.cell(row=row, column=1, value=texto)
            row += 1
        else:
            row += 1

    ws.column_dimensions['A'].width = 100

    return ws
