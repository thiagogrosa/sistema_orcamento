from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

wb = Workbook()

# ============================================
# ESTILOS GLOBAIS
# ============================================
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2E86AB')
title_font = Font(bold=True, size=14, color='2E86AB')
subtitle_font = Font(bold=True, size=12)
section_font = Font(bold=True, size=11, color='FFFFFF')
section_fill = PatternFill('solid', fgColor='5D6D7E')
code_font = Font(name='Consolas', size=10)
currency_format = 'R$ #,##0.00'
qty_format = '#,##0.00'
hours_format = '#,##0'

# Cores por categoria de material
cat_colors = {
    'Tubulação': 'D6EAF8',
    'Isolamento': 'D5F5E3',
    'Elétrica': 'FCF3CF',
    'Eletroduto': 'FCF3CF',
    'Dreno': 'D1F2EB',
    'Suporte': 'FAE5D3',
    'Acabamento': 'FADBD8',
    'Gás/Solda': 'E5E8E8',
    'Alvenaria': 'F5EEF8',
    'Proteção': 'FCF3CF',
}

# Cores para tipos nas composições
mo_fill = PatternFill('solid', fgColor='EAFAF1')
fer_fill = PatternFill('solid', fgColor='FEF9E7')
eqp_fill = PatternFill('solid', fgColor='EBF5FB')
comp_header_fill = PatternFill('solid', fgColor='2E86AB')
total_fill = PatternFill('solid', fgColor='BDC3C7')
input_fill = PatternFill('solid', fgColor='FFFACD')

# ============================================
# ABA 1 - INSTRUÇÕES
# ============================================
ws_inst = wb.active
ws_inst.title = 'INSTRUCOES'

# Conteúdo das instruções
instrucoes_conteudo = [
    ('TÍTULO', 'PLANILHA DE CUSTOS - INSTALAÇÃO DE EQUIPAMENTOS SPLIT'),
    ('', ''),
    ('SEÇÃO', '1. VISÃO GERAL'),
    ('TEXTO', 'Esta planilha foi desenvolvida para auxiliar na composição de custos e elaboração de orçamentos'),
    ('TEXTO', 'para serviços de instalação de equipamentos de ar-condicionado tipo Split.'),
    ('TEXTO', ''),
    ('TEXTO', 'A estrutura modular permite:'),
    ('TEXTO', '• Calcular custos precisos de materiais, mão de obra e ferramentas'),
    ('TEXTO', '• Criar composições de serviços reutilizáveis'),
    ('TEXTO', '• Montar orçamentos completos de forma rápida'),
    ('TEXTO', '• Atualizar preços de forma centralizada'),
    ('', ''),
    ('SEÇÃO', '2. ESTRUTURA DAS ABAS'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'MATERIAIS'),
    ('TEXTO', 'Contém todos os insumos consumíveis: tubos, cabos, conexões, isolamentos, etc.'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'MAO_DE_OBRA'),
    ('TEXTO', 'Profissionais e seus custos por hora: técnico, ajudante, eletricista, pedreiro.'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'NEGOCIO'),
    ('TEXTO', 'Configurações de BDI (Benefícios e Despesas Indiretas) ou Markup para cálculo'),
    ('TEXTO', 'do multiplicador aplicado aos custos diretos. Define a margem de lucro e despesas'),
    ('TEXTO', 'indiretas que compõem o preço de venda final.'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'FERRAMENTAS'),
    ('TEXTO', 'Ferramentas com cálculo de depreciação. O custo por hora é calculado automaticamente'),
    ('TEXTO', 'dividindo o valor de aquisição pela vida útil em horas.'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'EQUIPAMENTOS'),
    ('TEXTO', 'Equipamentos de climatização: splits de diversas capacidades, bombas de dreno.'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'COMPOSICOES'),
    ('TEXTO', 'Serviços compostos por materiais, mão de obra e ferramentas. O total de cada'),
    ('TEXTO', 'composição é calculado automaticamente e aparece na linha do código.'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'ORCAMENTO'),
    ('TEXTO', 'Aba para montagem de orçamentos. Permite selecionar composições, equipamentos,'),
    ('TEXTO', 'materiais ou mão de obra avulsa e calcular o total automaticamente.'),
    ('', ''),
    ('SEÇÃO', '3. SISTEMA DE NOMENCLATURA'),
    ('TEXTO', ''),
    ('TEXTO', 'Os códigos seguem um padrão intuitivo para facilitar a identificação:'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'TUBULAÇÃO DE COBRE'),
    ('TEXTO', 'Formato: TUB_[POLEGADA]_[TIPO]'),
    ('TEXTO', '• FLEX = Flexível (espessura ~1/32" ou 0,79mm)'),
    ('TEXTO', '• RIG = Rígido (espessura ~1/16" ou 1,58mm)'),
    ('TEXTO', 'Exemplos: TUB_14_FLEX (1/4" flexível), TUB_34_RIG (3/4" rígido)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'ISOLAMENTO TÉRMICO'),
    ('TEXTO', 'Formato: ISO_[POLEGADA]_[TIPO]_E[ESPESSURA]'),
    ('TEXTO', '• ELA = Elastomérico (Armaflex, K-Flex)'),
    ('TEXTO', '• POL = Polietileno blindado'),
    ('TEXTO', '• E9, E10, E13, E19, E25 = Espessura em mm'),
    ('TEXTO', 'Exemplos: ISO_14_ELA_E9 (p/ 1/4", elastom. 9mm), ISO_38_POL_E10 (p/ 3/8", polie. 10mm)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'CABOS ELÉTRICOS'),
    ('TEXTO', 'Formato: CAB_[TIPO]_[SEÇÃO]'),
    ('TEXTO', 'Exemplos: CAB_PP_25 (PP 3x2,5mm²), CAB_FLEX_40 (Flexível 4mm²), CAB_COM (comunicação)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'DRENO'),
    ('TEXTO', 'Formato: DRN_[TIPO]_[DIÂMETRO]'),
    ('TEXTO', 'Exemplos: DRN_CRIS_34 (mangueira cristal 3/4"), DRN_PVC_25 (tubo PVC 25mm)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'SUPORTES'),
    ('TEXTO', 'Formato: SUP_[TIPO]_[ESPECIFICAÇÃO]'),
    ('TEXTO', 'Exemplos: SUP_MF_400 (mão francesa 400mm), SUP_CALCO (calços borracha)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'ACABAMENTO'),
    ('TEXTO', 'Formato: ACA_[TIPO]_[ESPECIFICAÇÃO]'),
    ('TEXTO', 'Exemplos: ACA_CAN_50 (canaleta 50mm), ACA_ESPUMA (espuma expansiva)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'GÁS E SOLDA'),
    ('TEXTO', 'Formatos: GAS_[TIPO] / SOL_[TIPO]'),
    ('TEXTO', 'Exemplos: GAS_R410A, GAS_R22, SOL_PRATA, SOL_FLUXO'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'DISJUNTORES'),
    ('TEXTO', 'Formato: DISJ_[M/B]_[AMPERAGEM]'),
    ('TEXTO', '• M = Monopolar, B = Bipolar'),
    ('TEXTO', 'Exemplos: DISJ_M_16 (mono 16A), DISJ_B_25 (bi 25A)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'MÃO DE OBRA'),
    ('TEXTO', 'Formato: MO_[FUNÇÃO]'),
    ('TEXTO', 'Exemplos: MO_TEC (técnico), MO_AJU (ajudante), MO_ELE (eletricista)'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'FERRAMENTAS'),
    ('TEXTO', 'Formato: FER_[TIPO]'),
    ('TEXTO', 'Exemplos: FER_VACUO, FER_MANIF, FER_PERF, FER_ESCADA'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'EQUIPAMENTOS'),
    ('TEXTO', 'Formato: EQP_[TIPO]_[CAPACIDADE]'),
    ('TEXTO', '• HW = Hi-Wall, PT = Piso-Teto, BOMB = Bomba dreno'),
    ('TEXTO', 'Exemplos: EQP_HW_9K, EQP_PT_36K, EQP_BOMB_P'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'COMPOSIÇÕES'),
    ('TEXTO', 'Formato: COMP_[SERVIÇO]_[ESPECIFICAÇÃO]'),
    ('TEXTO', 'Exemplos: COMP_INST_9K, COMP_DRN_PVC, COMP_FURO, COMP_ALV_3M'),
    ('', ''),
    ('SEÇÃO', '4. COMO INSERIR NOVOS ITENS'),
    ('TEXTO', ''),
    ('TEXTO', 'IMPORTANTE: Cada aba possui uma coluna "Seleção" oculta (última coluna) que'),
    ('TEXTO', 'concatena código + descrição para os dropdowns. Esta coluna usa fórmulas,'),
    ('TEXTO', 'então ao adicionar novos itens, copie a fórmula da linha anterior.'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'Materiais, Mão de Obra, Equipamentos:'),
    ('TEXTO', '1. Vá até a última linha preenchida da aba correspondente'),
    ('TEXTO', '2. Insira uma nova linha abaixo'),
    ('TEXTO', '3. Preencha o código seguindo a nomenclatura padrão'),
    ('TEXTO', '4. Preencha os demais campos'),
    ('TEXTO', '5. Copie a fórmula da coluna Seleção (última coluna, oculta) da linha anterior'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'Ferramentas:'),
    ('TEXTO', '1. Preencha código, categoria, descrição, valor de aquisição e vida útil em HORAS'),
    ('TEXTO', '2. A coluna Custo/Hora será calculada automaticamente (=Valor/Vida Útil)'),
    ('TEXTO', '3. Copie a fórmula da coluna Seleção da linha anterior'),
    ('TEXTO', ''),
    ('SUBTÍTULO', 'Composições:'),
    ('TEXTO', '1. Insira uma linha de cabeçalho com código e descrição da composição'),
    ('TEXTO', '2. Nas linhas abaixo, selecione o TIPO (MAT, MO, FER, EQP) no dropdown'),
    ('TEXTO', '3. Selecione o código do item no dropdown da coluna seguinte'),
    ('TEXTO', '4. Informe a quantidade'),
    ('TEXTO', '5. Os campos descrição, unidade, preço e subtotal são automáticos'),
    ('TEXTO', '6. Na linha do cabeçalho, a coluna TOTAL soma todos os subtotais'),
    ('TEXTO', '7. Copie a fórmula da coluna Seleção (L) da linha de cabeçalho anterior'),
    ('', ''),
    ('SEÇÃO', '5. COMO CONFIGURAR BDI/MARKUP'),
    ('TEXTO', ''),
    ('TEXTO', '1. Acesse a aba NEGOCIO'),
    ('TEXTO', '2. Escolha o método de cálculo:'),
    ('TEXTO', '   • BDI: Para composição detalhada de custos indiretos'),
    ('TEXTO', '   • MARKUP: Para aplicação de percentual único'),
    ('TEXTO', '3. Se escolher BDI, ajuste os componentes:'),
    ('TEXTO', '   • Administração Central: custos de escritório, gerência'),
    ('TEXTO', '   • Despesas Financeiras: juros, taxas bancárias'),
    ('TEXTO', '   • Lucro: margem de lucro desejada'),
    ('TEXTO', '   • Impostos: ISS, PIS, COFINS (varia por município)'),
    ('TEXTO', '   • Riscos e Imprevistos: contingências'),
    ('TEXTO', '4. Se escolher MARKUP, defina o percentual único'),
    ('TEXTO', '5. O multiplicador é calculado automaticamente'),
    ('TEXTO', '6. O multiplicador é aplicado automaticamente no orçamento'),
    ('TEXTO', ''),
    ('SEÇÃO', '6. COMO MONTAR UM ORÇAMENTO'),
    ('TEXTO', ''),
    ('TEXTO', '1. Configure BDI/MARKUP na aba NEGOCIO (veja seção anterior)'),
    ('TEXTO', '2. Preencha os dados do cliente no cabeçalho'),
    ('TEXTO', '3. Na tabela de itens, selecione o TIPO:'),
    ('TEXTO', '   • COMP = Composição de serviço'),
    ('TEXTO', '   • EQP = Equipamento'),
    ('TEXTO', '   • MAT = Material avulso'),
    ('TEXTO', '   • MO = Mão de obra avulsa'),
    ('TEXTO', '   • FER = Ferramenta (custo por hora)'),
    ('TEXTO', '4. Selecione o código correspondente'),
    ('TEXTO', '5. Informe a quantidade e variável (metros de linha, cabo, etc. - se aplicável)'),
    ('TEXTO', '6. A descrição mostra automaticamente a metragem: "com 5 metros de linha frigorígena"'),
    ('TEXTO', '7. Os campos descrição, preço e total são calculados automaticamente'),
    ('TEXTO', '8. O BDI/MARKUP é aplicado automaticamente sobre o subtotal'),
    ('TEXTO', '9. Aplique desconto se necessário'),
    ('TEXTO', '10. O TOTAL GERAL é calculado automaticamente'),
    ('', ''),
    ('SEÇÃO', '7. DICAS E BOAS PRÁTICAS'),
    ('TEXTO', ''),
    ('TEXTO', '• Mantenha os preços atualizados nas abas de origem (MATERIAIS, MAO_DE_OBRA, etc.)'),
    ('TEXTO', '• Todas as composições e orçamentos serão atualizados automaticamente'),
    ('TEXTO', '• Use a aba PROMPTS para criar novos itens com ajuda de IA'),
    ('TEXTO', '• Faça backup da planilha periodicamente'),
    ('TEXTO', '• Revise as composições semestralmente para ajustar quantidades e tempos'),
]

row = 1
for tipo, texto in instrucoes_conteudo:
    if tipo == 'TÍTULO':
        ws_inst.merge_cells(f'A{row}:H{row}')
        cell = ws_inst.cell(row=row, column=1, value=texto)
        cell.font = Font(bold=True, size=16, color='2E86AB')
        cell.alignment = Alignment(horizontal='center')
        row += 1
    elif tipo == 'SEÇÃO':
        ws_inst.merge_cells(f'A{row}:H{row}')
        cell = ws_inst.cell(row=row, column=1, value=texto)
        cell.font = section_font
        cell.fill = section_fill
        row += 1
    elif tipo == 'SUBTÍTULO':
        cell = ws_inst.cell(row=row, column=1, value=texto)
        cell.font = subtitle_font
        row += 1
    elif tipo == 'TEXTO':
        cell = ws_inst.cell(row=row, column=1, value=texto)
        row += 1
    else:
        row += 1

ws_inst.column_dimensions['A'].width = 100

# ============================================
# ABA 2 - PROMPTS
# ============================================
ws_prompts = wb.create_sheet('PROMPTS')

prompts_conteudo = [
    ('TÍTULO', 'PROMPTS PARA CRIAÇÃO DE NOVOS ITENS COM LLM'),
    ('', ''),
    ('TEXTO', 'Copie o prompt correspondente, substitua o texto entre colchetes e use com ChatGPT, Claude ou outro LLM.'),
    ('', ''),
    ('SEÇÃO', 'PROMPT 1 - CRIAR MATERIAL'),
    ('', ''),
    ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar um novo material na minha planilha de custos.

CONTEXTO DA NOMENCLATURA:
- Tubos: TUB_[POLEGADA]_[FLEX/RIG]
- Isolamentos: ISO_[POLEGADA]_[ELA/POL]_E[ESPESSURA_MM]
- Cabos: CAB_[TIPO]_[SEÇÃO]
- Dreno: DRN_[TIPO]_[ESPECIFICAÇÃO]
- Suportes: SUP_[TIPO]_[ESPECIFICAÇÃO]
- Acabamento: ACA_[TIPO]_[ESPECIFICAÇÃO]
- Gás: GAS_[TIPO] / Solda: SOL_[TIPO]
- Disjuntores: DISJ_[M/B]_[AMPERAGEM]
- Alvenaria: ALV_[TIPO]

MATERIAL A CADASTRAR: [DESCREVA O MATERIAL AQUI]

Responda em formato de tabela com:
1. Código sugerido (seguindo a nomenclatura)
2. Categoria
3. Descrição completa
4. Unidade de medida
5. Preço estimado de mercado (Porto Alegre/RS)'''),
    ('', ''),
    ('SEÇÃO', 'PROMPT 2 - CRIAR MÃO DE OBRA'),
    ('', ''),
    ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar uma nova função de mão de obra.

CONTEXTO DA NOMENCLATURA:
- Formato: MO_[FUNÇÃO]
- Exemplos: MO_TEC (técnico), MO_AJU (ajudante), MO_ELE (eletricista)

FUNÇÃO A CADASTRAR: [DESCREVA A FUNÇÃO AQUI]

Responda em formato de tabela com:
1. Código sugerido
2. Categoria (Instalação, Elétrica, Civil, Adicional, Deslocamento)
3. Descrição completa
4. Unidade (H, VZ, DIA)
5. Custo estimado por unidade (Porto Alegre/RS)'''),
    ('', ''),
    ('SEÇÃO', 'PROMPT 3 - CRIAR FERRAMENTA'),
    ('', ''),
    ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar uma nova ferramenta com cálculo de depreciação.

CONTEXTO DA NOMENCLATURA:
- Formato: FER_[TIPO]
- Exemplos: FER_VACUO, FER_MANIF, FER_PERF

FERRAMENTA A CADASTRAR: [DESCREVA A FERRAMENTA AQUI]

Responda em formato de tabela com:
1. Código sugerido
2. Categoria (Vácuo, Manifold, Solda, Furação, Acesso, Teste, Elétrica, Diversos)
3. Descrição completa
4. Valor de aquisição estimado (R$)
5. Vida útil estimada em HORAS de uso
6. Justificativa da vida útil'''),
    ('', ''),
    ('SEÇÃO', 'PROMPT 4 - CRIAR EQUIPAMENTO'),
    ('', ''),
    ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso cadastrar um novo equipamento de climatização.

CONTEXTO DA NOMENCLATURA:
- Splits Hi-Wall: EQP_HW_[CAPACIDADE]K
- Splits Piso-Teto: EQP_PT_[CAPACIDADE]K
- Cassete: EQP_CASS_[CAPACIDADE]K
- Bombas: EQP_BOMB_[P/M/G]

EQUIPAMENTO A CADASTRAR: [DESCREVA O EQUIPAMENTO AQUI]

Responda em formato de tabela com:
1. Código sugerido
2. Categoria (Split Hi-Wall, Split Piso-Teto, Cassete, Bomba Dreno, etc.)
3. Descrição completa
4. Capacidade em BTUs (se aplicável)
5. Unidade (UN)
6. Preço estimado de mercado (Porto Alegre/RS)'''),
    ('', ''),
    ('SEÇÃO', 'PROMPT 5 - CRIAR COMPOSIÇÃO'),
    ('', ''),
    ('PROMPT', '''Você é um assistente especializado em HVAC. Preciso criar uma nova composição de serviço.

CONTEXTO DA NOMENCLATURA:
- Formato: COMP_[SERVIÇO]_[ESPECIFICAÇÃO]
- Exemplos: COMP_INST_9K, COMP_DRN_PVC, COMP_FURO

MATERIAIS DISPONÍVEIS (exemplos):
TUB_14_FLEX, TUB_38_FLEX, ISO_14_ELA_E9, CAB_PP_25, DRN_CRIS_34, SUP_MF_400, ACA_CAN_50, GAS_R410A, SOL_PRATA, DISJ_M_16

MÃO DE OBRA DISPONÍVEL:
MO_TEC (técnico R$65/h), MO_AJU (ajudante R$35/h), MO_ELE (eletricista R$55/h), MO_PED (pedreiro R$50/h)

FERRAMENTAS DISPONÍVEIS:
FER_VACUO, FER_MANIF, FER_SOLDA, FER_PERF, FER_SERRA_65, FER_ESCADA, FER_MULT, FER_MANUAL

SERVIÇO A CRIAR: [DESCREVA O SERVIÇO AQUI]

Responda com:
1. Código sugerido
2. Descrição completa do serviço
3. Lista de insumos em formato de tabela:
   | Tipo | Código | Quantidade | Unidade |
4. Tempo estimado de execução
5. Observações técnicas importantes'''),
]

row = 1
for tipo, texto in prompts_conteudo:
    if tipo == 'TÍTULO':
        ws_prompts.merge_cells(f'A{row}:H{row}')
        cell = ws_prompts.cell(row=row, column=1, value=texto)
        cell.font = Font(bold=True, size=16, color='2E86AB')
        cell.alignment = Alignment(horizontal='center')
        row += 1
    elif tipo == 'SEÇÃO':
        ws_prompts.merge_cells(f'A{row}:H{row}')
        cell = ws_prompts.cell(row=row, column=1, value=texto)
        cell.font = section_font
        cell.fill = section_fill
        row += 1
    elif tipo == 'PROMPT':
        ws_prompts.merge_cells(f'A{row}:H{row + texto.count(chr(10))}')
        cell = ws_prompts.cell(row=row, column=1, value=texto)
        cell.font = Font(name='Consolas', size=10)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.fill = PatternFill('solid', fgColor='F8F9F9')
        row += texto.count(chr(10)) + 2
    elif tipo == 'TEXTO':
        cell = ws_prompts.cell(row=row, column=1, value=texto)
        row += 1
    else:
        row += 1

ws_prompts.column_dimensions['A'].width = 120

# ============================================
# ABA 3 - NEGÓCIO (BDI / MARKUP)
# ============================================
ws_neg = wb.create_sheet('NEGOCIO')

# Título
ws_neg.merge_cells('A1:F1')
cell_titulo_neg = ws_neg.cell(row=1, column=1, value='CONFIGURAÇÕES DE NEGÓCIO - BDI E MARKUP')
cell_titulo_neg.font = Font(bold=True, size=16, color='2E86AB')
cell_titulo_neg.alignment = Alignment(horizontal='center')

# Explicação
ws_neg.cell(row=3, column=1, value='Esta aba permite configurar o multiplicador aplicado ao custo direto para obter o preço de venda.')
ws_neg.merge_cells('A3:F3')

# Seção: Método
ws_neg.merge_cells('A5:F5')
cell_metodo_label = ws_neg.cell(row=5, column=1, value='MÉTODO DE CÁLCULO')
cell_metodo_label.font = section_font
cell_metodo_label.fill = section_fill

ws_neg.cell(row=7, column=1, value='Método:').font = Font(bold=True)
ws_neg.cell(row=7, column=2, value='BDI')
ws_neg.cell(row=7, column=2).fill = input_fill
ws_neg.cell(row=7, column=2).border = thin_border
ws_neg.cell(row=7, column=3, value='(Escolha: BDI ou MARKUP)').font = Font(italic=True, size=9)

# Validação para método
dv_metodo = DataValidation(type="list", formula1='"BDI,MARKUP"', allow_blank=False)
ws_neg.add_data_validation(dv_metodo)
dv_metodo.add('B7')

# Seção: Componentes do BDI
ws_neg.merge_cells('A9:F9')
cell_bdi_label = ws_neg.cell(row=9, column=1, value='COMPONENTES DO BDI')
cell_bdi_label.font = section_font
cell_bdi_label.fill = section_fill

# Cabeçalho da tabela BDI
bdi_headers_row = 11
ws_neg.cell(row=bdi_headers_row, column=1, value='Componente').font = Font(bold=True)
ws_neg.cell(row=bdi_headers_row, column=2, value='%').font = Font(bold=True)
ws_neg.cell(row=bdi_headers_row, column=3, value='Descrição').font = Font(bold=True)

for col in range(1, 4):
    ws_neg.cell(row=bdi_headers_row, column=col).fill = PatternFill('solid', fgColor='D5D8DC')
    ws_neg.cell(row=bdi_headers_row, column=col).border = thin_border
    ws_neg.cell(row=bdi_headers_row, column=col).alignment = Alignment(horizontal='center')

# Componentes do BDI
bdi_components = [
    ('Administração Central', 5.0, 'Custos de escritório, gerência, RH, etc.'),
    ('Despesas Financeiras', 1.2, 'Juros, taxas bancárias, capital de giro'),
    ('Lucro', 8.0, 'Margem de lucro desejada'),
    ('Impostos', 13.65, 'ISS, PIS, COFINS, etc. (pode variar por município)'),
    ('Riscos e Imprevistos', 2.0, 'Contingências, garantias, seguros'),
    ('Outros', 0.0, 'Outros custos indiretos'),
]

current_bdi_row = bdi_headers_row + 1
for comp_name, comp_value, comp_desc in bdi_components:
    ws_neg.cell(row=current_bdi_row, column=1, value=comp_name).border = thin_border

    cell_value = ws_neg.cell(row=current_bdi_row, column=2, value=comp_value)
    cell_value.border = thin_border
    cell_value.fill = input_fill
    cell_value.number_format = '0.00'
    cell_value.alignment = Alignment(horizontal='center')

    ws_neg.cell(row=current_bdi_row, column=3, value=comp_desc).border = thin_border
    ws_neg.cell(row=current_bdi_row, column=3).font = Font(italic=True, size=9)

    current_bdi_row += 1

# Total BDI
ws_neg.cell(row=current_bdi_row, column=1, value='TOTAL BDI (%)').font = Font(bold=True)
ws_neg.cell(row=current_bdi_row, column=1).border = thin_border
ws_neg.cell(row=current_bdi_row, column=1).fill = PatternFill('solid', fgColor='BDC3C7')

formula_total_bdi = f'=SUM(B{bdi_headers_row + 1}:B{current_bdi_row - 1})'
cell_total_bdi = ws_neg.cell(row=current_bdi_row, column=2, value=formula_total_bdi)
cell_total_bdi.border = thin_border
cell_total_bdi.fill = PatternFill('solid', fgColor='BDC3C7')
cell_total_bdi.number_format = '0.00'
cell_total_bdi.font = Font(bold=True)
cell_total_bdi.alignment = Alignment(horizontal='center')

ws_neg.cell(row=current_bdi_row, column=3, value='Soma de todos os componentes').border = thin_border
ws_neg.cell(row=current_bdi_row, column=3).fill = PatternFill('solid', fgColor='BDC3C7')
ws_neg.cell(row=current_bdi_row, column=3).font = Font(italic=True, size=9)

total_bdi_row = current_bdi_row

# Seção: Markup Simples
markup_section_row = total_bdi_row + 2
ws_neg.merge_cells(f'A{markup_section_row}:F{markup_section_row}')
cell_markup_label = ws_neg.cell(row=markup_section_row, column=1, value='MARKUP SIMPLES (Alternativa ao BDI)')
cell_markup_label.font = section_font
cell_markup_label.fill = section_fill

markup_input_row = markup_section_row + 2
ws_neg.cell(row=markup_input_row, column=1, value='Markup (%):').font = Font(bold=True)
ws_neg.cell(row=markup_input_row, column=2, value=35.0)
ws_neg.cell(row=markup_input_row, column=2).fill = input_fill
ws_neg.cell(row=markup_input_row, column=2).border = thin_border
ws_neg.cell(row=markup_input_row, column=2).number_format = '0.00'
ws_neg.cell(row=markup_input_row, column=2).alignment = Alignment(horizontal='center')
ws_neg.cell(row=markup_input_row, column=3, value='Percentual único sobre o custo direto').font = Font(italic=True, size=9)

# Seção: Multiplicador Final
mult_section_row = markup_input_row + 3
ws_neg.merge_cells(f'A{mult_section_row}:F{mult_section_row}')
cell_mult_label = ws_neg.cell(row=mult_section_row, column=1, value='MULTIPLICADOR FINAL')
cell_mult_label.font = Font(bold=True, size=14, color='FFFFFF')
cell_mult_label.fill = PatternFill('solid', fgColor='27AE60')
cell_mult_label.alignment = Alignment(horizontal='center')

mult_calc_row = mult_section_row + 2
ws_neg.cell(row=mult_calc_row, column=1, value='Multiplicador:').font = Font(bold=True, size=12)

# Fórmula que escolhe entre BDI ou MARKUP baseado no método selecionado
formula_multiplicador = f'=IF(B7="BDI",1+(B{total_bdi_row}/100),IF(B7="MARKUP",1+(B{markup_input_row}/100),1))'
cell_multiplicador = ws_neg.cell(row=mult_calc_row, column=2, value=formula_multiplicador)
cell_multiplicador.font = Font(bold=True, size=14, color='FFFFFF')
cell_multiplicador.fill = PatternFill('solid', fgColor='27AE60')
cell_multiplicador.number_format = '0.0000'
cell_multiplicador.alignment = Alignment(horizontal='center')
cell_multiplicador.border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

ws_neg.cell(row=mult_calc_row, column=3, value='Aplicado no orçamento').font = Font(italic=True, size=10)

# Explicação
expl_row = mult_calc_row + 2
ws_neg.merge_cells(f'A{expl_row}:F{expl_row}')
ws_neg.cell(row=expl_row, column=1, value='COMO FUNCIONA:')
ws_neg.cell(row=expl_row, column=1).font = Font(bold=True, size=11)

expl_content = [
    'BDI: Preço de Venda = Custo Direto × (1 + BDI%/100)',
    '      Exemplo: Custo R$ 1.000,00 com BDI de 30% → Preço = R$ 1.000 × 1,30 = R$ 1.300,00',
    '',
    'MARKUP: Preço de Venda = Custo Direto × (1 + Markup%/100)',
    '        Exemplo: Custo R$ 1.000,00 com Markup de 35% → Preço = R$ 1.000 × 1,35 = R$ 1.350,00',
    '',
    'O multiplicador calculado acima é aplicado automaticamente no orçamento após o subtotal.',
]

expl_content_row = expl_row + 1
for idx, line in enumerate(expl_content):
    ws_neg.merge_cells(f'A{expl_content_row + idx}:F{expl_content_row + idx}')
    ws_neg.cell(row=expl_content_row + idx, column=1, value=line)
    if line.startswith('BDI:') or line.startswith('MARKUP:'):
        ws_neg.cell(row=expl_content_row + idx, column=1).font = Font(bold=True)

# Ajustar larguras
ws_neg.column_dimensions['A'].width = 25
ws_neg.column_dimensions['B'].width = 12
ws_neg.column_dimensions['C'].width = 55
ws_neg.column_dimensions['D'].width = 12
ws_neg.column_dimensions['E'].width = 12
ws_neg.column_dimensions['F'].width = 12

# ============================================
# ABA 4 - MATERIAIS
# ============================================
ws_mat = wb.create_sheet('MATERIAIS')

mat_headers = ['Código', 'Categoria', 'Descrição', 'Unidade', 'Preço (R$)', 'Seleção']
for col, header in enumerate(mat_headers, 1):
    cell = ws_mat.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

materiais = [
    # Tubulação de Cobre
    ('TUB_14_FLEX', 'Tubulação', 'Tubo de cobre 1/4" flexível', 'M', 18.00),
    ('TUB_38_FLEX', 'Tubulação', 'Tubo de cobre 3/8" flexível', 'M', 28.00),
    ('TUB_12_FLEX', 'Tubulação', 'Tubo de cobre 1/2" flexível', 'M', 42.00),
    ('TUB_58_FLEX', 'Tubulação', 'Tubo de cobre 5/8" flexível', 'M', 55.00),
    ('TUB_58_RIG', 'Tubulação', 'Tubo de cobre 5/8" rígido', 'M', 62.00),
    ('TUB_34_FLEX', 'Tubulação', 'Tubo de cobre 3/4" flexível', 'M', 72.00),
    ('TUB_34_RIG', 'Tubulação', 'Tubo de cobre 3/4" rígido', 'M', 80.00),
    ('TUB_78_RIG', 'Tubulação', 'Tubo de cobre 7/8" rígido', 'M', 95.00),
    ('TUB_118_RIG', 'Tubulação', 'Tubo de cobre 1.1/8" rígido', 'M', 135.00),
    ('TUB_138_RIG', 'Tubulação', 'Tubo de cobre 1.3/8" rígido', 'M', 175.00),
    # Isolamento Elastomérico
    ('ISO_14_ELA_E9', 'Isolamento', 'Isolamento elastomérico p/ 1/4" esp. 9mm', 'M', 8.00),
    ('ISO_14_ELA_E13', 'Isolamento', 'Isolamento elastomérico p/ 1/4" esp. 13mm', 'M', 12.00),
    ('ISO_38_ELA_E9', 'Isolamento', 'Isolamento elastomérico p/ 3/8" esp. 9mm', 'M', 10.00),
    ('ISO_38_ELA_E13', 'Isolamento', 'Isolamento elastomérico p/ 3/8" esp. 13mm', 'M', 14.00),
    ('ISO_12_ELA_E9', 'Isolamento', 'Isolamento elastomérico p/ 1/2" esp. 9mm', 'M', 12.00),
    ('ISO_12_ELA_E13', 'Isolamento', 'Isolamento elastomérico p/ 1/2" esp. 13mm', 'M', 16.00),
    ('ISO_58_ELA_E9', 'Isolamento', 'Isolamento elastomérico p/ 5/8" esp. 9mm', 'M', 14.00),
    ('ISO_58_ELA_E19', 'Isolamento', 'Isolamento elastomérico p/ 5/8" esp. 19mm', 'M', 22.00),
    ('ISO_34_ELA_E9', 'Isolamento', 'Isolamento elastomérico p/ 3/4" esp. 9mm', 'M', 16.00),
    ('ISO_34_ELA_E19', 'Isolamento', 'Isolamento elastomérico p/ 3/4" esp. 19mm', 'M', 25.00),
    ('ISO_78_ELA_E19', 'Isolamento', 'Isolamento elastomérico p/ 7/8" esp. 19mm', 'M', 28.00),
    ('ISO_118_ELA_E25', 'Isolamento', 'Isolamento elastomérico p/ 1.1/8" esp. 25mm', 'M', 38.00),
    ('ISO_138_ELA_E25', 'Isolamento', 'Isolamento elastomérico p/ 1.3/8" esp. 25mm', 'M', 48.00),
    # Isolamento Polietileno
    ('ISO_14_POL_E10', 'Isolamento', 'Isolamento polietileno p/ 1/4" esp. 10mm', 'M', 6.00),
    ('ISO_38_POL_E10', 'Isolamento', 'Isolamento polietileno p/ 3/8" esp. 10mm', 'M', 7.00),
    ('ISO_12_POL_E10', 'Isolamento', 'Isolamento polietileno p/ 1/2" esp. 10mm', 'M', 8.00),
    ('ISO_58_POL_E10', 'Isolamento', 'Isolamento polietileno p/ 5/8" esp. 10mm', 'M', 9.00),
    ('ISO_34_POL_E10', 'Isolamento', 'Isolamento polietileno p/ 3/4" esp. 10mm', 'M', 10.00),
    ('ACA_FITA_TER', 'Isolamento', 'Fita isolante térmica (rolo 30m)', 'UN', 25.00),
    # Cabos Elétricos
    ('CAB_PP_15', 'Elétrica', 'Cabo PP 3x1,5mm²', 'M', 6.50),
    ('CAB_PP_25', 'Elétrica', 'Cabo PP 3x2,5mm²', 'M', 9.00),
    ('CAB_PP_40', 'Elétrica', 'Cabo PP 3x4mm²', 'M', 14.00),
    ('CAB_PP_60', 'Elétrica', 'Cabo PP 3x6mm²', 'M', 22.00),
    ('CAB_FLEX_25', 'Elétrica', 'Cabo flexível 2,5mm² (fase/neutro/terra)', 'M', 3.50),
    ('CAB_FLEX_40', 'Elétrica', 'Cabo flexível 4mm² (fase/neutro/terra)', 'M', 5.50),
    ('CAB_FLEX_60', 'Elétrica', 'Cabo flexível 6mm² (fase/neutro/terra)', 'M', 8.00),
    ('CAB_COM', 'Elétrica', 'Cabo de comunicação 2x0,75mm²', 'M', 3.00),
    # Proteção Elétrica
    ('DISJ_M_10', 'Proteção', 'Disjuntor monopolar 10A', 'UN', 18.00),
    ('DISJ_M_16', 'Proteção', 'Disjuntor monopolar 16A', 'UN', 18.00),
    ('DISJ_M_20', 'Proteção', 'Disjuntor monopolar 20A', 'UN', 18.00),
    ('DISJ_B_20', 'Proteção', 'Disjuntor bipolar 20A', 'UN', 45.00),
    ('DISJ_B_25', 'Proteção', 'Disjuntor bipolar 25A', 'UN', 48.00),
    ('DISJ_B_32', 'Proteção', 'Disjuntor bipolar 32A', 'UN', 52.00),
    ('DISJ_B_40', 'Proteção', 'Disjuntor bipolar 40A', 'UN', 58.00),
    ('DISJ_CX', 'Proteção', 'Caixa de sobrepor para disjuntor', 'UN', 25.00),
    # Eletrodutos
    ('COND_CORR_34', 'Eletroduto', 'Eletroduto corrugado 3/4"', 'M', 2.50),
    ('COND_CORR_100', 'Eletroduto', 'Eletroduto corrugado 1"', 'M', 3.50),
    ('COND_CURV_34', 'Eletroduto', 'Curva 90° eletroduto 3/4"', 'UN', 1.50),
    ('COND_ABRC_34', 'Eletroduto', 'Abraçadeira para eletroduto 3/4"', 'UN', 0.80),
    # Dreno
    ('DRN_CRIS_34', 'Dreno', 'Mangueira cristal 3/4"', 'M', 4.00),
    ('DRN_CRIS_100', 'Dreno', 'Mangueira cristal 1"', 'M', 6.00),
    ('DRN_PVC_25', 'Dreno', 'Tubo PVC esgoto 25mm', 'M', 8.00),
    ('DRN_PVC_32', 'Dreno', 'Tubo PVC esgoto 32mm', 'M', 10.00),
    ('DRN_CURV_25', 'Dreno', 'Curva 90° PVC 25mm', 'UN', 3.50),
    ('DRN_LUVA_25', 'Dreno', 'Luva PVC 25mm', 'UN', 2.00),
    ('DRN_SIFAO', 'Dreno', 'Sifão sanfonado universal', 'UN', 15.00),
    ('DRN_COLA', 'Dreno', 'Cola PVC (tubo 75g)', 'UN', 12.00),
    ('DRN_ABRC', 'Dreno', 'Abraçadeira nylon 3/4"', 'UN', 0.50),
    # Suportes
    ('SUP_MF_400', 'Suporte', 'Suporte mão francesa 400mm (par)', 'PAR', 85.00),
    ('SUP_MF_500', 'Suporte', 'Suporte mão francesa 500mm (par)', 'PAR', 110.00),
    ('SUP_MF_600', 'Suporte', 'Suporte mão francesa 600mm (par)', 'PAR', 140.00),
    ('SUP_CALCO', 'Suporte', 'Calço borracha antivibração (jg 4pç)', 'JG', 35.00),
    ('SUP_EVAP', 'Suporte', 'Suporte evaporadora universal', 'UN', 45.00),
    ('SUP_PARF_100', 'Suporte', 'Parafuso sextavado 3/8"x100mm c/ bucha', 'UN', 4.50),
    ('SUP_PARF_150', 'Suporte', 'Parafuso sextavado 3/8"x150mm c/ bucha', 'UN', 6.00),
    ('SUP_BUCHA', 'Suporte', 'Bucha nylon S10', 'UN', 0.80),
    ('SUP_CHIP', 'Suporte', 'Parafuso chipboard 6x60mm', 'UN', 0.50),
    # Acabamento
    ('ACA_CAN_50', 'Acabamento', 'Canaleta PVC 50x50mm', 'M', 18.00),
    ('ACA_CAN_70', 'Acabamento', 'Canaleta PVC 70x70mm', 'M', 25.00),
    ('ACA_CAN_100', 'Acabamento', 'Canaleta PVC 100x50mm', 'M', 32.00),
    ('ACA_CURV_50', 'Acabamento', 'Curva 90° canaleta 50mm', 'UN', 12.00),
    ('ACA_TAMP_50', 'Acabamento', 'Tampa canaleta 50mm', 'M', 8.00),
    ('ACA_ESPUMA', 'Acabamento', 'Espuma expansiva 500ml', 'UN', 28.00),
    ('ACA_MASSA', 'Acabamento', 'Massa de calafetar 400g', 'UN', 18.00),
    ('ACA_FITA_ISO', 'Acabamento', 'Fita isolante preta 20m', 'UN', 8.00),
    ('ACA_FITA_AUT', 'Acabamento', 'Fita autofusão 10m', 'UN', 25.00),
    # Gás e Solda
    ('GAS_R410A', 'Gás/Solda', 'Gás refrigerante R-410A', 'KG', 95.00),
    ('GAS_R22', 'Gás/Solda', 'Gás refrigerante R-22', 'KG', 75.00),
    ('GAS_R32', 'Gás/Solda', 'Gás refrigerante R-32', 'KG', 85.00),
    ('GAS_N2', 'Gás/Solda', 'Nitrogênio (carga teste)', 'UN', 35.00),
    ('SOL_PRATA', 'Gás/Solda', 'Vareta solda prata 5%', 'UN', 18.00),
    ('SOL_FLUXO', 'Gás/Solda', 'Fluxo para solda 100g', 'UN', 22.00),
    # Alvenaria
    ('ALV_ARG', 'Alvenaria', 'Argamassa pronta 20kg', 'SC', 28.00),
    ('ALV_TIJ', 'Alvenaria', 'Tijolo cerâmico 6 furos', 'UN', 1.20),
    ('ALV_GESSO', 'Alvenaria', 'Gesso cola 5kg', 'SC', 18.00),
]

for row_idx, mat in enumerate(materiais, 2):
    for col_idx, value in enumerate(mat, 1):
        cell = ws_mat.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 1:
            cell.font = code_font
            cell.alignment = Alignment(horizontal='left')
        elif col_idx == 2 or col_idx == 4:
            cell.alignment = Alignment(horizontal='center')
        elif col_idx == 5:
            cell.number_format = currency_format

        cat = mat[1]
        if cat in cat_colors:
            cell.fill = PatternFill('solid', fgColor=cat_colors[cat])

    # Coluna F: Seleção (fórmula dinâmica CÓDIGO - DESCRIÇÃO)
    cell_sel = ws_mat.cell(row=row_idx, column=6, value=f'=A{row_idx}&" - "&C{row_idx}')
    cell_sel.border = thin_border

ws_mat.column_dimensions['A'].width = 18
ws_mat.column_dimensions['B'].width = 12
ws_mat.column_dimensions['C'].width = 45
ws_mat.column_dimensions['D'].width = 10
ws_mat.column_dimensions['E'].width = 14
ws_mat.column_dimensions['F'].width = 55
ws_mat.column_dimensions['F'].hidden = True  # Ocultar coluna auxiliar

# ============================================
# ABA 4 - MÃO DE OBRA
# ============================================
ws_mo = wb.create_sheet('MAO_DE_OBRA')

mo_headers = ['Código', 'Categoria', 'Descrição', 'Unidade', 'Custo (R$)', 'Seleção']
for col, header in enumerate(mo_headers, 1):
    cell = ws_mo.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

mao_de_obra = [
    ('MO_TEC', 'Instalação', 'Técnico em refrigeração', 'H', 65.00),
    ('MO_AJU', 'Instalação', 'Ajudante de instalação', 'H', 35.00),
    ('MO_ELE', 'Elétrica', 'Eletricista', 'H', 55.00),
    ('MO_PED', 'Civil', 'Pedreiro', 'H', 50.00),
    ('MO_SERV', 'Civil', 'Servente de pedreiro', 'H', 30.00),
    ('MO_ALT', 'Adicional', 'Adicional trabalho em altura (>3m)', 'H', 25.00),
    ('MO_FAC', 'Adicional', 'Adicional trabalho fachada', 'H', 35.00),
    ('MO_DESL_20', 'Deslocamento', 'Deslocamento equipe (até 20km)', 'VZ', 80.00),
    ('MO_DESL_50', 'Deslocamento', 'Deslocamento equipe (20-50km)', 'VZ', 150.00),
]

for row_idx, mo in enumerate(mao_de_obra, 2):
    for col_idx, value in enumerate(mo, 1):
        cell = ws_mo.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 1:
            cell.font = code_font
        elif col_idx == 2 or col_idx == 4:
            cell.alignment = Alignment(horizontal='center')
        elif col_idx == 5:
            cell.number_format = currency_format

    # Coluna F: Seleção (fórmula dinâmica CÓDIGO - DESCRIÇÃO)
    cell_sel = ws_mo.cell(row=row_idx, column=6, value=f'=A{row_idx}&" - "&C{row_idx}')
    cell_sel.border = thin_border

ws_mo.column_dimensions['A'].width = 14
ws_mo.column_dimensions['B'].width = 14
ws_mo.column_dimensions['C'].width = 40
ws_mo.column_dimensions['D'].width = 10
ws_mo.column_dimensions['E'].width = 14
ws_mo.column_dimensions['F'].width = 50
ws_mo.column_dimensions['F'].hidden = True  # Ocultar coluna auxiliar

# ============================================
# ABA 5 - FERRAMENTAS
# ============================================
ws_fer = wb.create_sheet('FERRAMENTAS')

fer_headers = ['Código', 'Categoria', 'Descrição', 'Valor Aquisição (R$)', 'Vida Útil (H)', 'Custo/Hora (R$)', 'Seleção']
for col, header in enumerate(fer_headers, 1):
    cell = ws_fer.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ferramentas = [
    ('FER_VACUO', 'Vácuo', 'Bomba de vácuo', 1500.00, 2000),
    ('FER_MANIF', 'Manifold', 'Manifold digital', 800.00, 3000),
    ('FER_SOLDA', 'Solda', 'Kit solda oxigênio/GLP', 1200.00, 1500),
    ('FER_PERF', 'Furação', 'Perfuratriz/martelete', 1800.00, 2500),
    ('FER_SERRA_65', 'Furação', 'Serra copo diamantada 65mm', 180.00, 150),
    ('FER_SERRA_80', 'Furação', 'Serra copo diamantada 80mm', 220.00, 150),
    ('FER_ESCADA', 'Acesso', 'Escada extensível 6m', 800.00, 3000),
    ('FER_ANDAIME', 'Acesso', 'Andaime cavalete (par)', 600.00, 2000),
    ('FER_ESTANQ', 'Teste', 'Kit teste estanqueidade', 350.00, 2000),
    ('FER_MULT', 'Elétrica', 'Multímetro/alicate amperímetro', 450.00, 3000),
    ('FER_MANUAL', 'Diversos', 'Ferramentas manuais (conjunto)', 1200.00, 4000),
    ('FER_FURAD', 'Furação', 'Furadeira de impacto', 650.00, 2000),
    ('FER_CORTA', 'Corte', 'Cortador de tubo cobre', 120.00, 1000),
    ('FER_FLANG', 'Dobra', 'Flangeador/alargador', 380.00, 2000),
    ('FER_BAL', 'Refrigeração', 'Balança digital refrigeração', 280.00, 2500),
]

for row_idx, fer in enumerate(ferramentas, 2):
    ws_fer.cell(row=row_idx, column=1, value=fer[0]).font = code_font
    ws_fer.cell(row=row_idx, column=1).border = thin_border
    
    ws_fer.cell(row=row_idx, column=2, value=fer[1]).border = thin_border
    ws_fer.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
    
    ws_fer.cell(row=row_idx, column=3, value=fer[2]).border = thin_border
    
    cell_valor = ws_fer.cell(row=row_idx, column=4, value=fer[3])
    cell_valor.border = thin_border
    cell_valor.number_format = currency_format
    
    cell_vida = ws_fer.cell(row=row_idx, column=5, value=fer[4])
    cell_vida.border = thin_border
    cell_vida.number_format = hours_format
    cell_vida.alignment = Alignment(horizontal='center')
    
    formula_custo = f'=IF(E{row_idx}>0,D{row_idx}/E{row_idx},0)'
    cell_custo_h = ws_fer.cell(row=row_idx, column=6, value=formula_custo)
    cell_custo_h.border = thin_border
    cell_custo_h.number_format = currency_format

    # Coluna G: Seleção (fórmula dinâmica CÓDIGO - DESCRIÇÃO)
    cell_sel = ws_fer.cell(row=row_idx, column=7, value=f'=A{row_idx}&" - "&C{row_idx}')
    cell_sel.border = thin_border

ws_fer.column_dimensions['A'].width = 16
ws_fer.column_dimensions['B'].width = 12
ws_fer.column_dimensions['C'].width = 35
ws_fer.column_dimensions['D'].width = 18
ws_fer.column_dimensions['E'].width = 14
ws_fer.column_dimensions['F'].width = 16
ws_fer.column_dimensions['G'].width = 45
ws_fer.column_dimensions['G'].hidden = True  # Ocultar coluna auxiliar

# ============================================
# ABA 6 - EQUIPAMENTOS
# ============================================
ws_eqp = wb.create_sheet('EQUIPAMENTOS')

eqp_headers = ['Código', 'Categoria', 'Descrição', 'Capacidade (BTU)', 'Unidade', 'Preço (R$)', 'Seleção']
for col, header in enumerate(eqp_headers, 1):
    cell = ws_eqp.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

equipamentos = [
    ('EQP_HW_9K', 'Split Hi-Wall', 'Split Hi-Wall Inverter 9.000 BTUs', 9000, 'UN', 1800.00),
    ('EQP_HW_12K', 'Split Hi-Wall', 'Split Hi-Wall Inverter 12.000 BTUs', 12000, 'UN', 2200.00),
    ('EQP_HW_18K', 'Split Hi-Wall', 'Split Hi-Wall Inverter 18.000 BTUs', 18000, 'UN', 3200.00),
    ('EQP_HW_22K', 'Split Hi-Wall', 'Split Hi-Wall Inverter 22.000 BTUs', 22000, 'UN', 3800.00),
    ('EQP_HW_24K', 'Split Hi-Wall', 'Split Hi-Wall Inverter 24.000 BTUs', 24000, 'UN', 4200.00),
    ('EQP_HW_30K', 'Split Hi-Wall', 'Split Hi-Wall Inverter 30.000 BTUs', 30000, 'UN', 5500.00),
    ('EQP_PT_36K', 'Split Piso-Teto', 'Split Piso-Teto 36.000 BTUs', 36000, 'UN', 7200.00),
    ('EQP_PT_48K', 'Split Piso-Teto', 'Split Piso-Teto 48.000 BTUs', 48000, 'UN', 9500.00),
    ('EQP_PT_60K', 'Split Piso-Teto', 'Split Piso-Teto 60.000 BTUs', 60000, 'UN', 12000.00),
    ('EQP_BOMB_P', 'Bomba Dreno', 'Bomba dreno mini (até 12.000 BTUs)', 0, 'UN', 320.00),
    ('EQP_BOMB_M', 'Bomba Dreno', 'Bomba dreno média (até 36.000 BTUs)', 0, 'UN', 480.00),
    ('EQP_BOMB_G', 'Bomba Dreno', 'Bomba dreno grande (até 60.000 BTUs)', 0, 'UN', 650.00),
]

for row_idx, eqp in enumerate(equipamentos, 2):
    ws_eqp.cell(row=row_idx, column=1, value=eqp[0]).font = code_font
    ws_eqp.cell(row=row_idx, column=1).border = thin_border
    
    ws_eqp.cell(row=row_idx, column=2, value=eqp[1]).border = thin_border
    ws_eqp.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
    
    ws_eqp.cell(row=row_idx, column=3, value=eqp[2]).border = thin_border
    
    cell_cap = ws_eqp.cell(row=row_idx, column=4, value=eqp[3] if eqp[3] > 0 else '-')
    cell_cap.border = thin_border
    cell_cap.alignment = Alignment(horizontal='center')
    if eqp[3] > 0:
        cell_cap.number_format = '#,##0'
    
    ws_eqp.cell(row=row_idx, column=5, value=eqp[4]).border = thin_border
    ws_eqp.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')
    
    cell_preco = ws_eqp.cell(row=row_idx, column=6, value=eqp[5])
    cell_preco.border = thin_border
    cell_preco.number_format = currency_format

    # Coluna G: Seleção (fórmula dinâmica CÓDIGO - DESCRIÇÃO)
    cell_sel = ws_eqp.cell(row=row_idx, column=7, value=f'=A{row_idx}&" - "&C{row_idx}')
    cell_sel.border = thin_border

ws_eqp.column_dimensions['A'].width = 16
ws_eqp.column_dimensions['B'].width = 14
ws_eqp.column_dimensions['C'].width = 40
ws_eqp.column_dimensions['D'].width = 16
ws_eqp.column_dimensions['E'].width = 10
ws_eqp.column_dimensions['F'].width = 14
ws_eqp.column_dimensions['G'].width = 50
ws_eqp.column_dimensions['G'].hidden = True  # Ocultar coluna auxiliar

# ============================================
# ABA 7 - COMPOSIÇÕES
# ============================================
ws_comp = wb.create_sheet('COMPOSICOES')

# Headers principais (A-L) - dados das composições
comp_headers = ['Código', 'Descrição', 'Tipo', 'Cód. Item', 'Descrição Item', 'Un', 'Qtd Base', 'Qtd/Metro', 'Preço Unit.', 'Sub. Base', 'Sub./Metro', 'Seleção']
for col, header in enumerate(comp_headers, 1):
    cell = ws_comp.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Headers de lookup (M-S) - tabela auxiliar para VLOOKUP no ORCAMENTO
lookup_headers = ['LK_CODIGO', 'LK_SELECAO', 'LK_TOTAL_BASE', 'LK_TOTAL_METRO', 'LK_DESCRICAO', 'LK_DESC_PRE', 'LK_DESC_POS']
for col, header in enumerate(lookup_headers, 13):  # Começa na coluna M (13)
    cell = ws_comp.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

composicoes = [
    {
        'codigo': 'COMP_INST_9K',
        'descricao': 'Instalação Split 9.000 BTUs',
        'desc_pre': 'Instalação Split 9.000 BTUs com ',
        'desc_pos': ' de linha frigorígena',
        'itens': [
            # (Tipo, Código, Qtd_Base, Qtd_por_Metro)
            ('MAT', 'TUB_14_FLEX', 0, 1.1),
            ('MAT', 'TUB_38_FLEX', 0, 1.1),
            ('MAT', 'ISO_14_ELA_E9', 0, 1.1),
            ('MAT', 'ISO_38_ELA_E9', 0, 1.1),
            ('MAT', 'ACA_FITA_TER', 0.15, 0),
            ('MAT', 'CAB_COM', 0, 1.15),
            ('MAT', 'SUP_EVAP', 1, 0),
            ('MAT', 'SUP_BUCHA', 6, 0),
            ('MAT', 'SUP_CHIP', 6, 0),
            ('MAT', 'ACA_ESPUMA', 0.3, 0),
            ('MAT', 'ACA_FITA_ISO', 0.25, 0),
            ('MAT', 'ACA_FITA_AUT', 0.2, 0),
            ('MAT', 'SOL_PRATA', 2, 0),
            ('MAT', 'SOL_FLUXO', 0.2, 0),
            ('MO', 'MO_TEC', 2, 0.15),
            ('MO', 'MO_AJU', 2, 0.15),
            ('FER', 'FER_VACUO', 0.5, 0),
            ('FER', 'FER_MANIF', 0.5, 0),
            ('FER', 'FER_SOLDA', 0.5, 0),
            ('FER', 'FER_MANUAL', 2.5, 0),
        ]
    },
    {
        'codigo': 'COMP_SUP_MF',
        'descricao': 'Adicional suporte mão francesa 400mm (condensadora)',
        'itens': [
            ('MAT', 'SUP_MF_400', 1, 0),
            ('MAT', 'SUP_PARF_100', 4, 0),
            ('MO', 'MO_TEC', 0.75, 0),
            ('MO', 'MO_AJU', 0.75, 0),
            ('FER', 'FER_PERF', 0.5, 0),
        ]
    },
    {
        'codigo': 'COMP_SUP_CALCO',
        'descricao': 'Adicional calços de borracha (condensadora)',
        'itens': [
            ('MAT', 'SUP_CALCO', 1, 0),
            ('MO', 'MO_TEC', 0.25, 0),
        ]
    },
    {
        'codigo': 'COMP_ELE',
        'descricao': 'Alimentação elétrica 220V mono',
        'desc_pre': 'Alimentação elétrica 220V mono com ',
        'desc_pos': ' de cabo e eletroduto',
        'itens': [
            ('MAT', 'CAB_PP_25', 0, 1.1),
            ('MAT', 'COND_CORR_34', 0, 1.1),
            ('MAT', 'COND_CURV_34', 2, 0),
            ('MAT', 'COND_ABRC_34', 0, 1.2),
            ('MAT', 'ACA_FITA_ISO', 0.25, 0),
            ('MO', 'MO_ELE', 1, 0.15),
            ('FER', 'FER_MULT', 0.5, 0),
        ]
    },
    {
        'codigo': 'COMP_DRN_CRIS',
        'descricao': 'Dreno mangueira cristal',
        'desc_pre': 'Dreno mangueira cristal com ',
        'desc_pos': '',
        'itens': [
            ('MAT', 'DRN_CRIS_34', 0, 1.1),
            ('MAT', 'DRN_ABRC', 0, 1.3),
            ('MO', 'MO_TEC', 0.25, 0.1),
        ]
    },
    {
        'codigo': 'COMP_DRN_PVC',
        'descricao': 'Dreno tubo PVC 25mm',
        'desc_pre': 'Dreno tubo PVC 25mm com ',
        'desc_pos': '',
        'itens': [
            ('MAT', 'DRN_PVC_25', 0, 1.1),
            ('MAT', 'DRN_CURV_25', 2, 0),
            ('MAT', 'DRN_LUVA_25', 0, 0.5),
            ('MAT', 'DRN_COLA', 0.3, 0),
            ('MO', 'MO_TEC', 0.5, 0.15),
        ]
    },
    {
        'codigo': 'COMP_FURO',
        'descricao': 'Furo em parede/laje/viga (até 20cm espessura)',
        'itens': [
            ('MAT', 'ACA_ESPUMA', 0.2, 0),
            ('MAT', 'ACA_MASSA', 0.3, 0),
            ('MO', 'MO_TEC', 0.75, 0),
            ('MO', 'MO_AJU', 0.75, 0),
            ('FER', 'FER_PERF', 0.5, 0),
            ('FER', 'FER_SERRA_65', 1, 0),
        ]
    },
    {
        'codigo': 'COMP_CAN_50',
        'descricao': 'Acabamento canaleta PVC 50mm',
        'desc_pre': 'Acabamento canaleta PVC 50mm com ',
        'desc_pos': '',
        'itens': [
            ('MAT', 'ACA_CAN_50', 0, 1.1),
            ('MAT', 'SUP_BUCHA', 0, 2),
            ('MAT', 'SUP_CHIP', 0, 2),
            ('MO', 'MO_TEC', 0, 0.25),
        ]
    },
    {
        'codigo': 'COMP_ALV',
        'descricao': 'Abertura e fechamento alvenaria',
        'desc_pre': 'Abertura e fechamento alvenaria com ',
        'desc_pos': ' para embutir linha',
        'itens': [
            ('MAT', 'ALV_ARG', 0, 0.5),
            ('MAT', 'ALV_TIJ', 0, 10),
            ('MO', 'MO_PED', 1, 1.25),
            ('MO', 'MO_SERV', 1, 1.25),
            ('FER', 'FER_MANUAL', 1, 1),
        ]
    },
    {
        'codigo': 'COMP_FACH',
        'descricao': 'Instalação condensadora em fachada (com suporte)',
        'itens': [
            ('MAT', 'SUP_MF_500', 1, 0),
            ('MAT', 'SUP_PARF_150', 4, 0),
            ('MO', 'MO_TEC', 2, 0),
            ('MO', 'MO_AJU', 2, 0),
            ('MO', 'MO_FAC', 2, 0),
            ('FER', 'FER_PERF', 1, 0),
            ('FER', 'FER_MANUAL', 2, 0),
        ]
    },
    {
        'codigo': 'COMP_ALT',
        'descricao': 'Adicional trabalho em altura (acima 3m)',
        'itens': [
            ('MO', 'MO_ALT', 2, 0),
            ('FER', 'FER_ESCADA', 2, 0),
        ]
    },
    {
        'codigo': 'COMP_BOMB_DRN',
        'descricao': 'Instalação bomba de dreno',
        'itens': [
            ('EQP', 'EQP_BOMB_P', 1, 0),
            ('MAT', 'DRN_CRIS_34', 2, 0),
            ('MAT', 'ACA_FITA_ISO', 0.1, 0),
            ('MO', 'MO_TEC', 1, 0),
        ]
    },
    {
        'codigo': 'COMP_DISJ',
        'descricao': 'Instalação disjuntor no QDC',
        'itens': [
            ('MAT', 'DISJ_M_16', 1, 0),
            ('MAT', 'CAB_FLEX_25', 3, 0),
            ('MAT', 'ACA_FITA_ISO', 0.15, 0),
            ('MO', 'MO_ELE', 1, 0),
            ('FER', 'FER_MULT', 0.25, 0),
        ]
    },
    {
        'codigo': 'COMP_DRN_ESG',
        'descricao': 'Conexão dreno em rede esgoto (com sifão)',
        'itens': [
            ('MAT', 'DRN_PVC_25', 1, 0),
            ('MAT', 'DRN_SIFAO', 1, 0),
            ('MAT', 'DRN_CURV_25', 1, 0),
            ('MAT', 'DRN_COLA', 0.2, 0),
            ('MO', 'MO_TEC', 0.75, 0),
        ]
    },
    {
        'codigo': 'COMP_DESINST',
        'descricao': 'Desinstalação de equipamento Split',
        'itens': [
            ('MO', 'MO_TEC', 1.5, 0),
            ('MO', 'MO_AJU', 1.5, 0),
            ('FER', 'FER_MANIF', 0.5, 0),
            ('FER', 'FER_MANUAL', 1.5, 0),
        ]
    },
    {
        'codigo': 'COMP_GAS_ADIC',
        'descricao': 'Carga adicional gás R-410A (por kg)',
        'itens': [
            ('MAT', 'GAS_R410A', 1, 0),
            ('MO', 'MO_TEC', 0.5, 0),
            ('FER', 'FER_MANIF', 0.25, 0),
            ('FER', 'FER_BAL', 0.25, 0),
        ]
    },
]

current_row = 2

# Dicionário para armazenar linha e range de cada composição (para lookup no orçamento)
comp_totals_info = []

for comp in composicoes:
    start_row = current_row
    num_itens = len(comp['itens'])
    header_row_comp = current_row

    # Linha de cabeçalho da composição
    cell_codigo = ws_comp.cell(row=current_row, column=1, value=comp['codigo'])
    cell_codigo.font = Font(bold=True, color='FFFFFF', name='Consolas')
    cell_codigo.fill = comp_header_fill
    cell_codigo.alignment = Alignment(horizontal='left', vertical='center')
    cell_codigo.border = thin_border

    cell_desc = ws_comp.cell(row=current_row, column=2, value=comp['descricao'])
    cell_desc.font = Font(bold=True, color='FFFFFF')
    cell_desc.fill = comp_header_fill
    cell_desc.alignment = Alignment(horizontal='left', vertical='center')
    cell_desc.border = thin_border

    # Colunas C até I vazias no cabeçalho
    for col in range(3, 10):
        cell = ws_comp.cell(row=current_row, column=col, value='')
        cell.fill = comp_header_fill
        cell.border = thin_border

    # Coluna L: Seleção (fórmula dinâmica CÓDIGO - DESCRIÇÃO)
    cell_sel = ws_comp.cell(row=current_row, column=12, value=f'=A{current_row}&" - "&B{current_row}')
    cell_sel.fill = comp_header_fill
    cell_sel.border = thin_border

    # TOTAL da composição na coluna J (será preenchido após os itens)
    # Placeholder - a fórmula será adicionada depois de saber o range dos itens

    current_row += 1
    first_item_row = current_row

    # Itens da composição
    for item in comp['itens']:
        tipo, codigo, qtd_base, qtd_metro = item

        # Colunas A e B vazias para itens
        ws_comp.cell(row=current_row, column=1, value='').border = thin_border
        ws_comp.cell(row=current_row, column=2, value='').border = thin_border

        # Tipo (coluna C)
        cell_tipo = ws_comp.cell(row=current_row, column=3, value=tipo)
        cell_tipo.alignment = Alignment(horizontal='center')
        cell_tipo.border = thin_border
        cell_tipo.fill = input_fill

        # Código Item (coluna D) - será dropdown híbrido
        cell_cod = ws_comp.cell(row=current_row, column=4, value=codigo)
        cell_cod.font = code_font
        cell_cod.alignment = Alignment(horizontal='left')
        cell_cod.border = thin_border
        cell_cod.fill = input_fill

        # Fórmulas dinâmicas baseadas no TIPO (coluna C)
        # Extrai código: se tem " - " usa LEFT(), senão usa o valor direto
        # IFERROR(LEFT(D,FIND(" - ",D)-1),D) = extrai código ou usa valor direto
        extract_code = f'IFERROR(LEFT(D{current_row},FIND(" - ",D{current_row})-1),D{current_row})'

        # Descrição Item (coluna E)
        formula_desc = f'''=IF(C{current_row}="","",IF(C{current_row}="MAT",VLOOKUP({extract_code},MATERIAIS!$A:$E,3,FALSE),IF(C{current_row}="MO",VLOOKUP({extract_code},MAO_DE_OBRA!$A:$E,3,FALSE),IF(C{current_row}="FER",VLOOKUP({extract_code},FERRAMENTAS!$A:$F,3,FALSE),IF(C{current_row}="EQP",VLOOKUP({extract_code},EQUIPAMENTOS!$A:$F,3,FALSE),"")))))'''
        cell_descr = ws_comp.cell(row=current_row, column=5, value=formula_desc)
        cell_descr.border = thin_border

        # Unidade (coluna F)
        formula_un = f'''=IF(C{current_row}="","",IF(C{current_row}="MAT",VLOOKUP({extract_code},MATERIAIS!$A:$E,4,FALSE),IF(C{current_row}="MO",VLOOKUP({extract_code},MAO_DE_OBRA!$A:$E,4,FALSE),IF(C{current_row}="FER","H",IF(C{current_row}="EQP",VLOOKUP({extract_code},EQUIPAMENTOS!$A:$F,5,FALSE),"")))))'''
        cell_un = ws_comp.cell(row=current_row, column=6, value=formula_un)
        cell_un.alignment = Alignment(horizontal='center')
        cell_un.border = thin_border

        # Qtd Base (coluna G)
        cell_qtd_base = ws_comp.cell(row=current_row, column=7, value=qtd_base if qtd_base > 0 else '')
        cell_qtd_base.number_format = qty_format
        cell_qtd_base.alignment = Alignment(horizontal='center')
        cell_qtd_base.border = thin_border
        cell_qtd_base.fill = input_fill

        # Qtd/Metro (coluna H)
        cell_qtd_metro = ws_comp.cell(row=current_row, column=8, value=qtd_metro if qtd_metro > 0 else '')
        cell_qtd_metro.number_format = qty_format
        cell_qtd_metro.alignment = Alignment(horizontal='center')
        cell_qtd_metro.border = thin_border
        cell_qtd_metro.fill = input_fill

        # Preço Unitário (coluna I)
        formula_preco = f'''=IF(C{current_row}="","",IF(C{current_row}="MAT",VLOOKUP({extract_code},MATERIAIS!$A:$E,5,FALSE),IF(C{current_row}="MO",VLOOKUP({extract_code},MAO_DE_OBRA!$A:$E,5,FALSE),IF(C{current_row}="FER",VLOOKUP({extract_code},FERRAMENTAS!$A:$F,6,FALSE),IF(C{current_row}="EQP",VLOOKUP({extract_code},EQUIPAMENTOS!$A:$F,6,FALSE),"")))))'''
        cell_preco = ws_comp.cell(row=current_row, column=9, value=formula_preco)
        cell_preco.number_format = currency_format
        cell_preco.border = thin_border

        # Subtotal Base (coluna J) = Qtd Base * Preço Unitário
        formula_subtotal_base = f'=IF(OR(G{current_row}="",I{current_row}=""),"",G{current_row}*I{current_row})'
        cell_subtotal_base = ws_comp.cell(row=current_row, column=10, value=formula_subtotal_base)
        cell_subtotal_base.number_format = currency_format
        cell_subtotal_base.border = thin_border

        # Subtotal/Metro (coluna K) = Qtd/Metro * Preço Unitário
        formula_subtotal_metro = f'=IF(OR(H{current_row}="",I{current_row}=""),"",H{current_row}*I{current_row})'
        cell_subtotal_metro = ws_comp.cell(row=current_row, column=11, value=formula_subtotal_metro)
        cell_subtotal_metro.number_format = currency_format
        cell_subtotal_metro.border = thin_border

        # Aplicar cores por tipo
        if tipo == 'MO':
            for col in range(1, 12):
                ws_comp.cell(row=current_row, column=col).fill = mo_fill
        elif tipo == 'FER':
            for col in range(1, 12):
                ws_comp.cell(row=current_row, column=col).fill = fer_fill
        elif tipo == 'EQP':
            for col in range(1, 12):
                ws_comp.cell(row=current_row, column=col).fill = eqp_fill

        current_row += 1

    last_item_row = current_row - 1

    # Agora adicionar os TOTAIS na linha de cabeçalho
    # Total Base (coluna J)
    formula_total_base = f'=SUM(J{first_item_row}:J{last_item_row})'
    cell_total_base = ws_comp.cell(row=header_row_comp, column=10, value=formula_total_base)
    cell_total_base.font = Font(bold=True, color='FFFFFF')
    cell_total_base.fill = comp_header_fill
    cell_total_base.number_format = currency_format
    cell_total_base.border = thin_border
    cell_total_base.alignment = Alignment(horizontal='right')

    # Total/Metro (coluna K)
    formula_total_metro = f'=SUM(K{first_item_row}:K{last_item_row})'
    cell_total_metro = ws_comp.cell(row=header_row_comp, column=11, value=formula_total_metro)
    cell_total_metro.font = Font(bold=True, color='FFFFFF')
    cell_total_metro.fill = comp_header_fill
    cell_total_metro.number_format = currency_format
    cell_total_metro.border = thin_border
    cell_total_metro.alignment = Alignment(horizontal='right')

    # Guardar info para referência no orçamento (incluindo descrição dinâmica)
    comp_totals_info.append({
        'codigo': comp['codigo'],
        'descricao': comp['descricao'],
        'desc_pre': comp.get('desc_pre', ''),
        'desc_pos': comp.get('desc_pos', ''),
        'header_row': header_row_comp,
        'first_item': first_item_row,
        'last_item': last_item_row
    })

    current_row += 1  # Linha em branco

# Preencher tabela de lookup (colunas M-S) - dados contíguos para VLOOKUP
for idx, comp_info in enumerate(comp_totals_info, 2):
    # M: Código
    ws_comp.cell(row=idx, column=13, value=comp_info['codigo'])
    # N: Seleção (CÓDIGO - DESCRIÇÃO)
    ws_comp.cell(row=idx, column=14, value=f"{comp_info['codigo']} - {comp_info['descricao']}")
    # O: Total base (fórmula referenciando coluna J da linha do header)
    ws_comp.cell(row=idx, column=15, value=f"=$J${comp_info['header_row']}")
    # P: Total por metro (fórmula referenciando coluna K da linha do header)
    ws_comp.cell(row=idx, column=16, value=f"=$K${comp_info['header_row']}")
    # Q: Descrição curta
    ws_comp.cell(row=idx, column=17, value=comp_info['descricao'])
    # R: desc_pre
    ws_comp.cell(row=idx, column=18, value=comp_info['desc_pre'])
    # S: desc_pos
    ws_comp.cell(row=idx, column=19, value=comp_info['desc_pos'])

# Ajustar larguras
ws_comp.column_dimensions['A'].width = 18
ws_comp.column_dimensions['B'].width = 55
ws_comp.column_dimensions['C'].width = 8
ws_comp.column_dimensions['D'].width = 18
ws_comp.column_dimensions['E'].width = 40
ws_comp.column_dimensions['F'].width = 8
ws_comp.column_dimensions['G'].width = 12
ws_comp.column_dimensions['H'].width = 12
ws_comp.column_dimensions['I'].width = 14
ws_comp.column_dimensions['J'].width = 12
ws_comp.column_dimensions['K'].width = 12
ws_comp.column_dimensions['L'].width = 65
ws_comp.column_dimensions['L'].hidden = True  # Ocultar coluna Seleção
# Colunas de lookup (M-S) - ocultas
ws_comp.column_dimensions['M'].width = 18
ws_comp.column_dimensions['M'].hidden = True
ws_comp.column_dimensions['N'].width = 55
ws_comp.column_dimensions['N'].hidden = True
ws_comp.column_dimensions['O'].width = 12
ws_comp.column_dimensions['O'].hidden = True
ws_comp.column_dimensions['P'].width = 12
ws_comp.column_dimensions['P'].hidden = True
ws_comp.column_dimensions['Q'].width = 45
ws_comp.column_dimensions['Q'].hidden = True
ws_comp.column_dimensions['R'].width = 45
ws_comp.column_dimensions['R'].hidden = True
ws_comp.column_dimensions['S'].width = 35
ws_comp.column_dimensions['S'].hidden = True

# Adicionar validação de dados para tipo (coluna C)
dv_tipo_comp = DataValidation(type="list", formula1='"MAT,MO,FER,EQP"', allow_blank=True)
dv_tipo_comp.error = 'Selecione um tipo válido'
dv_tipo_comp.errorTitle = 'Tipo inválido'
ws_comp.add_data_validation(dv_tipo_comp)
dv_tipo_comp.add(f'C2:C500')

# Validação dinâmica para Cód. Item (coluna D) usando INDIRECT
# A fórmula =INDIRECT("LISTA_"&C2) constrói o nome do intervalo baseado no tipo
for row_idx in range(2, 501):
    dv_item_comp = DataValidation(
        type="list",
        formula1=f'=INDIRECT("LISTA_"&$C${row_idx})',
        allow_blank=True
    )
    dv_item_comp.error = 'Primeiro selecione o TIPO, depois escolha o item'
    dv_item_comp.errorTitle = 'Item inválido'
    dv_item_comp.showErrorMessage = True
    dv_item_comp.showDropDown = False
    ws_comp.add_data_validation(dv_item_comp)
    dv_item_comp.add(f'D{row_idx}')

# ============================================
# ABA 8 - ORÇAMENTO
# ============================================
ws_orc = wb.create_sheet('ORCAMENTO')

# Cabeçalho do orçamento
ws_orc.merge_cells('A1:H1')
cell_titulo = ws_orc.cell(row=1, column=1, value='ORÇAMENTO DE CLIMATIZAÇÃO')
cell_titulo.font = Font(bold=True, size=16, color='2E86AB')
cell_titulo.alignment = Alignment(horizontal='center')

# Campos do cabeçalho
ws_orc.cell(row=3, column=1, value='Cliente:').font = Font(bold=True)
ws_orc.merge_cells('B3:E3')
ws_orc.cell(row=3, column=2).fill = input_fill
ws_orc.cell(row=3, column=2).border = thin_border

ws_orc.cell(row=4, column=1, value='Endereço:').font = Font(bold=True)
ws_orc.merge_cells('B4:E4')
ws_orc.cell(row=4, column=2).fill = input_fill
ws_orc.cell(row=4, column=2).border = thin_border

ws_orc.cell(row=5, column=1, value='Data:').font = Font(bold=True)
ws_orc.cell(row=5, column=2).fill = input_fill
ws_orc.cell(row=5, column=2).border = thin_border

ws_orc.cell(row=5, column=3, value='Validade:').font = Font(bold=True)
ws_orc.cell(row=5, column=4).fill = input_fill
ws_orc.cell(row=5, column=4).border = thin_border

ws_orc.cell(row=6, column=1, value='Responsável:').font = Font(bold=True)
ws_orc.merge_cells('B6:C6')
ws_orc.cell(row=6, column=2).fill = input_fill
ws_orc.cell(row=6, column=2).border = thin_border

ws_orc.cell(row=6, column=4, value='Telefone:').font = Font(bold=True)
ws_orc.merge_cells('E6:F6')
ws_orc.cell(row=6, column=5).fill = input_fill
ws_orc.cell(row=6, column=5).border = thin_border

# Tabela de itens - Estrutura com dropdown condicional e descrição
# Coluna C tem dropdown híbrido CÓDIGO - DESCRIÇÃO que filtra por Tipo
orc_headers = ['Item', 'Tipo', 'Serviço/Item', 'Descrição', 'Qtd', 'Variável', 'Preço Unit.', 'Total']
header_row = 9
for col, header in enumerate(orc_headers, 1):
    cell = ws_orc.cell(row=header_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Linhas para preenchimento (30 linhas)
for row_idx in range(header_row + 1, header_row + 31):
    item_num = row_idx - header_row

    # Coluna A: Item
    cell_item = ws_orc.cell(row=row_idx, column=1, value=item_num)
    cell_item.alignment = Alignment(horizontal='center')
    cell_item.border = thin_border

    # Coluna B: Tipo
    cell_tipo = ws_orc.cell(row=row_idx, column=2, value='')
    cell_tipo.alignment = Alignment(horizontal='center')
    cell_tipo.border = thin_border
    cell_tipo.fill = input_fill

    # Coluna C: Serviço/Item - Dropdown híbrido condicional (CÓDIGO - DESCRIÇÃO)
    cell_selecao = ws_orc.cell(row=row_idx, column=3, value='')
    cell_selecao.font = Font(size=8)
    cell_selecao.alignment = Alignment(horizontal='left')
    cell_selecao.border = thin_border
    cell_selecao.fill = input_fill

    # Coluna D: Descrição - Fórmula dinâmica com plural
    # Para COMP: usa desc_pre + variável + "metro(s)" + desc_pos (se variável preenchida)
    # Para COMP sem variável: usa descrição curta
    # Para outros: busca descrição na tabela correspondente
    extract_code = f'IFERROR(LEFT(C{row_idx},FIND(" - ",C{row_idx})-1),C{row_idx})'
    # Fórmula COMP: Se desc_pre vazio ou variável vazia, usa descrição curta; senão monta dinâmica
    # COMPOSICOES: M=código, N=seleção, O=total_base, P=total_metro, Q=descricao, R=desc_pre, S=desc_pos
    # Índices no range M:S: 1=código, 2=seleção, 3=total_base, 4=total_metro, 5=descricao, 6=desc_pre, 7=desc_pos
    formula_comp_desc = f'''IF(OR(VLOOKUP({extract_code},COMPOSICOES!$M:$S,6,FALSE)="",F{row_idx}=""),VLOOKUP({extract_code},COMPOSICOES!$M:$S,5,FALSE),VLOOKUP({extract_code},COMPOSICOES!$M:$S,6,FALSE)&F{row_idx}&IF(F{row_idx}=1," metro"," metros")&VLOOKUP({extract_code},COMPOSICOES!$M:$S,7,FALSE))'''
    formula_desc = f'''=IF(OR(B{row_idx}="",C{row_idx}=""),"",IF(B{row_idx}="COMP",{formula_comp_desc},IF(B{row_idx}="EQP",VLOOKUP({extract_code},EQUIPAMENTOS!$A:$F,3,FALSE),IF(B{row_idx}="MAT",VLOOKUP({extract_code},MATERIAIS!$A:$E,3,FALSE),IF(B{row_idx}="MO",VLOOKUP({extract_code},MAO_DE_OBRA!$A:$E,3,FALSE),IF(B{row_idx}="FER",VLOOKUP({extract_code},FERRAMENTAS!$A:$F,3,FALSE),""))))))'''
    cell_desc = ws_orc.cell(row=row_idx, column=4, value=formula_desc)
    cell_desc.font = Font(size=10)
    cell_desc.alignment = Alignment(horizontal='left', wrap_text=True)
    cell_desc.border = thin_border

    # Coluna E: Quantidade
    cell_qtd = ws_orc.cell(row=row_idx, column=5, value='')
    cell_qtd.alignment = Alignment(horizontal='center')
    cell_qtd.border = thin_border
    cell_qtd.fill = input_fill
    cell_qtd.number_format = qty_format

    # Coluna F: Variável (metros de linha, cabo, duto, etc.)
    cell_metros = ws_orc.cell(row=row_idx, column=6, value='')
    cell_metros.alignment = Alignment(horizontal='center')
    cell_metros.border = thin_border
    cell_metros.fill = input_fill
    cell_metros.number_format = qty_format

    # Coluna G: Preço Unitário (fórmula)
    # Para COMP: Total Base + (Total/Metro × Metros)
    # Para outros: busca preço na tabela correspondente
    # COMPOSICOES: M=código, O=total_base (índice 3), P=total_metro (índice 4)
    formula_preco = f'''=IF(OR(B{row_idx}="",C{row_idx}=""),"",IF(B{row_idx}="COMP",VLOOKUP({extract_code},COMPOSICOES!$M:$S,3,FALSE)+VLOOKUP({extract_code},COMPOSICOES!$M:$S,4,FALSE)*IF(F{row_idx}="",0,F{row_idx}),IF(B{row_idx}="EQP",VLOOKUP({extract_code},EQUIPAMENTOS!$A:$F,6,FALSE),IF(B{row_idx}="MAT",VLOOKUP({extract_code},MATERIAIS!$A:$E,5,FALSE),IF(B{row_idx}="MO",VLOOKUP({extract_code},MAO_DE_OBRA!$A:$E,5,FALSE),IF(B{row_idx}="FER",VLOOKUP({extract_code},FERRAMENTAS!$A:$F,6,FALSE),""))))))'''
    cell_preco = ws_orc.cell(row=row_idx, column=7, value=formula_preco)
    cell_preco.border = thin_border
    cell_preco.number_format = currency_format

    # Coluna H: Total (fórmula) = Quantidade * Preço Unitário
    formula_total = f'=IF(OR(E{row_idx}="",G{row_idx}=""),"",E{row_idx}*G{row_idx})'
    cell_total = ws_orc.cell(row=row_idx, column=8, value=formula_total)
    cell_total.border = thin_border
    cell_total.number_format = currency_format

# Validação de dados para Tipo
dv_tipo = DataValidation(type="list", formula1='"COMP,EQP,MAT,MO,FER"', allow_blank=True)
dv_tipo.error = 'Selecione um tipo válido'
dv_tipo.errorTitle = 'Tipo inválido'
ws_orc.add_data_validation(dv_tipo)
dv_tipo.add(f'B{header_row + 1}:B{header_row + 30}')

# Validação dinâmica para Serviço/Item usando INDIRECT
# A fórmula =INDIRECT("LISTA_"&B10) constrói dinamicamente o nome do intervalo
# baseado no tipo selecionado na coluna B (COMP, EQP, MAT, MO, FER)
# Cada linha precisa de sua própria referência à célula de tipo correspondente
for row_idx in range(header_row + 1, header_row + 31):
    # Criar validação individual por linha com referência à célula de tipo da mesma linha
    dv_item = DataValidation(
        type="list",
        formula1=f'=INDIRECT("LISTA_"&$B${row_idx})',
        allow_blank=True
    )
    dv_item.error = 'Primeiro selecione o TIPO, depois escolha o item da lista'
    dv_item.errorTitle = 'Item inválido'
    dv_item.showErrorMessage = True
    dv_item.showDropDown = False  # False = mostrar dropdown (confuso, mas é assim no openpyxl)
    ws_orc.add_data_validation(dv_item)
    dv_item.add(f'C{row_idx}')

# Coluna C (Seleção) visível para permitir uso do dropdown condicional
# A coluna mostra o dropdown híbrido CÓDIGO - DESCRIÇÃO

# Totalizadores
total_row = header_row + 32

# Subtotal de Custos Diretos
ws_orc.merge_cells(f'A{total_row}:F{total_row}')
cell_subtotal_label = ws_orc.cell(row=total_row, column=1, value='SUBTOTAL DE CUSTOS DIRETOS')
cell_subtotal_label.font = Font(bold=True)
cell_subtotal_label.alignment = Alignment(horizontal='right')
cell_subtotal_label.border = thin_border

ws_orc.cell(row=total_row, column=7, value='').border = thin_border

formula_subtotal = f'=SUM(H{header_row + 1}:H{header_row + 30})'
cell_subtotal = ws_orc.cell(row=total_row, column=8, value=formula_subtotal)
cell_subtotal.font = Font(bold=True)
cell_subtotal.border = thin_border
cell_subtotal.number_format = currency_format

subtotal_row = total_row

# BDI/Markup
total_row += 1
ws_orc.merge_cells(f'A{total_row}:F{total_row}')
formula_bdi_label = f'=IF(NEGOCIO!B7="BDI","+ BDI (" & TEXT(NEGOCIO!B{total_bdi_row},"0.00") & "%)","+ MARKUP (" & TEXT(NEGOCIO!B{markup_input_row},"0.00") & "%)")'
cell_bdi_label = ws_orc.cell(row=total_row, column=1, value=formula_bdi_label)
cell_bdi_label.font = Font(bold=True)
cell_bdi_label.alignment = Alignment(horizontal='right')
cell_bdi_label.border = thin_border

ws_orc.cell(row=total_row, column=7, value='').border = thin_border

formula_bdi_value = f'=H{subtotal_row}*(NEGOCIO!B{mult_calc_row}-1)'
cell_bdi_value = ws_orc.cell(row=total_row, column=8, value=formula_bdi_value)
cell_bdi_value.font = Font(bold=True)
cell_bdi_value.border = thin_border
cell_bdi_value.number_format = currency_format

bdi_row = total_row

# Subtotal com BDI/Markup
total_row += 1
ws_orc.merge_cells(f'A{total_row}:F{total_row}')
cell_subtotal_bdi_label = ws_orc.cell(row=total_row, column=1, value='SUBTOTAL (COM MARGEM)')
cell_subtotal_bdi_label.font = Font(bold=True)
cell_subtotal_bdi_label.alignment = Alignment(horizontal='right')
cell_subtotal_bdi_label.border = thin_border
cell_subtotal_bdi_label.fill = PatternFill('solid', fgColor='D5F5E3')

ws_orc.cell(row=total_row, column=7, value='').border = thin_border
ws_orc.cell(row=total_row, column=7).fill = PatternFill('solid', fgColor='D5F5E3')

formula_subtotal_bdi = f'=H{subtotal_row}+H{bdi_row}'
cell_subtotal_bdi = ws_orc.cell(row=total_row, column=8, value=formula_subtotal_bdi)
cell_subtotal_bdi.font = Font(bold=True)
cell_subtotal_bdi.border = thin_border
cell_subtotal_bdi.number_format = currency_format
cell_subtotal_bdi.fill = PatternFill('solid', fgColor='D5F5E3')

subtotal_bdi_row = total_row

# Desconto
total_row += 1
ws_orc.merge_cells(f'A{total_row}:F{total_row}')
cell_desc_label = ws_orc.cell(row=total_row, column=1, value='- DESCONTO')
cell_desc_label.font = Font(bold=True)
cell_desc_label.alignment = Alignment(horizontal='right')
cell_desc_label.border = thin_border

cell_desc_perc = ws_orc.cell(row=total_row, column=7, value=0)
cell_desc_perc.border = thin_border
cell_desc_perc.fill = input_fill
cell_desc_perc.number_format = '0.0%'
cell_desc_perc.alignment = Alignment(horizontal='center')

formula_desconto = f'=-H{subtotal_bdi_row}*G{total_row}'
cell_desconto = ws_orc.cell(row=total_row, column=8, value=formula_desconto)
cell_desconto.border = thin_border
cell_desconto.number_format = currency_format

desconto_row = total_row

# Total Geral
total_row += 1
ws_orc.merge_cells(f'A{total_row}:F{total_row}')
cell_total_label = ws_orc.cell(row=total_row, column=1, value='TOTAL GERAL')
cell_total_label.font = Font(bold=True, size=12)
cell_total_label.alignment = Alignment(horizontal='right')
cell_total_label.fill = total_fill
cell_total_label.border = thin_border

for col in range(2, 8):
    ws_orc.cell(row=total_row, column=col).fill = total_fill
    ws_orc.cell(row=total_row, column=col).border = thin_border

ws_orc.cell(row=total_row, column=7, value='').fill = total_fill

formula_total_geral = f'=H{subtotal_bdi_row}+H{desconto_row}'
cell_total_geral = ws_orc.cell(row=total_row, column=8, value=formula_total_geral)
cell_total_geral.font = Font(bold=True, size=12)
cell_total_geral.fill = total_fill
cell_total_geral.border = thin_border
cell_total_geral.number_format = currency_format

# Observações
obs_row = total_row + 2
ws_orc.cell(row=obs_row, column=1, value='Observações:').font = Font(bold=True)
ws_orc.merge_cells(f'A{obs_row + 1}:H{obs_row + 4}')
cell_obs = ws_orc.cell(row=obs_row + 1, column=1, value='')
cell_obs.fill = input_fill
cell_obs.border = thin_border
cell_obs.alignment = Alignment(vertical='top', wrap_text=True)

# Ajustar larguras
# Coluna A: Item, B: Tipo, C: Serviço/Item (dropdown), D: Descrição, E: Qtd, F: Variável, G: Preço, H: Total
ws_orc.column_dimensions['A'].width = 6
ws_orc.column_dimensions['B'].width = 8
ws_orc.column_dimensions['C'].width = 55  # Dropdown híbrido CÓDIGO - DESCRIÇÃO
ws_orc.column_dimensions['D'].width = 65  # Descrição principal
ws_orc.column_dimensions['E'].width = 8
ws_orc.column_dimensions['F'].width = 10
ws_orc.column_dimensions['G'].width = 14
ws_orc.column_dimensions['H'].width = 14

# Altura das linhas de observação
for row in range(obs_row + 1, obs_row + 5):
    ws_orc.row_dimensions[row].height = 20

# ============================================
# CONFIGURAÇÕES FINAIS
# ============================================

# Congelar primeira linha em abas de dados
ws_mat.freeze_panes = 'A2'
ws_mo.freeze_panes = 'A2'
ws_fer.freeze_panes = 'A2'
ws_eqp.freeze_panes = 'A2'
ws_comp.freeze_panes = 'A2'
ws_orc.freeze_panes = 'A10'

# Criar intervalos nomeados para dropdowns híbridos (CÓDIGO - DESCRIÇÃO)
# Apontam diretamente para a coluna "Seleção" de cada aba
comp_list_last_row = len(composicoes) + 1
eqp_list_last_row = len(equipamentos) + 1
mat_list_last_row = len(materiais) + 1
mo_list_last_row = len(mao_de_obra) + 1
fer_list_last_row = len(ferramentas) + 1

# Intervalos nomeados apontando para as colunas de Seleção nas respectivas abas
# COMP usa coluna N (LK_SELECAO) na tabela de lookup da aba COMPOSICOES
wb.defined_names['LISTA_COMP'] = DefinedName('LISTA_COMP', attr_text=f'COMPOSICOES!$N$2:$N${comp_list_last_row}')
# MAT, MO, FER, EQP apontam diretamente para suas abas
wb.defined_names['LISTA_EQP'] = DefinedName('LISTA_EQP', attr_text=f'EQUIPAMENTOS!$G$2:$G${eqp_list_last_row}')
wb.defined_names['LISTA_MAT'] = DefinedName('LISTA_MAT', attr_text=f'MATERIAIS!$F$2:$F${mat_list_last_row}')
wb.defined_names['LISTA_MO'] = DefinedName('LISTA_MO', attr_text=f'MAO_DE_OBRA!$F$2:$F${mo_list_last_row}')
wb.defined_names['LISTA_FER'] = DefinedName('LISTA_FER', attr_text=f'FERRAMENTAS!$G$2:$G${fer_list_last_row}')

# Salvar arquivo
import os
output_path = os.path.join(os.path.dirname(__file__), 'Composicoes_Split_v3.xlsx')
wb.save(output_path)
print(f'Arquivo salvo em: {output_path}')
