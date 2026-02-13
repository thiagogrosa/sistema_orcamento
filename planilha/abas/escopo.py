"""
Aba ESCOPO - Selecao de servicos para o orcamento.

Esta aba contem apenas a selecao de servicos/itens, sem dados do cliente.
As linhas 1-5 sao reservadas para botoes (Criar Composicao, etc.).

IMPORTANTE: As colunas F (Descricao) e G (Preco Unit.) sao preenchidas por
formulas de spill no template.xlsm. O Python so preenche:
- Estrutura geral (headers, totalizadores)
- Validacoes de dados
- Formatacao

Estrutura de colunas:
- A: Item (numero sequencial)
- B: Tipo (COMP, EQP, MAT, MO, FER)
- C: Servico/Item (dropdown dinamico)
- D: Qtd (quantidade - usuario preenche)
- E: Variavel (metros de linha - usuario preenche)
- F: Descricao (preenchida por spill)
- G: Preco Unit. (preenchido por spill)
- H: Total (formula simples)
"""
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


def _aplicar_estrutura(ws, estilos, config_negocio):
    """
    Aplica toda a estrutura da aba ESCOPO.
    """
    thin_border = estilos['thin_border']
    header_font = estilos['header_font']
    header_fill = estilos['header_fill']
    currency_format = estilos['currency_format']
    qty_format = estilos['qty_format']
    input_fill = estilos['input_fill']
    total_fill = estilos['total_fill']

    # Linhas 1-5: Reservadas para botoes (deixar vazio)
    # O usuario pode adicionar botoes manualmente no Excel

    # Tabela de itens - headers na linha 6
    # Nova ordem: Item, Tipo, Servico/Item, Qtd, Variavel, Descricao, Preco Unit., Total
    escopo_headers = ['Item', 'Tipo', 'Servico/Item', 'Qtd', 'Variavel', 'Descricao', 'Preco Unit.', 'Total']
    header_row = 6
    for col, header in enumerate(escopo_headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 30 linhas de itens (linhas 7-36)
    for row_idx in range(header_row + 1, header_row + 31):
        item_num = row_idx - header_row

        # Coluna A: Item
        cell_item = ws.cell(row=row_idx, column=1, value=item_num)
        cell_item.alignment = Alignment(horizontal='center')
        cell_item.border = thin_border

        # Coluna B: Tipo
        cell_tipo = ws.cell(row=row_idx, column=2, value='')
        cell_tipo.alignment = Alignment(horizontal='center')
        cell_tipo.border = thin_border
        cell_tipo.fill = input_fill

        # Coluna C: Servico/Item (DADO - usuario preenche)
        cell_selecao = ws.cell(row=row_idx, column=3, value='')
        cell_selecao.font = Font(size=8)
        cell_selecao.alignment = Alignment(horizontal='left')
        cell_selecao.border = thin_border
        cell_selecao.fill = input_fill

        # Coluna D: Quantidade (DADO - usuario preenche)
        cell_qtd = ws.cell(row=row_idx, column=4, value='')
        cell_qtd.alignment = Alignment(horizontal='center')
        cell_qtd.border = thin_border
        cell_qtd.fill = input_fill
        cell_qtd.number_format = qty_format

        # Coluna E: Variavel (DADO - usuario preenche)
        cell_metros = ws.cell(row=row_idx, column=5, value='')
        cell_metros.alignment = Alignment(horizontal='center')
        cell_metros.border = thin_border
        cell_metros.fill = input_fill
        cell_metros.number_format = qty_format

        # Coluna F: Descricao (SPILL preenchera via template)
        cell_desc = ws.cell(row=row_idx, column=6, value='')
        cell_desc.font = Font(size=10)
        cell_desc.alignment = Alignment(horizontal='left', wrap_text=True)
        cell_desc.border = thin_border

        # Coluna G: Preco Unitario (SPILL preenchera via template)
        cell_preco = ws.cell(row=row_idx, column=7, value='')
        cell_preco.border = thin_border
        cell_preco.number_format = currency_format

        # Coluna H: Total (formula simples - Qtd esta na coluna D agora)
        formula_total = f'=IF(OR(D{row_idx}="",G{row_idx}=""),"",D{row_idx}*G{row_idx})'
        cell_total = ws.cell(row=row_idx, column=8, value=formula_total)
        cell_total.border = thin_border
        cell_total.number_format = currency_format

    # Validacao para Tipo
    dv_tipo = DataValidation(type="list", formula1='"COMP,EQP,MAT,MO,FER"', allow_blank=True)
    dv_tipo.error = 'Selecione um tipo valido'
    dv_tipo.errorTitle = 'Tipo invalido'
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f'B{header_row + 1}:B{header_row + 30}')

    # Validacao dinamica para Servico/Item
    for row_idx in range(header_row + 1, header_row + 31):
        dv_item = DataValidation(
            type="list",
            formula1=f'=INDIRECT("LISTA_"&$B${row_idx})',
            allow_blank=True
        )
        dv_item.error = 'Primeiro selecione o TIPO, depois escolha o item da lista'
        dv_item.errorTitle = 'Item invalido'
        dv_item.showErrorMessage = True
        dv_item.showDropDown = False
        ws.add_data_validation(dv_item)
        dv_item.add(f'C{row_idx}')

    # Totalizadores (linha 38)
    total_row = header_row + 32  # 38

    # Subtotal (ja com margem aplicada)
    ws.merge_cells(f'A{total_row}:F{total_row}')
    cell_subtotal_label = ws.cell(row=total_row, column=1, value='SUBTOTAL (COM MARGEM APLICADA)')
    cell_subtotal_label.font = Font(bold=True)
    cell_subtotal_label.alignment = Alignment(horizontal='right')
    cell_subtotal_label.border = thin_border
    cell_subtotal_label.fill = PatternFill('solid', fgColor='D5F5E3')

    ws.cell(row=total_row, column=7, value='').border = thin_border
    ws.cell(row=total_row, column=7).fill = PatternFill('solid', fgColor='D5F5E3')

    formula_subtotal = f'=SUM(H{header_row + 1}:H{header_row + 30})'
    cell_subtotal = ws.cell(row=total_row, column=8, value=formula_subtotal)
    cell_subtotal.font = Font(bold=True)
    cell_subtotal.border = thin_border
    cell_subtotal.number_format = currency_format
    cell_subtotal.fill = PatternFill('solid', fgColor='D5F5E3')

    subtotal_row = total_row

    # Desconto (linha 39)
    total_row += 1
    ws.merge_cells(f'A{total_row}:F{total_row}')
    cell_desc_label = ws.cell(row=total_row, column=1, value='- DESCONTO')
    cell_desc_label.font = Font(bold=True)
    cell_desc_label.alignment = Alignment(horizontal='right')
    cell_desc_label.border = thin_border

    cell_desc_perc = ws.cell(row=total_row, column=7, value=0)
    cell_desc_perc.border = thin_border
    cell_desc_perc.fill = input_fill
    cell_desc_perc.number_format = '0.0%'
    cell_desc_perc.alignment = Alignment(horizontal='center')

    formula_desconto = f'=-H{subtotal_row}*G{total_row}'
    cell_desconto = ws.cell(row=total_row, column=8, value=formula_desconto)
    cell_desconto.border = thin_border
    cell_desconto.number_format = currency_format

    desconto_row = total_row

    # Total Geral (linha 40)
    total_row += 1
    ws.merge_cells(f'A{total_row}:F{total_row}')
    cell_total_label = ws.cell(row=total_row, column=1, value='TOTAL GERAL')
    cell_total_label.font = Font(bold=True, size=12)
    cell_total_label.alignment = Alignment(horizontal='right')
    cell_total_label.fill = total_fill
    cell_total_label.border = thin_border

    for col in range(2, 8):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = thin_border

    ws.cell(row=total_row, column=7, value='').fill = total_fill

    formula_total_geral = f'=H{subtotal_row}+H{desconto_row}'
    cell_total_geral = ws.cell(row=total_row, column=8, value=formula_total_geral)
    cell_total_geral.font = Font(bold=True, size=12)
    cell_total_geral.fill = total_fill
    cell_total_geral.border = thin_border
    cell_total_geral.number_format = currency_format

    # Observacoes (linha 42)
    obs_row = total_row + 2
    ws.cell(row=obs_row, column=1, value='Observacoes:').font = Font(bold=True)
    ws.merge_cells(f'A{obs_row + 1}:H{obs_row + 4}')
    cell_obs = ws.cell(row=obs_row + 1, column=1, value='')
    cell_obs.fill = input_fill
    cell_obs.border = thin_border
    cell_obs.alignment = Alignment(vertical='top', wrap_text=True)

    # Ajustar larguras das colunas
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 65
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    # Altura das linhas de observacao
    for row in range(obs_row + 1, obs_row + 5):
        ws.row_dimensions[row].height = 20

    # Congelar apos header (linha 7 em diante rola)
    ws.freeze_panes = 'A7'

    return ws


def criar(wb, estilos, config_negocio):
    """
    Cria a aba ESCOPO.

    Args:
        wb: Workbook
        estilos: dict de estilos
        config_negocio: dict com mult_rows para formulas de multiplicadores
    """
    ws = wb.create_sheet('ESCOPO')
    return _aplicar_estrutura(ws, estilos, config_negocio)


def preencher(wb, estilos, config_negocio):
    """
    Preenche a aba ESCOPO existente no template.

    A logica e identica ao criar(), mas usa aba existente.
    Se a aba nao existir, ela sera criada.
    Se existir ORCAMENTO (nome antigo), renomeia para ESCOPO.
    """
    if 'ESCOPO' not in wb.sheetnames:
        if 'ORCAMENTO' in wb.sheetnames:
            # Renomear aba antiga
            ws = wb['ORCAMENTO']
            ws.title = 'ESCOPO'
        else:
            ws = wb.create_sheet('ESCOPO')
    else:
        ws = wb['ESCOPO']

    # Desfazer mesclagens existentes antes de limpar
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Limpar conteudo existente
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    return _aplicar_estrutura(ws, estilos, config_negocio)
