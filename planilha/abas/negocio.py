"""
Aba NEGOCIO - Configuracao de BDI e Markup por tipo de item.

Estrutura:
- BDI Geral: Usado para MAT, MO, FER
- BDI Equipamentos: Usado apenas para EQP (diferenciado)
- Markup por tipo: Alternativa simples ao BDI
- Multiplicadores calculados: Um para cada tipo (MAT, MO, FER, EQP)
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


def _criar_secao_bdi(ws, estilos, start_row, titulo, is_eqp=False):
    """
    Cria uma secao de BDI (geral ou diferenciado).

    Retorna a linha do total BDI.
    """
    thin_border = estilos['thin_border']
    input_fill = estilos['input_fill']

    # Titulo da secao
    ws.merge_cells(f'A{start_row}:F{start_row}')
    cell_label = ws.cell(row=start_row, column=1, value=titulo)
    cell_label.font = estilos['section_font']
    cell_label.fill = estilos['section_fill']

    # Cabecalho da tabela
    headers_row = start_row + 2
    ws.cell(row=headers_row, column=1, value='Componente').font = Font(bold=True)
    ws.cell(row=headers_row, column=2, value='%').font = Font(bold=True)
    ws.cell(row=headers_row, column=3, value='Descricao').font = Font(bold=True)

    for col in range(1, 4):
        ws.cell(row=headers_row, column=col).fill = PatternFill('solid', fgColor='D5D8DC')
        ws.cell(row=headers_row, column=col).border = thin_border
        ws.cell(row=headers_row, column=col).alignment = Alignment(horizontal='center')

    # Componentes do BDI
    # Valores diferentes para EQP (tipicamente menor margem em equipamentos)
    if is_eqp:
        bdi_components = [
            ('Administracao Central', 3.0, 'Custos de escritorio, gerencia, RH, etc.'),
            ('Despesas Financeiras', 0.8, 'Juros, taxas bancarias, capital de giro'),
            ('Lucro', 5.0, 'Margem de lucro desejada'),
            ('Impostos', 13.65, 'ISS, PIS, COFINS, etc.'),
            ('Riscos e Imprevistos', 1.0, 'Contingencias, garantias, seguros'),
            ('Outros', 0.0, 'Outros custos indiretos'),
        ]
    else:
        bdi_components = [
            ('Administracao Central', 5.0, 'Custos de escritorio, gerencia, RH, etc.'),
            ('Despesas Financeiras', 1.2, 'Juros, taxas bancarias, capital de giro'),
            ('Lucro', 8.0, 'Margem de lucro desejada'),
            ('Impostos', 13.65, 'ISS, PIS, COFINS, etc. (pode variar por municipio)'),
            ('Riscos e Imprevistos', 2.0, 'Contingencias, garantias, seguros'),
            ('Outros', 0.0, 'Outros custos indiretos'),
        ]

    current_row = headers_row + 1
    for comp_name, comp_value, comp_desc in bdi_components:
        ws.cell(row=current_row, column=1, value=comp_name).border = thin_border

        cell_value = ws.cell(row=current_row, column=2, value=comp_value)
        cell_value.border = thin_border
        cell_value.fill = input_fill
        cell_value.number_format = '0.00'
        cell_value.alignment = Alignment(horizontal='center')

        ws.cell(row=current_row, column=3, value=comp_desc).border = thin_border
        ws.cell(row=current_row, column=3).font = Font(italic=True, size=9)

        current_row += 1

    # Total BDI
    ws.cell(row=current_row, column=1, value='TOTAL BDI (%)').font = Font(bold=True)
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=1).fill = PatternFill('solid', fgColor='BDC3C7')

    formula_total = f'=SUM(B{headers_row + 1}:B{current_row - 1})'
    cell_total = ws.cell(row=current_row, column=2, value=formula_total)
    cell_total.border = thin_border
    cell_total.fill = PatternFill('solid', fgColor='BDC3C7')
    cell_total.number_format = '0.00'
    cell_total.font = Font(bold=True)
    cell_total.alignment = Alignment(horizontal='center')

    ws.cell(row=current_row, column=3, value='Soma de todos os componentes').border = thin_border
    ws.cell(row=current_row, column=3).fill = PatternFill('solid', fgColor='BDC3C7')
    ws.cell(row=current_row, column=3).font = Font(italic=True, size=9)

    return current_row


def _criar_secao_markup_por_tipo(ws, estilos, start_row):
    """
    Cria a secao de markup simples por tipo de item.

    Retorna dict com as linhas de cada tipo.
    """
    thin_border = estilos['thin_border']
    input_fill = estilos['input_fill']

    # Titulo da secao
    ws.merge_cells(f'A{start_row}:F{start_row}')
    cell_label = ws.cell(row=start_row, column=1, value='MARKUP SIMPLES POR TIPO (Alternativa ao BDI)')
    cell_label.font = estilos['section_font']
    cell_label.fill = estilos['section_fill']

    # Cabecalho da tabela
    headers_row = start_row + 2
    ws.cell(row=headers_row, column=1, value='Tipo').font = Font(bold=True)
    ws.cell(row=headers_row, column=2, value='Markup (%)').font = Font(bold=True)
    ws.cell(row=headers_row, column=3, value='Descricao').font = Font(bold=True)

    for col in range(1, 4):
        ws.cell(row=headers_row, column=col).fill = PatternFill('solid', fgColor='D5D8DC')
        ws.cell(row=headers_row, column=col).border = thin_border
        ws.cell(row=headers_row, column=col).alignment = Alignment(horizontal='center')

    # Markup por tipo
    markup_items = [
        ('MAT', 35.0, 'Materiais'),
        ('MO', 50.0, 'Mao de Obra'),
        ('FER', 30.0, 'Ferramentas'),
        ('EQP', 20.0, 'Equipamentos'),
    ]

    rows = {}
    current_row = headers_row + 1
    for tipo, valor, desc in markup_items:
        ws.cell(row=current_row, column=1, value=tipo).border = thin_border
        ws.cell(row=current_row, column=1).font = Font(bold=True, name='Consolas')
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')

        cell_value = ws.cell(row=current_row, column=2, value=valor)
        cell_value.border = thin_border
        cell_value.fill = input_fill
        cell_value.number_format = '0.00'
        cell_value.alignment = Alignment(horizontal='center')

        ws.cell(row=current_row, column=3, value=desc).border = thin_border
        ws.cell(row=current_row, column=3).font = Font(italic=True, size=9)

        rows[tipo] = current_row
        current_row += 1

    return rows


def _criar_secao_multiplicadores(ws, estilos, start_row, total_bdi_geral_row, total_bdi_eqp_row, markup_rows):
    """
    Cria a secao de multiplicadores calculados por tipo.

    Retorna dict com as linhas de cada multiplicador.
    """
    thin_border = estilos['thin_border']

    # Titulo da secao
    ws.merge_cells(f'A{start_row}:F{start_row}')
    cell_label = ws.cell(row=start_row, column=1, value='MULTIPLICADORES CALCULADOS')
    cell_label.font = Font(bold=True, size=14, color='FFFFFF')
    cell_label.fill = PatternFill('solid', fgColor='27AE60')
    cell_label.alignment = Alignment(horizontal='center')

    # Cabecalho da tabela
    headers_row = start_row + 2
    headers = ['Tipo', 'Multiplicador', 'Metodo Usado', 'Formula']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=headers_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='27AE60')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Multiplicadores por tipo
    # MAT, MO, FER usam BDI Geral; EQP usa BDI EQP
    mult_items = [
        ('MAT', total_bdi_geral_row, markup_rows['MAT'], 'BDI Geral ou Markup MAT'),
        ('MO', total_bdi_geral_row, markup_rows['MO'], 'BDI Geral ou Markup MO'),
        ('FER', total_bdi_geral_row, markup_rows['FER'], 'BDI Geral ou Markup FER'),
        ('EQP', total_bdi_eqp_row, markup_rows['EQP'], 'BDI EQP ou Markup EQP'),
    ]

    rows = {}
    current_row = headers_row + 1
    for tipo, bdi_row, markup_row, desc in mult_items:
        # Coluna A: Tipo
        ws.cell(row=current_row, column=1, value=tipo).border = thin_border
        ws.cell(row=current_row, column=1).font = Font(bold=True, name='Consolas')
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')

        # Coluna B: Formula do multiplicador
        formula = f'=IF($B$7="BDI",1+(B{bdi_row}/100),IF($B$7="MARKUP",1+(B{markup_row}/100),1))'
        cell_mult = ws.cell(row=current_row, column=2, value=formula)
        cell_mult.border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        cell_mult.font = Font(bold=True, size=12)
        cell_mult.fill = PatternFill('solid', fgColor='D5F5E3')
        cell_mult.number_format = '0.0000'
        cell_mult.alignment = Alignment(horizontal='center')

        # Coluna C: Metodo usado (formula dinamica)
        formula_metodo = f'=IF($B$7="BDI","BDI {"EQP" if tipo == "EQP" else "Geral"} ("&TEXT(B{bdi_row},"0.00")&"%)","MARKUP ("&TEXT(B{markup_row},"0.00")&"%)")'
        ws.cell(row=current_row, column=3, value=formula_metodo).border = thin_border
        ws.cell(row=current_row, column=3).font = Font(italic=True, size=9)

        # Coluna D: Descricao
        ws.cell(row=current_row, column=4, value=desc).border = thin_border
        ws.cell(row=current_row, column=4).font = Font(italic=True, size=9)

        rows[tipo] = current_row
        current_row += 1

    return rows


def _criar_secao_csv_config(ws, estilos, start_row):
    """
    Cria a secao de configuracao de importacao CSV.

    Posicionada nas colunas A-C apos o conteudo principal.
    """
    thin_border = estilos['thin_border']
    input_fill = estilos['input_fill']

    # Titulo da secao
    ws.merge_cells(f'A{start_row}:C{start_row}')
    cell_label = ws.cell(row=start_row, column=1, value='IMPORTACAO DE CATALOGOS CSV')
    cell_label.font = Font(bold=True, size=11, color='FFFFFF')
    cell_label.fill = PatternFill('solid', fgColor='8E44AD')  # Roxo
    cell_label.alignment = Alignment(horizontal='center')

    # Diretorio CSV
    ws.cell(row=start_row + 2, column=1, value='Diretorio CSV:').font = Font(bold=True)
    cell_path = ws.cell(row=start_row + 2, column=2, value=r'.\dados_csv\\')
    cell_path.fill = input_fill
    cell_path.border = thin_border

    # Ultima importacao
    ws.cell(row=start_row + 3, column=1, value='Ultima Importacao:').font = Font(bold=True)
    cell_timestamp = ws.cell(row=start_row + 3, column=2, value='')
    cell_timestamp.border = thin_border
    cell_timestamp.font = Font(italic=True, size=9)

    # Status de validade
    ws.cell(row=start_row + 4, column=1, value='Status:').font = Font(bold=True)
    cell_status = ws.cell(row=start_row + 4, column=2, value='')
    cell_status.border = thin_border
    cell_status.font = Font(bold=True)

    # Instrucoes
    ws.merge_cells(f'A{start_row + 6}:C{start_row + 6}')
    ws.cell(row=start_row + 6, column=1, value='Use as macros VBA para importar:')
    ws.cell(row=start_row + 6, column=1).font = Font(italic=True, size=9)

    instrucoes = [
        'ImportarTodosCatalogos - Importa todos os catalogos',
        'VerificarValidade - Verifica se dados estao vencidos',
        'AbrirPastaCSV - Abre pasta de CSVs no Explorer',
    ]
    for idx, instr in enumerate(instrucoes):
        ws.cell(row=start_row + 7 + idx, column=1, value=instr)
        ws.cell(row=start_row + 7 + idx, column=1).font = Font(size=9)

    return {
        'path_row': start_row + 2,
        'timestamp_row': start_row + 3,
        'status_row': start_row + 4,
    }


def _aplicar_estrutura(ws, estilos):
    """
    Aplica toda a estrutura da aba NEGOCIO.

    Retorna dict com todas as referencias de linhas.
    """
    thin_border = estilos['thin_border']
    input_fill = estilos['input_fill']

    # Titulo
    ws.merge_cells('A1:F1')
    cell_titulo = ws.cell(row=1, column=1, value='CONFIGURACOES DE NEGOCIO - BDI E MARKUP POR TIPO')
    cell_titulo.font = Font(bold=True, size=16, color='2E86AB')
    cell_titulo.alignment = Alignment(horizontal='center')

    # Explicacao
    ws.cell(row=3, column=1, value='Esta aba permite configurar o multiplicador aplicado ao custo direto por tipo de item.')
    ws.merge_cells('A3:F3')

    # Secao: Metodo
    ws.merge_cells('A5:F5')
    cell_metodo_label = ws.cell(row=5, column=1, value='METODO DE CALCULO')
    cell_metodo_label.font = estilos['section_font']
    cell_metodo_label.fill = estilos['section_fill']

    ws.cell(row=7, column=1, value='Metodo:').font = Font(bold=True)
    ws.cell(row=7, column=2, value='BDI')
    ws.cell(row=7, column=2).fill = input_fill
    ws.cell(row=7, column=2).border = thin_border
    ws.cell(row=7, column=3, value='(Escolha: BDI ou MARKUP)').font = Font(italic=True, size=9)

    # Validacao para metodo
    dv_metodo = DataValidation(type="list", formula1='"BDI,MARKUP"', allow_blank=False)
    ws.add_data_validation(dv_metodo)
    dv_metodo.add('B7')

    # Secao: BDI Geral (MAT, MO, FER)
    total_bdi_geral_row = _criar_secao_bdi(
        ws, estilos,
        start_row=9,
        titulo='BDI GERAL (Materiais, Mao de Obra, Ferramentas)',
        is_eqp=False
    )

    # Secao: BDI Equipamentos (diferenciado)
    bdi_eqp_start = total_bdi_geral_row + 2
    total_bdi_eqp_row = _criar_secao_bdi(
        ws, estilos,
        start_row=bdi_eqp_start,
        titulo='BDI EQUIPAMENTOS (Diferenciado)',
        is_eqp=True
    )

    # Secao: Markup por tipo
    markup_start = total_bdi_eqp_row + 2
    markup_rows = _criar_secao_markup_por_tipo(ws, estilos, markup_start)

    # Secao: Multiplicadores calculados
    mult_start = markup_rows['EQP'] + 2
    mult_rows = _criar_secao_multiplicadores(
        ws, estilos, mult_start,
        total_bdi_geral_row, total_bdi_eqp_row, markup_rows
    )

    # Explicacao final
    expl_row = mult_rows['EQP'] + 2
    ws.merge_cells(f'A{expl_row}:F{expl_row}')
    ws.cell(row=expl_row, column=1, value='COMO FUNCIONA:')
    ws.cell(row=expl_row, column=1).font = Font(bold=True, size=11)

    expl_content = [
        'Os multiplicadores sao aplicados automaticamente nas composicoes por tipo de item.',
        '',
        'BDI: Usa o BDI Geral para MAT, MO, FER e o BDI Equipamentos para EQP.',
        'MARKUP: Usa o percentual especifico de cada tipo.',
        '',
        'O preco final ja inclui a margem e e importado diretamente no orcamento.',
    ]

    for idx, line in enumerate(expl_content):
        row_num = expl_row + 1 + idx
        ws.merge_cells(f'A{row_num}:F{row_num}')
        ws.cell(row=row_num, column=1, value=line)
        if line.startswith('BDI:') or line.startswith('MARKUP:'):
            ws.cell(row=row_num, column=1).font = Font(bold=True)

    # Secao: Configuracao CSV (posicionada apos o conteudo principal)
    csv_start_row = expl_row + len(expl_content) + 3
    csv_config = _criar_secao_csv_config(ws, estilos, start_row=csv_start_row)

    # Ajustar larguras
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 30

    return {
        'ws': ws,
        'total_bdi_geral_row': total_bdi_geral_row,
        'total_bdi_eqp_row': total_bdi_eqp_row,
        'markup_rows': markup_rows,
        'mult_rows': mult_rows,
        'csv_config': csv_config,
    }


def criar(wb, estilos):
    """
    Cria a aba NEGOCIO com configuracoes de BDI/Markup por tipo.

    Retorna dict com:
    - ws: worksheet
    - total_bdi_geral_row: linha do total BDI geral
    - total_bdi_eqp_row: linha do total BDI equipamentos
    - markup_rows: dict com linhas de markup por tipo
    - mult_rows: dict com linhas de multiplicadores por tipo
    """
    ws = wb.create_sheet('NEGOCIO')
    return _aplicar_estrutura(ws, estilos)


def preencher(wb, estilos):
    """
    Preenche a aba NEGOCIO existente no template.

    Retorna dict com as mesmas chaves do criar().
    """
    ws = wb['NEGOCIO']

    # Desfazer mesclagens existentes antes de limpar
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    # Limpar conteúdo existente
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    return _aplicar_estrutura(ws, estilos)
