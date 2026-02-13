#!/usr/bin/env python3
"""
Gerador de Planilha de Custos HVAC Split.

Este script orquestra a criação da planilha Excel completa,
importando os módulos de estilos, dados e abas.

Modos de operação:
1. Sem template: Cria planilha .xlsx do zero
2. Com template: Carrega template.xlsm com VBA e preenche dados, salvando como .xlsm

Uso: python3 criar_planilha.py
"""
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from estilos import criar_estilos
from dados import MATERIAIS, MAO_DE_OBRA, FERRAMENTAS, EQUIPAMENTOS, COMPOSICOES
from abas import instrucoes, prompts, negocio, catalogos, composicoes, cliente, escopo


def _atualizar_intervalo_nomeado(wb, nome, referencia):
    """Atualiza ou cria intervalo nomeado."""
    if nome in wb.defined_names:
        del wb.defined_names[nome]
    wb.defined_names[nome] = DefinedName(nome, attr_text=referencia)


def main():
    base_dir = os.path.dirname(__file__)
    template_path = os.path.join(base_dir, 'template.xlsm')
    usar_template = os.path.exists(template_path)

    # Criar estilos
    estilos = criar_estilos()

    if usar_template:
        print(f'Usando template: {template_path}')
        # Carregar template preservando VBA
        wb = load_workbook(template_path, keep_vba=True)

        # Preencher abas existentes
        instrucoes.preencher(wb, estilos)
        prompts.preencher(wb, estilos)
        config_neg = negocio.preencher(wb, estilos)
        config_cat = catalogos.preencher(wb, estilos, MATERIAIS, MAO_DE_OBRA, FERRAMENTAS, EQUIPAMENTOS)
        config_comp = composicoes.preencher(wb, estilos, COMPOSICOES, config_neg,
                                           MATERIAIS, MAO_DE_OBRA, FERRAMENTAS, EQUIPAMENTOS)
        cliente.preencher(wb, estilos)
        escopo.preencher(wb, estilos, config_neg)

        output_path = os.path.join(base_dir, 'Composicoes_Split_v3.xlsm')
    else:
        print('Template não encontrado, criando planilha do zero...')
        # Criar workbook do zero
        wb = Workbook()

        # Criar abas (ordem importa para referências)
        instrucoes.criar(wb, estilos)
        prompts.criar(wb, estilos)
        config_neg = negocio.criar(wb, estilos)
        config_cat = catalogos.criar(wb, estilos, MATERIAIS, MAO_DE_OBRA, FERRAMENTAS, EQUIPAMENTOS)
        config_comp = composicoes.criar(wb, estilos, COMPOSICOES, config_neg,
                                        MATERIAIS, MAO_DE_OBRA, FERRAMENTAS, EQUIPAMENTOS)
        cliente.criar(wb, estilos)
        escopo.criar(wb, estilos, config_neg)

        output_path = os.path.join(base_dir, 'Composicoes_Split_v3.xlsx')

    # Atualizar intervalos nomeados para dropdowns
    # LISTA_COMP agora na coluna N (Selecao)
    _atualizar_intervalo_nomeado(
        wb, 'LISTA_COMP',
        f"COMPOSICOES!$N$2:$N${config_comp['last_used_row']}"
    )
    _atualizar_intervalo_nomeado(
        wb, 'LISTA_MAT',
        f"MATERIAIS!$F$2:$F${config_cat['mat_last_row']}"
    )
    _atualizar_intervalo_nomeado(
        wb, 'LISTA_MO',
        f"MAO_DE_OBRA!$F$2:$F${config_cat['mo_last_row']}"
    )
    _atualizar_intervalo_nomeado(
        wb, 'LISTA_FER',
        f"FERRAMENTAS!$G$2:$G${config_cat['fer_last_row']}"
    )
    _atualizar_intervalo_nomeado(
        wb, 'LISTA_EQP',
        f"EQUIPAMENTOS!$G$2:$G${config_cat['eqp_last_row']}"
    )

    # Salvar arquivo
    wb.save(output_path)
    print(f'Arquivo salvo em: {output_path}')


if __name__ == '__main__':
    main()
