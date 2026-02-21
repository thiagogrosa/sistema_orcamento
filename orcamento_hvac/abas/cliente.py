"""
Aba CLIENTE - Dados do cliente para o orcamento.

Esta aba contem os campos de identificacao do cliente:
- Cliente (nome/razao social)
- Endereco
- Data e Validade
- Responsavel e Telefone
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _aplicar_estrutura(ws, estilos):
    """
    Aplica toda a estrutura da aba CLIENTE.
    """
    thin_border = estilos['thin_border']
    input_fill = estilos['input_fill']

    # Titulo
    ws.merge_cells('A1:F1')
    cell_titulo = ws.cell(row=1, column=1, value='DADOS DO CLIENTE')
    cell_titulo.font = Font(bold=True, size=16, color='2E86AB')
    cell_titulo.alignment = Alignment(horizontal='center')

    # Campo Cliente (linha 3)
    ws.cell(row=3, column=1, value='Cliente:').font = Font(bold=True)
    ws.merge_cells('B3:F3')
    cell_cliente = ws.cell(row=3, column=2)
    cell_cliente.fill = input_fill
    cell_cliente.border = thin_border
    # Aplicar borda em todas as celulas mescladas
    for col in range(2, 7):
        ws.cell(row=3, column=col).border = thin_border

    # Campo Endereco (linha 4)
    ws.cell(row=4, column=1, value='Endereco:').font = Font(bold=True)
    ws.merge_cells('B4:F4')
    cell_endereco = ws.cell(row=4, column=2)
    cell_endereco.fill = input_fill
    cell_endereco.border = thin_border
    for col in range(2, 7):
        ws.cell(row=4, column=col).border = thin_border

    # Campos Data e Validade (linha 5)
    ws.cell(row=5, column=1, value='Data:').font = Font(bold=True)
    cell_data = ws.cell(row=5, column=2)
    cell_data.fill = input_fill
    cell_data.border = thin_border

    ws.cell(row=5, column=3, value='Validade:').font = Font(bold=True)
    ws.merge_cells('D5:F5')
    cell_validade = ws.cell(row=5, column=4)
    cell_validade.fill = input_fill
    cell_validade.border = thin_border
    for col in range(4, 7):
        ws.cell(row=5, column=col).border = thin_border

    # Campos Responsavel e Telefone (linha 6)
    ws.cell(row=6, column=1, value='Responsavel:').font = Font(bold=True)
    ws.merge_cells('B6:C6')
    cell_resp = ws.cell(row=6, column=2)
    cell_resp.fill = input_fill
    cell_resp.border = thin_border
    ws.cell(row=6, column=3).border = thin_border

    ws.cell(row=6, column=4, value='Telefone:').font = Font(bold=True)
    ws.merge_cells('E6:F6')
    cell_tel = ws.cell(row=6, column=5)
    cell_tel.fill = input_fill
    cell_tel.border = thin_border
    ws.cell(row=6, column=6).border = thin_border

    # Ajustar larguras das colunas
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    return ws


def criar(wb, estilos):
    """
    Cria a aba CLIENTE.

    Args:
        wb: Workbook
        estilos: dict de estilos
    """
    ws = wb.create_sheet('CLIENTE')
    return _aplicar_estrutura(ws, estilos)


def preencher(wb, estilos):
    """
    Preenche a aba CLIENTE existente no template.

    A logica e identica ao criar(), mas usa aba existente.
    Se a aba nao existir, ela sera criada.
    """
    if 'CLIENTE' not in wb.sheetnames:
        ws = wb.create_sheet('CLIENTE')
    else:
        ws = wb['CLIENTE']

        # Desfazer mesclagens existentes antes de limpar
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))

        # Limpar conteudo existente
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.value = None

    return _aplicar_estrutura(ws, estilos)
