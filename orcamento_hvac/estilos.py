"""
Estilos globais para a planilha de custos HVAC Split.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def criar_estilos():
    """Retorna dicionário com todos os estilos da planilha."""

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        # Bordas
        'thin_border': thin_border,

        # Fontes
        'header_font': Font(bold=True, color='FFFFFF', size=11),
        'title_font': Font(bold=True, size=14, color='2E86AB'),
        'subtitle_font': Font(bold=True, size=12),
        'section_font': Font(bold=True, size=11, color='FFFFFF'),
        'code_font': Font(name='Consolas', size=10),

        # Preenchimentos de cabeçalho
        'header_fill': PatternFill('solid', fgColor='2E86AB'),
        'section_fill': PatternFill('solid', fgColor='5D6D7E'),

        # Preenchimentos por tipo
        'mo_fill': PatternFill('solid', fgColor='EAFAF1'),
        'fer_fill': PatternFill('solid', fgColor='FEF9E7'),
        'eqp_fill': PatternFill('solid', fgColor='EBF5FB'),
        'comp_header_fill': PatternFill('solid', fgColor='2E86AB'),
        'total_fill': PatternFill('solid', fgColor='BDC3C7'),
        'input_fill': PatternFill('solid', fgColor='FFFACD'),

        # Cores por categoria de material
        'cat_colors': {
            'Tubulação': 'D6EAF8',
            'Isolamento': 'D5F5E3',
            'Elétrica': 'FCF3CF',
            'Eletroduto': 'FCF3CF',
            'Dreno': 'D1F2EB',
            'Suporte': 'FAE5D3',
            'Acabamento': 'FADBD8',
            'Gás/Solda': 'E5E8E8',
            'Alvenaria': 'F5EEF8',
            'Proteção': 'FCF3CF',
        },

        # Formatos numéricos
        'currency_format': 'R$ #,##0.00',
        'qty_format': '#,##0.00',
        'hours_format': '#,##0',

        # Alinhamentos comuns
        'center': Alignment(horizontal='center', vertical='center'),
        'wrap': Alignment(wrap_text=True, vertical='top'),
    }
