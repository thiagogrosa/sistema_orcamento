"""
Abas de catálogos: MATERIAIS, MAO_DE_OBRA, FERRAMENTAS, EQUIPAMENTOS.
Estrutura atualizada com colunas de validade por item.
"""
from openpyxl.styles import Font, PatternFill, Alignment


def criar(wb, estilos, materiais, mao_de_obra, ferramentas, equipamentos):
    """
    Cria as abas de catálogos.

    Retorna dict com:
    - mat_last_row, mo_last_row, fer_last_row, eqp_last_row
    """
    thin_border = estilos['thin_border']
    header_font = estilos['header_font']
    header_fill = estilos['header_fill']
    code_font = estilos['code_font']
    currency_format = estilos['currency_format']
    qty_format = estilos['qty_format']
    hours_format = estilos['hours_format']
    cat_colors = estilos['cat_colors']
    input_fill = estilos['input_fill']
    date_format = 'YYYY-MM-DD'

    # ========== MATERIAIS ==========
    # Colunas: A-Código, B-Categoria, C-Descrição, D-Unidade, E-Preço,
    #          F-Atualizado Em, G-Validade (dias), H-Seleção
    ws_mat = wb.create_sheet('MATERIAIS')

    mat_headers = ['Código', 'Categoria', 'Descrição', 'Unidade', 'Preço (R$)',
                   'Atualizado Em', 'Validade (dias)', 'Seleção']
    for col, header in enumerate(mat_headers, 1):
        cell = ws_mat.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, mat in enumerate(materiais, 2):
        # mat: (codigo, categoria, descricao, unidade, preco, data_atualizacao, validade_dias)
        for col_idx, value in enumerate(mat, 1):
            cell = ws_mat.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = code_font
                cell.alignment = Alignment(horizontal='left')
            elif col_idx == 2 or col_idx == 4:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 5:
                cell.number_format = currency_format
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal='center')

            cat = mat[1]
            if cat in cat_colors:
                cell.fill = PatternFill('solid', fgColor=cat_colors[cat])

        # Coluna H: Seleção
        cell_sel = ws_mat.cell(row=row_idx, column=8, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    ws_mat.column_dimensions['A'].width = 18
    ws_mat.column_dimensions['B'].width = 12
    ws_mat.column_dimensions['C'].width = 45
    ws_mat.column_dimensions['D'].width = 10
    ws_mat.column_dimensions['E'].width = 14
    ws_mat.column_dimensions['F'].width = 14
    ws_mat.column_dimensions['G'].width = 14
    ws_mat.column_dimensions['H'].width = 55
    ws_mat.column_dimensions['H'].hidden = True
    ws_mat.freeze_panes = 'A2'

    # ========== MÃO DE OBRA ==========
    # Colunas: A-Código, B-Categoria, C-Descrição, D-Unidade, E-Custo,
    #          F-Atualizado Em, G-Validade (dias), H-Seleção
    ws_mo = wb.create_sheet('MAO_DE_OBRA')

    mo_headers = ['Código', 'Categoria', 'Descrição', 'Unidade', 'Custo (R$)',
                  'Atualizado Em', 'Validade (dias)', 'Seleção']
    for col, header in enumerate(mo_headers, 1):
        cell = ws_mo.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, mo in enumerate(mao_de_obra, 2):
        # mo: (codigo, categoria, descricao, unidade, custo, data_atualizacao, validade_dias)
        for col_idx, value in enumerate(mo, 1):
            cell = ws_mo.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = code_font
            elif col_idx == 2 or col_idx == 4:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 5:
                cell.number_format = currency_format
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal='center')

        # Coluna H: Seleção
        cell_sel = ws_mo.cell(row=row_idx, column=8, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    ws_mo.column_dimensions['A'].width = 14
    ws_mo.column_dimensions['B'].width = 14
    ws_mo.column_dimensions['C'].width = 40
    ws_mo.column_dimensions['D'].width = 10
    ws_mo.column_dimensions['E'].width = 14
    ws_mo.column_dimensions['F'].width = 14
    ws_mo.column_dimensions['G'].width = 14
    ws_mo.column_dimensions['H'].width = 50
    ws_mo.column_dimensions['H'].hidden = True
    ws_mo.freeze_panes = 'A2'

    # ========== FERRAMENTAS ==========
    # Colunas: A-Código, B-Categoria, C-Descrição, D-Valor Aquisição, E-Vida Útil,
    #          F-Custo/Hora, G-Atualizado Em, H-Validade (dias), I-Seleção
    ws_fer = wb.create_sheet('FERRAMENTAS')

    fer_headers = ['Código', 'Categoria', 'Descrição', 'Valor Aquisição (R$)',
                   'Vida Útil (H)', 'Custo/Hora (R$)', 'Atualizado Em',
                   'Validade (dias)', 'Seleção']
    for col, header in enumerate(fer_headers, 1):
        cell = ws_fer.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, fer in enumerate(ferramentas, 2):
        # fer: (codigo, categoria, descricao, valor_aquisicao, vida_util, data_atualizacao, validade_dias)
        ws_fer.cell(row=row_idx, column=1, value=fer[0]).font = code_font
        ws_fer.cell(row=row_idx, column=1).border = thin_border

        ws_fer.cell(row=row_idx, column=2, value=fer[1]).border = thin_border
        ws_fer.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

        ws_fer.cell(row=row_idx, column=3, value=fer[2]).border = thin_border

        cell_valor = ws_fer.cell(row=row_idx, column=4, value=fer[3])
        cell_valor.border = thin_border
        cell_valor.number_format = currency_format

        cell_vida = ws_fer.cell(row=row_idx, column=5, value=fer[4])
        cell_vida.border = thin_border
        cell_vida.number_format = hours_format
        cell_vida.alignment = Alignment(horizontal='center')

        # Coluna F: Custo/Hora (fórmula)
        formula_custo = f'=IF(E{row_idx}>0,D{row_idx}/E{row_idx},0)'
        cell_custo_h = ws_fer.cell(row=row_idx, column=6, value=formula_custo)
        cell_custo_h.border = thin_border
        cell_custo_h.number_format = currency_format

        # Coluna G: Atualizado Em
        cell_atualizado = ws_fer.cell(row=row_idx, column=7, value=fer[5])
        cell_atualizado.border = thin_border
        cell_atualizado.alignment = Alignment(horizontal='center')

        # Coluna H: Validade (dias)
        cell_validade = ws_fer.cell(row=row_idx, column=8, value=fer[6])
        cell_validade.border = thin_border
        cell_validade.alignment = Alignment(horizontal='center')

        # Coluna I: Seleção
        cell_sel = ws_fer.cell(row=row_idx, column=9, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    ws_fer.column_dimensions['A'].width = 16
    ws_fer.column_dimensions['B'].width = 12
    ws_fer.column_dimensions['C'].width = 35
    ws_fer.column_dimensions['D'].width = 18
    ws_fer.column_dimensions['E'].width = 14
    ws_fer.column_dimensions['F'].width = 16
    ws_fer.column_dimensions['G'].width = 14
    ws_fer.column_dimensions['H'].width = 14
    ws_fer.column_dimensions['I'].width = 45
    ws_fer.column_dimensions['I'].hidden = True
    ws_fer.freeze_panes = 'A2'

    # ========== EQUIPAMENTOS ==========
    # Colunas: A-Código, B-Categoria, C-Descrição, D-Capacidade, E-Unidade,
    #          F-Preço, G-Atualizado Em, H-Validade (dias), I-Seleção
    ws_eqp = wb.create_sheet('EQUIPAMENTOS')

    eqp_headers = ['Código', 'Categoria', 'Descrição', 'Capacidade (BTU)',
                   'Unidade', 'Preço (R$)', 'Atualizado Em', 'Validade (dias)', 'Seleção']
    for col, header in enumerate(eqp_headers, 1):
        cell = ws_eqp.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, eqp in enumerate(equipamentos, 2):
        # eqp: (codigo, categoria, descricao, capacidade, unidade, preco, data_atualizacao, validade_dias)
        ws_eqp.cell(row=row_idx, column=1, value=eqp[0]).font = code_font
        ws_eqp.cell(row=row_idx, column=1).border = thin_border

        ws_eqp.cell(row=row_idx, column=2, value=eqp[1]).border = thin_border
        ws_eqp.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

        ws_eqp.cell(row=row_idx, column=3, value=eqp[2]).border = thin_border

        cell_cap = ws_eqp.cell(row=row_idx, column=4, value=eqp[3] if eqp[3] > 0 else '-')
        cell_cap.border = thin_border
        cell_cap.alignment = Alignment(horizontal='center')
        if eqp[3] > 0:
            cell_cap.number_format = '#,##0'

        ws_eqp.cell(row=row_idx, column=5, value=eqp[4]).border = thin_border
        ws_eqp.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')

        cell_preco = ws_eqp.cell(row=row_idx, column=6, value=eqp[5])
        cell_preco.border = thin_border
        cell_preco.number_format = currency_format

        # Coluna G: Atualizado Em
        cell_atualizado = ws_eqp.cell(row=row_idx, column=7, value=eqp[6])
        cell_atualizado.border = thin_border
        cell_atualizado.alignment = Alignment(horizontal='center')

        # Coluna H: Validade (dias)
        cell_validade = ws_eqp.cell(row=row_idx, column=8, value=eqp[7])
        cell_validade.border = thin_border
        cell_validade.alignment = Alignment(horizontal='center')

        # Coluna I: Seleção
        cell_sel = ws_eqp.cell(row=row_idx, column=9, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    ws_eqp.column_dimensions['A'].width = 16
    ws_eqp.column_dimensions['B'].width = 14
    ws_eqp.column_dimensions['C'].width = 40
    ws_eqp.column_dimensions['D'].width = 16
    ws_eqp.column_dimensions['E'].width = 10
    ws_eqp.column_dimensions['F'].width = 14
    ws_eqp.column_dimensions['G'].width = 14
    ws_eqp.column_dimensions['H'].width = 14
    ws_eqp.column_dimensions['I'].width = 50
    ws_eqp.column_dimensions['I'].hidden = True
    ws_eqp.freeze_panes = 'A2'

    return {
        'ws_mat': ws_mat,
        'ws_mo': ws_mo,
        'ws_fer': ws_fer,
        'ws_eqp': ws_eqp,
        'mat_last_row': len(materiais) + 1,
        'mo_last_row': len(mao_de_obra) + 1,
        'fer_last_row': len(ferramentas) + 1,
        'eqp_last_row': len(equipamentos) + 1,
    }


def _limpar_dados(ws, start_row):
    """Limpa dados a partir de uma linha, mantendo headers."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def preencher(wb, estilos, materiais, mao_de_obra, ferramentas, equipamentos):
    """
    Preenche as abas de catálogos existentes no template.

    Retorna dict com last_row info.
    """
    thin_border = estilos['thin_border']
    header_font = estilos['header_font']
    header_fill = estilos['header_fill']
    code_font = estilos['code_font']
    currency_format = estilos['currency_format']
    qty_format = estilos['qty_format']
    hours_format = estilos['hours_format']
    cat_colors = estilos['cat_colors']
    input_fill = estilos['input_fill']

    # ========== MATERIAIS ==========
    ws_mat = wb['MATERIAIS']
    _limpar_dados(ws_mat, 2)

    for row_idx, mat in enumerate(materiais, 2):
        # mat: (codigo, categoria, descricao, unidade, preco, data_atualizacao, validade_dias)
        for col_idx, value in enumerate(mat, 1):
            cell = ws_mat.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = code_font
                cell.alignment = Alignment(horizontal='left')
            elif col_idx == 2 or col_idx == 4:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 5:
                cell.number_format = currency_format
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal='center')

            cat = mat[1]
            if cat in cat_colors:
                cell.fill = PatternFill('solid', fgColor=cat_colors[cat])

        # Coluna H: Seleção
        cell_sel = ws_mat.cell(row=row_idx, column=8, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    # ========== MÃO DE OBRA ==========
    ws_mo = wb['MAO_DE_OBRA']
    _limpar_dados(ws_mo, 2)

    for row_idx, mo in enumerate(mao_de_obra, 2):
        # mo: (codigo, categoria, descricao, unidade, custo, data_atualizacao, validade_dias)
        for col_idx, value in enumerate(mo, 1):
            cell = ws_mo.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = code_font
            elif col_idx == 2 or col_idx == 4:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 5:
                cell.number_format = currency_format
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal='center')

        # Coluna H: Seleção
        cell_sel = ws_mo.cell(row=row_idx, column=8, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    # ========== FERRAMENTAS ==========
    ws_fer = wb['FERRAMENTAS']
    _limpar_dados(ws_fer, 2)

    for row_idx, fer in enumerate(ferramentas, 2):
        # fer: (codigo, categoria, descricao, valor_aquisicao, vida_util, data_atualizacao, validade_dias)
        ws_fer.cell(row=row_idx, column=1, value=fer[0]).font = code_font
        ws_fer.cell(row=row_idx, column=1).border = thin_border

        ws_fer.cell(row=row_idx, column=2, value=fer[1]).border = thin_border
        ws_fer.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

        ws_fer.cell(row=row_idx, column=3, value=fer[2]).border = thin_border

        cell_valor = ws_fer.cell(row=row_idx, column=4, value=fer[3])
        cell_valor.border = thin_border
        cell_valor.number_format = currency_format

        cell_vida = ws_fer.cell(row=row_idx, column=5, value=fer[4])
        cell_vida.border = thin_border
        cell_vida.number_format = hours_format
        cell_vida.alignment = Alignment(horizontal='center')

        formula_custo = f'=IF(E{row_idx}>0,D{row_idx}/E{row_idx},0)'
        cell_custo_h = ws_fer.cell(row=row_idx, column=6, value=formula_custo)
        cell_custo_h.border = thin_border
        cell_custo_h.number_format = currency_format

        # Coluna G: Atualizado Em
        cell_atualizado = ws_fer.cell(row=row_idx, column=7, value=fer[5])
        cell_atualizado.border = thin_border
        cell_atualizado.alignment = Alignment(horizontal='center')

        # Coluna H: Validade (dias)
        cell_validade = ws_fer.cell(row=row_idx, column=8, value=fer[6])
        cell_validade.border = thin_border
        cell_validade.alignment = Alignment(horizontal='center')

        # Coluna I: Seleção
        cell_sel = ws_fer.cell(row=row_idx, column=9, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    # ========== EQUIPAMENTOS ==========
    ws_eqp = wb['EQUIPAMENTOS']
    _limpar_dados(ws_eqp, 2)

    for row_idx, eqp in enumerate(equipamentos, 2):
        # eqp: (codigo, categoria, descricao, capacidade, unidade, preco, data_atualizacao, validade_dias)
        ws_eqp.cell(row=row_idx, column=1, value=eqp[0]).font = code_font
        ws_eqp.cell(row=row_idx, column=1).border = thin_border

        ws_eqp.cell(row=row_idx, column=2, value=eqp[1]).border = thin_border
        ws_eqp.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

        ws_eqp.cell(row=row_idx, column=3, value=eqp[2]).border = thin_border

        cell_cap = ws_eqp.cell(row=row_idx, column=4, value=eqp[3] if eqp[3] > 0 else '-')
        cell_cap.border = thin_border
        cell_cap.alignment = Alignment(horizontal='center')
        if eqp[3] > 0:
            cell_cap.number_format = '#,##0'

        ws_eqp.cell(row=row_idx, column=5, value=eqp[4]).border = thin_border
        ws_eqp.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')

        cell_preco = ws_eqp.cell(row=row_idx, column=6, value=eqp[5])
        cell_preco.border = thin_border
        cell_preco.number_format = currency_format

        # Coluna G: Atualizado Em
        cell_atualizado = ws_eqp.cell(row=row_idx, column=7, value=eqp[6])
        cell_atualizado.border = thin_border
        cell_atualizado.alignment = Alignment(horizontal='center')

        # Coluna H: Validade (dias)
        cell_validade = ws_eqp.cell(row=row_idx, column=8, value=eqp[7])
        cell_validade.border = thin_border
        cell_validade.alignment = Alignment(horizontal='center')

        # Coluna I: Seleção
        cell_sel = ws_eqp.cell(row=row_idx, column=9, value=f'=A{row_idx}&" - "&C{row_idx}')
        cell_sel.border = thin_border

    return {
        'ws_mat': ws_mat,
        'ws_mo': ws_mo,
        'ws_fer': ws_fer,
        'ws_eqp': ws_eqp,
        'mat_last_row': len(materiais) + 1,
        'mo_last_row': len(mao_de_obra) + 1,
        'fer_last_row': len(ferramentas) + 1,
        'eqp_last_row': len(equipamentos) + 1,
    }
