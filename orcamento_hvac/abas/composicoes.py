"""
Aba COMPOSICOES - Servicos compostos de materiais, mao de obra e ferramentas.

Estrutura de colunas (A-R):
- A: Codigo da composicao
- B: Descricao
- C: Tipo (MAT/MO/FER/EQP)
- D: Cod. Item (dropdown com CODIGO - DESCRICAO)
- E: Unidade (SPILL - template)
- F: Qtd Base
- G: Qtd Var (variavel por metro)
- H: Preco Unit. (SPILL - template)
- I: Sub. Base (SPILL - template)
- J: Sub. Var (SPILL - template)
- K: Mult. (SPILL - template)
- L: Base c/ Margem (SPILL - template)
- M: Var c/ Margem (SPILL - template)
- N: Selecao (oculta, para dropdown)
- O: Desc. Pre (descricao dinamica)
- P: Desc. Pos (descricao dinamica)
- Q: Unid. Sing (oculta, unidade singular para desc dinamica)
- R: Unid. Plur (oculta, unidade plural para desc dinamica)

IMPORTANTE: Colunas E, H, I, J, K, L, M sao preenchidas por
formulas de spill no template.xlsm. O Python so preenche DADOS nas
colunas A, B, C, D, F, G, N, O, P, Q, R.
"""
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


def _aplicar_cor_tipo(ws, row, tipo, estilos, num_cols=18):
    """Aplica cor de fundo baseada no tipo do item (colunas A-R)."""
    fill = None
    if tipo == 'MO':
        fill = estilos['mo_fill']
    elif tipo == 'FER':
        fill = estilos['fer_fill']
    elif tipo == 'EQP':
        fill = estilos['eqp_fill']

    if fill:
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = fill


def _criar_lookup_catalogos(materiais, mao_de_obra, ferramentas, equipamentos):
    """
    Cria dicionário de lookup: código -> "CÓDIGO - DESCRIÇÃO"

    Args:
        materiais: Lista de tuplas (codigo, categoria, descricao, unidade, preco)
        mao_de_obra: Lista de tuplas (codigo, categoria, descricao, unidade, preco)
        ferramentas: Lista de tuplas (codigo, categoria, descricao, unidade, custo, preco_hora)
        equipamentos: Lista de tuplas (codigo, categoria, descricao, unidade, preco_aquisicao, preco_hora)

    Returns:
        Dict com código -> "CÓDIGO - DESCRIÇÃO"
    """
    lookup = {}

    # Materiais: (codigo, categoria, descricao, unidade, preco)
    for item in materiais:
        codigo = item[0]
        descricao = item[2]
        lookup[codigo] = f"{codigo} - {descricao}"

    # Mão de obra: (codigo, categoria, descricao, unidade, preco)
    for item in mao_de_obra:
        codigo = item[0]
        descricao = item[2]
        lookup[codigo] = f"{codigo} - {descricao}"

    # Ferramentas: (codigo, categoria, descricao, unidade, custo, preco_hora)
    for item in ferramentas:
        codigo = item[0]
        descricao = item[2]
        lookup[codigo] = f"{codigo} - {descricao}"

    # Equipamentos: (codigo, categoria, descricao, unidade, preco_aquisicao, preco_hora)
    for item in equipamentos:
        codigo = item[0]
        descricao = item[2]
        lookup[codigo] = f"{codigo} - {descricao}"

    return lookup


def criar(wb, estilos, dados_composicoes, config_negocio=None,
          materiais=None, mao_de_obra=None, ferramentas=None, equipamentos=None):
    """
    Cria a aba COMPOSICOES.

    Args:
        wb: Workbook
        estilos: Dict com estilos
        dados_composicoes: Lista de composicoes
        config_negocio: Dict retornado por negocio.criar() com mult_rows
        materiais: Lista de materiais para lookup
        mao_de_obra: Lista de mão de obra para lookup
        ferramentas: Lista de ferramentas para lookup
        equipamentos: Lista de equipamentos para lookup

    Retorna dict com:
    - ws: worksheet
    - comp_totals_info: lista com info de cada composicao
    - last_used_row: ultima linha usada
    """
    if config_negocio is None:
        config_negocio = {'mult_rows': {'MAT': 42, 'MO': 43, 'FER': 44, 'EQP': 45}}

    # Criar lookup para código -> "CÓDIGO - DESCRIÇÃO"
    lookup = {}
    if materiais and mao_de_obra and ferramentas and equipamentos:
        lookup = _criar_lookup_catalogos(materiais, mao_de_obra, ferramentas, equipamentos)

    ws = wb.create_sheet('COMPOSICOES')
    thin_border = estilos['thin_border']
    header_font = estilos['header_font']
    header_fill = estilos['header_fill']
    code_font = estilos['code_font']
    currency_format = estilos['currency_format']
    qty_format = estilos['qty_format']
    input_fill = estilos['input_fill']
    comp_header_fill = estilos['comp_header_fill']

    # Headers (A-R)
    comp_headers = [
        'Codigo', 'Descricao', 'Tipo', 'Cod. Item', 'Un',
        'Qtd Base', 'Qtd Var', 'Preco Unit.', 'Sub. Base', 'Sub. Var',
        'Mult.', 'Base c/ Margem', 'Var c/ Margem',
        'Selecao', 'Desc. Pre', 'Desc. Pos',
        'Unid. Sing', 'Unid. Plur'  # Q, R - unidades para descricao dinamica
    ]
    for col, header in enumerate(comp_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    current_row = 2
    comp_totals_info = []

    for comp in dados_composicoes:
        header_row_comp = current_row

        # Linha de cabecalho da composicao
        cell_codigo = ws.cell(row=current_row, column=1, value=comp['codigo'])
        cell_codigo.font = Font(bold=True, color='FFFFFF', name='Consolas')
        cell_codigo.fill = comp_header_fill
        cell_codigo.alignment = Alignment(horizontal='left', vertical='center')
        cell_codigo.border = thin_border

        cell_desc = ws.cell(row=current_row, column=2, value=comp['descricao'])
        cell_desc.font = Font(bold=True, color='FFFFFF')
        cell_desc.fill = comp_header_fill
        cell_desc.alignment = Alignment(horizontal='left', vertical='center')
        cell_desc.border = thin_border

        # Colunas C ate M vazias no cabecalho (spill preenchera)
        for col in range(3, 14):
            cell = ws.cell(row=current_row, column=col, value='')
            cell.fill = comp_header_fill
            cell.border = thin_border

        # Coluna N: Selecao (formula)
        cell_sel = ws.cell(row=current_row, column=14, value=f'=A{current_row}&" - "&B{current_row}')
        cell_sel.fill = comp_header_fill
        cell_sel.border = thin_border

        # Coluna O: Desc. Pre
        cell_desc_pre = ws.cell(row=current_row, column=15, value=comp.get('desc_pre', ''))
        cell_desc_pre.fill = comp_header_fill
        cell_desc_pre.border = thin_border
        cell_desc_pre.font = Font(color='FFFFFF')

        # Coluna P: Desc. Pos
        cell_desc_pos = ws.cell(row=current_row, column=16, value=comp.get('desc_pos', ''))
        cell_desc_pos.fill = comp_header_fill
        cell_desc_pos.border = thin_border
        cell_desc_pos.font = Font(color='FFFFFF')

        # Coluna Q: Unid. Sing
        cell_unid_sing = ws.cell(row=current_row, column=17, value=comp.get('unid_sing', ''))
        cell_unid_sing.fill = comp_header_fill
        cell_unid_sing.border = thin_border
        cell_unid_sing.font = Font(color='FFFFFF')

        # Coluna R: Unid. Plur
        cell_unid_plur = ws.cell(row=current_row, column=18, value=comp.get('unid_plur', ''))
        cell_unid_plur.fill = comp_header_fill
        cell_unid_plur.border = thin_border
        cell_unid_plur.font = Font(color='FFFFFF')

        current_row += 1
        first_item_row = current_row

        # Itens da composicao
        for item in comp['itens']:
            tipo, codigo, qtd_base, qtd_metro = item

            # Colunas A, B: Vazias para itens
            ws.cell(row=current_row, column=1, value='').border = thin_border
            ws.cell(row=current_row, column=2, value='').border = thin_border

            # Coluna C: Tipo (DADO)
            cell_tipo = ws.cell(row=current_row, column=3, value=tipo)
            cell_tipo.alignment = Alignment(horizontal='center')
            cell_tipo.border = thin_border
            cell_tipo.fill = input_fill

            # Coluna D: Codigo Item com descrição (DADO)
            # Usa lookup para obter "CÓDIGO - DESCRIÇÃO", ou só o código se não encontrar
            valor_col_d = lookup.get(codigo, codigo)
            cell_cod = ws.cell(row=current_row, column=4, value=valor_col_d)
            cell_cod.font = code_font
            cell_cod.alignment = Alignment(horizontal='left')
            cell_cod.border = thin_border
            cell_cod.fill = input_fill

            # Coluna E: Vazia (SPILL preenchera Un)
            cell_un = ws.cell(row=current_row, column=5, value='')
            cell_un.border = thin_border
            cell_un.alignment = Alignment(horizontal='center')

            # Coluna F: Qtd Base (DADO)
            cell_qtd_base = ws.cell(row=current_row, column=6, value=qtd_base if qtd_base > 0 else '')
            cell_qtd_base.number_format = qty_format
            cell_qtd_base.alignment = Alignment(horizontal='center')
            cell_qtd_base.border = thin_border
            cell_qtd_base.fill = input_fill

            # Coluna G: Qtd Var (DADO)
            cell_qtd_var = ws.cell(row=current_row, column=7, value=qtd_metro if qtd_metro > 0 else '')
            cell_qtd_var.number_format = qty_format
            cell_qtd_var.alignment = Alignment(horizontal='center')
            cell_qtd_var.border = thin_border
            cell_qtd_var.fill = input_fill

            # Colunas H-M: Vazias (SPILL preenchera)
            for col in range(8, 14):
                cell = ws.cell(row=current_row, column=col, value='')
                cell.border = thin_border
                if col in [8, 9, 10, 12, 13]:  # Preco, Sub.Base, Sub.Var, Base c/Margem, Var c/Margem
                    cell.number_format = currency_format
                elif col == 11:  # Mult.
                    cell.number_format = '0.0000'
                    cell.alignment = Alignment(horizontal='center')

            # Colunas N, O, P: Vazias nas linhas de item
            for col in range(14, 17):
                ws.cell(row=current_row, column=col, value='').border = thin_border

            # Colunas Q, R: Vazias nas linhas de item (unidades)
            for col in range(17, 19):
                ws.cell(row=current_row, column=col, value='').border = thin_border

            # Cores por tipo
            _aplicar_cor_tipo(ws, current_row, tipo, estilos, num_cols=18)

            current_row += 1

        last_item_row = current_row - 1

        # Nota: Os totais nas colunas I, J, L, M eram SUM manuais.
        # Agora os totais sao calculados automaticamente pelas formulas
        # de spill nas colunas L e M (SUMIFS baseado na coluna A).
        # O ESCOPO busca totais nas colunas L e M.

        comp_totals_info.append({
            'codigo': comp['codigo'],
            'header_row': header_row_comp,
        })

        current_row += 1

    # Ajustar larguras (colunas A-R)
    ws.column_dimensions['A'].width = 18   # Codigo
    ws.column_dimensions['B'].width = 55   # Descricao
    ws.column_dimensions['C'].width = 8    # Tipo
    ws.column_dimensions['D'].width = 55   # Cod. Item (CÓDIGO - DESCRIÇÃO)
    ws.column_dimensions['E'].width = 8    # Un (SPILL)
    ws.column_dimensions['F'].width = 12   # Qtd Base
    ws.column_dimensions['G'].width = 12   # Qtd Var
    ws.column_dimensions['H'].width = 14   # Preco Unit. (SPILL)
    ws.column_dimensions['I'].width = 12   # Sub. Base (SPILL)
    ws.column_dimensions['J'].width = 12   # Sub. Var (SPILL)
    ws.column_dimensions['K'].width = 10   # Mult. (SPILL)
    ws.column_dimensions['L'].width = 14   # Base c/ Margem (SPILL)
    ws.column_dimensions['M'].width = 14   # Var c/ Margem (SPILL)
    ws.column_dimensions['N'].width = 65   # Selecao
    ws.column_dimensions['N'].hidden = True  # Selecao oculta
    ws.column_dimensions['O'].width = 45   # Desc. Pre
    ws.column_dimensions['P'].width = 35   # Desc. Pos
    ws.column_dimensions['Q'].width = 12   # Unid. Sing
    ws.column_dimensions['Q'].hidden = True  # Unid.Sing oculta (para lookup)
    ws.column_dimensions['R'].width = 12   # Unid. Plur
    ws.column_dimensions['R'].hidden = True  # Unid.Plur oculta (para lookup)

    ws.freeze_panes = 'A2'

    # Validacao para tipo
    dv_tipo = DataValidation(type="list", formula1='"MAT,MO,FER,EQP"', allow_blank=True)
    dv_tipo.error = 'Selecione um tipo valido'
    dv_tipo.errorTitle = 'Tipo invalido'
    ws.add_data_validation(dv_tipo)
    dv_tipo.add('C2:C500')

    # Validacao dinamica para Cod. Item
    for row_idx in range(2, 501):
        dv_item = DataValidation(
            type="list",
            formula1=f'=INDIRECT("LISTA_"&$C${row_idx})',
            allow_blank=True
        )
        dv_item.error = 'Primeiro selecione o TIPO, depois escolha o item'
        dv_item.errorTitle = 'Item invalido'
        dv_item.showErrorMessage = True
        dv_item.showDropDown = False
        ws.add_data_validation(dv_item)
        dv_item.add(f'D{row_idx}')

    last_used_row = current_row - 1

    return {
        'ws': ws,
        'comp_totals_info': comp_totals_info,
        'last_used_row': last_used_row,
    }


def preencher(wb, estilos, dados_composicoes, config_negocio=None,
              materiais=None, mao_de_obra=None, ferramentas=None, equipamentos=None):
    """
    Preenche a aba COMPOSICOES existente no template.

    IMPORTANTE: O template deve ter formulas de spill nas colunas E, H, I, J, K, L, M.
    Esta funcao so preenche DADOS nas colunas A, B, C, D, F, G, N, O, P, Q, R.

    Retorna dict com last_used_row.
    """
    if config_negocio is None:
        config_negocio = {'mult_rows': {'MAT': 42, 'MO': 43, 'FER': 44, 'EQP': 45}}

    # Criar lookup para código -> "CÓDIGO - DESCRIÇÃO"
    lookup = {}
    if materiais and mao_de_obra and ferramentas and equipamentos:
        lookup = _criar_lookup_catalogos(materiais, mao_de_obra, ferramentas, equipamentos)

    ws = wb['COMPOSICOES']
    thin_border = estilos['thin_border']
    code_font = estilos['code_font']
    currency_format = estilos['currency_format']
    qty_format = estilos['qty_format']
    input_fill = estilos['input_fill']
    comp_header_fill = estilos['comp_header_fill']

    # Limpar dados existentes (manter header na linha 1)
    # NOTA: Nao limpar colunas com formulas de spill (E, H-M)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            # Limpar apenas colunas de dados (A-D, F-G, N-R)
            if cell.column in [1, 2, 3, 4, 6, 7, 14, 15, 16, 17, 18]:
                cell.value = None

    current_row = 2
    comp_totals_info = []

    for comp in dados_composicoes:
        header_row_comp = current_row

        # Linha de cabecalho da composicao
        cell_codigo = ws.cell(row=current_row, column=1, value=comp['codigo'])
        cell_codigo.font = Font(bold=True, color='FFFFFF', name='Consolas')
        cell_codigo.fill = comp_header_fill
        cell_codigo.alignment = Alignment(horizontal='left', vertical='center')
        cell_codigo.border = thin_border

        cell_desc = ws.cell(row=current_row, column=2, value=comp['descricao'])
        cell_desc.font = Font(bold=True, color='FFFFFF')
        cell_desc.fill = comp_header_fill
        cell_desc.alignment = Alignment(horizontal='left', vertical='center')
        cell_desc.border = thin_border

        # Colunas C ate M: Formatacao do cabecalho (spill preenchera valores)
        for col in range(3, 14):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = comp_header_fill
            cell.border = thin_border

        # Coluna N: Selecao (formula)
        cell_sel = ws.cell(row=current_row, column=14, value=f'=A{current_row}&" - "&B{current_row}')
        cell_sel.fill = comp_header_fill
        cell_sel.border = thin_border

        # Coluna O: Desc. Pre
        cell_desc_pre = ws.cell(row=current_row, column=15, value=comp.get('desc_pre', ''))
        cell_desc_pre.fill = comp_header_fill
        cell_desc_pre.border = thin_border
        cell_desc_pre.font = Font(color='FFFFFF')

        # Coluna P: Desc. Pos
        cell_desc_pos = ws.cell(row=current_row, column=16, value=comp.get('desc_pos', ''))
        cell_desc_pos.fill = comp_header_fill
        cell_desc_pos.border = thin_border
        cell_desc_pos.font = Font(color='FFFFFF')

        # Coluna Q: Unid. Sing
        cell_unid_sing = ws.cell(row=current_row, column=17, value=comp.get('unid_sing', ''))
        cell_unid_sing.fill = comp_header_fill
        cell_unid_sing.border = thin_border
        cell_unid_sing.font = Font(color='FFFFFF')

        # Coluna R: Unid. Plur
        cell_unid_plur = ws.cell(row=current_row, column=18, value=comp.get('unid_plur', ''))
        cell_unid_plur.fill = comp_header_fill
        cell_unid_plur.border = thin_border
        cell_unid_plur.font = Font(color='FFFFFF')

        current_row += 1
        first_item_row = current_row

        for item in comp['itens']:
            tipo, codigo, qtd_base, qtd_metro = item

            # Colunas A, B: Vazias para itens
            ws.cell(row=current_row, column=1, value='').border = thin_border
            ws.cell(row=current_row, column=2, value='').border = thin_border

            # Coluna C: Tipo (DADO)
            cell_tipo = ws.cell(row=current_row, column=3, value=tipo)
            cell_tipo.alignment = Alignment(horizontal='center')
            cell_tipo.border = thin_border
            cell_tipo.fill = input_fill

            # Coluna D: Codigo Item com descrição (DADO)
            # Usa lookup para obter "CÓDIGO - DESCRIÇÃO", ou só o código se não encontrar
            valor_col_d = lookup.get(codigo, codigo)
            cell_cod = ws.cell(row=current_row, column=4, value=valor_col_d)
            cell_cod.font = code_font
            cell_cod.alignment = Alignment(horizontal='left')
            cell_cod.border = thin_border
            cell_cod.fill = input_fill

            # Coluna E: Formatacao (SPILL preenchera Un)
            cell_un = ws.cell(row=current_row, column=5)
            cell_un.border = thin_border
            cell_un.alignment = Alignment(horizontal='center')

            # Coluna F: Qtd Base (DADO)
            cell_qtd_base = ws.cell(row=current_row, column=6, value=qtd_base if qtd_base > 0 else '')
            cell_qtd_base.number_format = qty_format
            cell_qtd_base.alignment = Alignment(horizontal='center')
            cell_qtd_base.border = thin_border
            cell_qtd_base.fill = input_fill

            # Coluna G: Qtd Var (DADO)
            cell_qtd_var = ws.cell(row=current_row, column=7, value=qtd_metro if qtd_metro > 0 else '')
            cell_qtd_var.number_format = qty_format
            cell_qtd_var.alignment = Alignment(horizontal='center')
            cell_qtd_var.border = thin_border
            cell_qtd_var.fill = input_fill

            # Colunas H-M: Formatacao (SPILL preenchera)
            for col in range(8, 14):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if col in [8, 9, 10, 12, 13]:  # Preco, Sub.Base, Sub.Var, Base c/Margem, Var c/Margem
                    cell.number_format = currency_format
                elif col == 11:  # Mult.
                    cell.number_format = '0.0000'
                    cell.alignment = Alignment(horizontal='center')

            # Colunas N, O, P: Vazias nas linhas de item
            for col in range(14, 17):
                ws.cell(row=current_row, column=col, value='').border = thin_border

            # Colunas Q, R: Vazias nas linhas de item (unidades)
            for col in range(17, 19):
                ws.cell(row=current_row, column=col, value='').border = thin_border

            # Cores por tipo
            _aplicar_cor_tipo(ws, current_row, tipo, estilos, num_cols=18)

            current_row += 1

        last_item_row = current_row - 1

        # Nota: Os totais sao calculados pelas formulas de spill em L e M
        comp_totals_info.append({
            'codigo': comp['codigo'],
            'header_row': header_row_comp,
        })

        current_row += 1

    last_used_row = current_row - 1

    return {
        'ws': ws,
        'comp_totals_info': comp_totals_info,
        'last_used_row': last_used_row,
    }
